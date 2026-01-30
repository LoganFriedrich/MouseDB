"""
Script 0: 0_Make_or_Fix_Sheets.py

Creates new cohort Excel files OR fixes existing files to have all required sheets.

Mode 1 (New Cohort): Creates a complete template with:
- All sheets needed for the experiment
- Dates pre-calculated from food deprivation start date
- SubjectIDs pre-filled
- Tray types and test phases pre-filled
- Only manually-entered data left blank (pellet scores, weights, surgery details)

Usage:
    python 0_Make_or_Fix_Sheets.py --new --cohort CNT_05 --start-date 2025-02-01 --mice 16
    python 0_Make_or_Fix_Sheets.py --new --cohort CNT_05 --start-date 2025-02-01  # defaults to 16 mice

Future Mode 2 (Fix Existing): Will standardize existing files
    python 0_Make_or_Fix_Sheets.py --fix path/to/existing_file.xlsx
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from pathlib import Path
import argparse
import sys
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def autoscale_columns(ws):
    """
    Autoscale column widths based on header text length
    """
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        header_value = header_cell.value
        
        if header_value:
            # Calculate width based on header text length
            # Add a little padding (1.2 multiplier) for readability
            width = len(str(header_value)) * 1.2
            # Minimum width of 8, maximum of 50
            width = max(8, min(50, width))
        else:
            width = 8
        
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

# =============================================================================
# SCRIPT SETUP - Set working directory to script location
# =============================================================================

# Get the directory where this script is located
# When used as a package, defaults to current working directory
SCRIPT_DIR = Path.cwd()

# Output subdirectory for generated files
OUTPUT_SUBDIR = SCRIPT_DIR / "generated"

# =============================================================================
# CONFIGURATION - Experimental Timeline
# =============================================================================

# Based on EXPERIMENTAL_TIMELINE.md
# Using naming convention from existing files: Training_Flat_1, Training_Pillar_2, etc.
TIMELINE = [
    # (day_offset, phase, tray_type, trays_per_day, notes)
    # Days 0-3: Ramp (no testing)
    # Days 4-6: Flat Training
    (4, "Training_Flat_1", "F", 4, "Learning task mechanics"),
    (5, "Training_Flat_2", "F", 4, "Learning task mechanics"),
    (6, "Training_Flat_3", "F", 4, "Learning task mechanics"),
    # Days 7-13: Pillar Training
    (7, "Training_Pillar_1", "P", 4, "Full difficulty training"),
    (8, "Training_Pillar_2", "P", 4, "Full difficulty training"),
    (9, "Training_Pillar_3", "P", 4, "Full difficulty training"),
    (10, "Training_Pillar_4", "P", 4, "Full difficulty training"),
    (11, "Training_Pillar_5", "P", 4, "Full difficulty training"),
    (12, "Training_Pillar_6", "P", 4, "Full difficulty training"),
    (13, "Training_Pillar_7", "P", 4, "Full difficulty training"),
    # Days 14-16: Last 3 (Pre-Injury Test)
    (14, "Pre-Injury_Test_Pillar_1", "P", 4, "Final baseline day 1"),
    (15, "Pre-Injury_Test_Pillar_2", "P", 4, "Final baseline day 2"),
    (16, "Pre-Injury_Test_Pillar_3", "P", 4, "Final baseline day 3"),
    # Day 17: Injury (no testing)
    # Days 18-24: Recovery (no testing)
    # Post-injury testing (Fridays)
    (25, "Post-Injury_Test_1", "P", 2, "DPI 9"),
    (32, "Post-Injury_Test_2", "P", 2, "DPI 16"),
    (39, "Post-Injury_Test_3", "P", 2, "DPI 23"),
    (46, "Post-Injury_Test_4", "P", 2, "DPI 30"),
    # Days 47-55: Rehab Easy (9 days)
    (47, "Rehab_Easy_1", "E", 4, "Rehab day 1"),
    (48, "Rehab_Easy_2", "E", 4, "Rehab day 2"),
    (49, "Rehab_Easy_3", "E", 4, "Rehab day 3"),
    (50, "Rehab_Easy_4", "E", 4, "Rehab day 4"),
    (51, "Rehab_Easy_5", "E", 4, "Rehab day 5"),
    (52, "Rehab_Easy_6", "E", 4, "Rehab day 6"),
    (53, "Rehab_Easy_7", "E", 4, "Rehab day 7"),
    (54, "Rehab_Easy_8", "E", 4, "Rehab day 8"),
    (55, "Rehab_Easy_9", "E", 4, "Rehab day 9"),
    # Days 56-62: Rehab Flat (7 days)
    (56, "Rehab_Flat_1", "F", 4, "Rehab day 1"),
    (57, "Rehab_Flat_2", "F", 4, "Rehab day 2"),
    (58, "Rehab_Flat_3", "F", 4, "Rehab day 3"),
    (59, "Rehab_Flat_4", "F", 4, "Rehab day 4"),
    (60, "Rehab_Flat_5", "F", 4, "Rehab day 5"),
    (61, "Rehab_Flat_6", "F", 4, "Rehab day 6"),
    (62, "Rehab_Flat_7", "F", 4, "Rehab day 7"),
    # Days 63-69: Rehab Pillar (7 days)
    (63, "Rehab_Pillar_1", "P", 4, "Rehab day 1"),
    (64, "Rehab_Pillar_2", "P", 4, "Rehab day 2"),
    (65, "Rehab_Pillar_3", "P", 4, "Rehab day 3"),
    (66, "Rehab_Pillar_4", "P", 4, "Rehab day 4"),
    (67, "Rehab_Pillar_5", "P", 4, "Rehab day 5"),
    (68, "Rehab_Pillar_6", "P", 4, "Rehab day 6"),
    (69, "Rehab_Pillar_7", "P", 4, "Rehab day 7"),
]

# Injury day offset from start
INJURY_DAY = 17
TRACING_DAY = 70
PERFUSION_DAY = 84

# Hard-coded values for this project
PROJECT_DEFAULTS = {
    "SpeciesTyp": "Mouse",
    "SpeciesStrainTyp": "C57BL/6J",
    "AnimalSourceNam": "Jackson Laboratory",
    "Laboratory": "Murray/Blackmore Lab",
    "StudyLeader": "Logan Friedrich",
    "Injury_device": "Infinite Horizon Impactor",
}


# =============================================================================
# SHEET GENERATORS
# =============================================================================

def generate_subject_ids(cohort_name, num_mice):
    """
    Generate SubjectIDs in format CNT_XX_YY
    
    Args:
        cohort_name: e.g., "CNT_05" 
        num_mice: number of mice (default 16)
    
    Returns:
        List of SubjectIDs like ['CNT_05_01', 'CNT_05_02', ...]
    """
    # Extract cohort number from name (e.g., "CNT_05" -> "05")
    # Handle both "CNT_05" and "05" formats
    if "_" in cohort_name:
        parts = cohort_name.split("_")
        cohort_num = parts[-1]  # Get last part
        prefix = "_".join(parts[:-1])  # Get everything before last part
    else:
        cohort_num = cohort_name
        prefix = "CNT"
    
    subject_ids = []
    for i in range(1, num_mice + 1):
        subject_ids.append(f"{prefix}_{cohort_num}_{i:02d}")
    
    return subject_ids


def create_0a_metadata(subject_ids):
    """
    Create 0a_Metadata sheet with animal-level information
    
    Columns: SubjectID, Date_of_Birth, Date_of_Death, Sex, Cohort, Notes
    """
    df = pd.DataFrame({
        "SubjectID": subject_ids,
        "Date_of_Birth": [None] * len(subject_ids),
        "Date_of_Death": [None] * len(subject_ids),
        "Sex": [None] * len(subject_ids),
        "Cohort": [subject_ids[0].rsplit("_", 1)[0]] * len(subject_ids),  # e.g., "CNT_05"
        "Notes": [None] * len(subject_ids),
    })
    return df


def create_0_injection_calculations():
    """
    Create 0_Injection_Calculations sheet for virus prep
    Matches existing structure with calculation formulas
    """
    # This will be written with formulas via openpyxl
    df = pd.DataFrame({
        "Date of surgery": [None],
        "Virus Name": [None],
        "Virus Box": [None],
        "Virus Source": [None],
        "Starting concentration in SciNot": [None],
        "Starting concentration (in AAV protect)": [None],  # Formula: =E2/L2
        "Target Concentration": [50],  # Default target
        "Total Volume": [None],
        "parts for this virus": [1],  # Default
        "parts of 1xAAV protect": [None],  # Formula: =H2-I2
        "Final concentration": [None],  # Formula: =(F2/H2)*I2
        "Conversion factor": [100000000000],  # 10^11
        "Final Concentration in SciNot": [None],  # Formula: =K2*L2
    })
    return df


def write_0_injection_calculations_with_formulas(ws, df):
    """
    Write 0_Injection_Calculations with calculation formulas
    """
    headers = list(df.columns)
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write row 2 with formulas
    row_idx = 2
    for col_idx, col_name in enumerate(headers, 1):
        if col_name == "Starting concentration (in AAV protect)":
            # Formula: =E2/L2 (Starting SciNot / Conversion factor)
            ws.cell(row=row_idx, column=col_idx, value="=E2/L2")
        elif col_name == "parts of 1xAAV protect":
            # Formula: =H2-I2 (Total Volume - parts for virus)
            ws.cell(row=row_idx, column=col_idx, value="=H2-I2")
        elif col_name == "Final concentration":
            # Formula: =(F2/H2)*I2
            ws.cell(row=row_idx, column=col_idx, value="=(F2/H2)*I2")
        elif col_name == "Final Concentration in SciNot":
            # Formula: =K2*L2
            ws.cell(row=row_idx, column=col_idx, value="=K2*L2")
        else:
            # Static value from dataframe
            val = df.iloc[0][col_name]
            if pd.notna(val):
                ws.cell(row=row_idx, column=col_idx, value=val)


def create_0_virus_preparation(cohort_name, num_groups=2, viruses_per_group=3):
    """
    Create 0_Virus_Preparation sheet with:
    1. Standard notes about virus preparation
    2. GROUP ASSIGNMENTS table (which animals get which injection group)
    3. VIRUS MIX sections for each group (mixing calculations)
    
    This sheet is the single source of truth for virus info.
    Script 1 reads this to populate 5_SC_Injection_Details.
    
    Args:
        cohort_name: e.g., "CNT_05" or "ENCR_01" - used for group naming
        num_groups: Number of injection group templates to create (default 2)
        viruses_per_group: Number of virus rows per group (default 3)
    
    Returns:
        dict with structure info for write_0_virus_preparation
    """
    # Extract prefix for group naming (CNT_05 -> CNT_05.01, ENCR_01 -> ENCR.01)
    if "_" in cohort_name:
        parts = cohort_name.split("_")
        # For CNT_05, we want CNT_05.01 style
        # For ENCR_01, we want ENCR.01 style
        if parts[0].upper() == "ENCR":
            group_prefix = parts[0].upper()  # Just "ENCR"
        else:
            group_prefix = cohort_name  # Full "CNT_05"
    else:
        group_prefix = cohort_name
    
    return {
        '_is_virus_prep': True,
        'cohort_name': cohort_name,
        'group_prefix': group_prefix,
        'num_groups': num_groups,
        'viruses_per_group': viruses_per_group
    }


def write_0_virus_preparation(ws, data):
    """
    Write 0_Virus_Preparation sheet with parseable structure.
    
    Structure:
    - Rows 2-10: Standard notes about virus preparation
    - Row 12: "=== GROUP ASSIGNMENTS ===" marker
    - Row 13: Headers (Group_ID, Animals, Surgery_Date, Target, Notes)
    - Row 14+: One row per group for animal assignments
    - Row X: "=== VIRUS MIXES ===" marker
    - Then for each group: group header + virus table
    
    Script 1 parses this by looking for the marker rows.
    """
    group_prefix = data['group_prefix']
    num_groups = data['num_groups']
    viruses_per_group = data['viruses_per_group']
    
    # Standard notes (rows 2-10)
    notes = [
        (2, 'A', '1. Animals must be set up in a standard way so that reporters and DLK and protection are all consistent'),
        (3, 'A', '2. There will be at least 4 viruses in every injection:'),
        (4, 'B', 'nuclear reporter'),
        (4, 'C', 'H2B-mScarlet'),
        (5, 'B', 'CAG-DLK'),
        (6, 'B', 'Protection'),
        (7, 'B', 'cytoplasmic reporter OR barcode (in eventual single nuclei)'),
        (8, 'A', '3. To simplify virus communication and calculations, I will translate 1x10^12 into 10. So 2.5x10^12 is now expressed as 25 and 1x10^13 is expressed as 100'),
        (9, 'A', '4. The control for the "protective factor" will always be Malat-BC8, brought to a titer of ~20 (2x10^12) in AAV protect'),
        (10, 'A', '5. The control for DLK will be Malat-BC-7, brought to a titer of ~360 (3.6x10^13) in AAV Protect'),
    ]
    
    # Write standard notes
    for row, col, text in notes:
        col_idx = ord(col) - ord('A') + 1
        ws.cell(row=row, column=col_idx, value=text)
    
    # === GROUP ASSIGNMENTS SECTION ===
    current_row = 12
    ws.cell(row=current_row, column=1, value='=== GROUP ASSIGNMENTS ===')
    current_row += 1
    
    # Headers for assignment table
    assignment_headers = ['Group_ID', 'Animals', 'Surgery_Date', 'Injection_Target', 'Notes']
    for col_idx, header in enumerate(assignment_headers, 1):
        ws.cell(row=current_row, column=col_idx, value=header)
    current_row += 1
    
    # Group assignment rows (user fills in Animals column)
    for group_num in range(1, num_groups + 1):
        group_name = f"{group_prefix}.{group_num:02d}"
        ws.cell(row=current_row, column=1, value=group_name)
        ws.cell(row=current_row, column=2, value='')  # Animals - user fills in e.g., "01-03" or "04,06-08"
        ws.cell(row=current_row, column=3, value='')  # Surgery_Date - user fills in
        ws.cell(row=current_row, column=4, value='Caudal C6')  # Default target
        ws.cell(row=current_row, column=5, value='')  # Notes
        current_row += 1
    
    # Blank row
    current_row += 2
    
    # === VIRUS MIXES SECTION ===
    ws.cell(row=current_row, column=1, value='=== VIRUS MIXES ===')
    current_row += 2
    
    # Create virus mix tables for each group
    for group_num in range(1, num_groups + 1):
        group_name = f"{group_prefix}.{group_num:02d}"
        
        # Group header with marker format for parsing
        ws.cell(row=current_row, column=1, value=f'[{group_name}]')
        current_row += 1
        
        # Column headers for the virus table
        virus_headers = ['Virus_Name', 'Source', 'Box', 'Start_Conc', 'Total_Vol', 'Parts', 'Final_Conc']
        for col_idx, header in enumerate(virus_headers, 1):
            ws.cell(row=current_row, column=col_idx, value=header)
        header_row = current_row
        current_row += 1
        
        # Virus rows with formulas
        first_virus_row = current_row
        for virus_idx in range(viruses_per_group):
            row = current_row
            
            # Column A: Virus name placeholder
            if virus_idx == 0:
                ws.cell(row=row, column=1, value='[Virus 1]')
            elif virus_idx == 1:
                ws.cell(row=row, column=1, value='[Virus 2]')
            elif virus_idx == viruses_per_group - 1:
                ws.cell(row=row, column=1, value='[Filler]')
            else:
                ws.cell(row=row, column=1, value=f'[Virus {virus_idx + 1}]')
            
            # Column B: Source (e.g., "Miami", "Bhullar")
            # Leave blank
            
            # Column C: Box number
            # Leave blank
            
            # Column D: Starting concentration (user fills in)
            # Leave blank
            
            # Column E: Total Volume (constant)
            ws.cell(row=row, column=5, value=10)
            
            # Column F: Formula for parts
            if virus_idx < viruses_per_group - 1:
                # Regular virus: parts = Final * Total / Starting
                # Formula: =G{row}*E{row}/D{row}
                ws.cell(row=row, column=6, value=f'=IFERROR(G{row}*E{row}/D{row},"")')
            else:
                # Filler virus: remaining parts = Total - sum of other parts
                part_refs = '+'.join([f'F{first_virus_row + i}' for i in range(viruses_per_group - 1)])
                ws.cell(row=row, column=6, value=f'=IFERROR(E{row}-({part_refs}),"")')
            
            # Column G: Final/Target concentration (user fills in, or N/A for filler)
            if virus_idx == viruses_per_group - 1:
                ws.cell(row=row, column=7, value='N/A')
            # Otherwise leave blank for user input
            
            current_row += 1
        
        # Add blank rows between groups
        current_row += 2


def scan_cohort_files_for_planning(script_dir, exclude_file=None):
    """
    Scan directory for existing cohort files and extract their start dates
    
    Returns:
        List of dicts with cohort info: [{'name': 'CNT_01', 'start_date': datetime, 'file': path}, ...]
    """
    cohorts = []
    
    # Find all Connectome tracking files
    for f in sorted(script_dir.glob('Connectome_*_Animal_Tracking*.xlsx')):
        # Skip the file we're currently creating (if specified)
        if exclude_file and f.name == exclude_file:
            continue
        
        # Skip temp files
        if f.name.startswith('~'):
            continue
            
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            
            start_date = None
            cohort_name = None
            
            # Try to get start date from 3b_Manual_Tray (first date - 4 days = Day 0)
            if '3b_Manual_Tray' in wb.sheetnames:
                ws = wb['3b_Manual_Tray']
                first_date = ws.cell(row=2, column=1).value
                if first_date:
                    if isinstance(first_date, datetime):
                        start_date = first_date - timedelta(days=4)  # First test is Day 4
                    
                # Get cohort name from first SubjectID
                first_id = ws.cell(row=2, column=2).value  # Animal column
                if first_id and '_' in str(first_id):
                    parts = str(first_id).split('_')
                    if len(parts) >= 2:
                        cohort_name = f"{parts[0]}_{parts[1]}"  # e.g., "CNT_01"
            
            wb.close()
            
            if start_date and cohort_name:
                cohorts.append({
                    'name': cohort_name,
                    'start_date': start_date,
                    'file': f.name
                })
                
        except Exception as e:
            print(f"  Warning: Could not read {f.name}: {e}")
    
    return cohorts


def create_1_experiment_planning(start_date, subject_ids, cohort_name=None, script_dir=None):
    """
    Create 1_Experiment_Planning as a multi-cohort Gantt chart
    
    Scans directory for other cohort files and builds a timeline showing
    all cohorts' phases for easy overlap detection.
    
    Format:
    - Row 1: Header with dates
    - Row 2+: One row per cohort, cells contain phase abbreviations
    
    Returns dict with planning data (written specially by write function)
    """
    
    # If no script_dir provided, use current directory
    if script_dir is None:
        script_dir = SCRIPT_DIR
    
    # Get phases with day ranges
    phases = [
        ('FD', 0, 3),      # Food Dep
        ('FT', 4, 6),      # Flat Training
        ('PT', 7, 13),     # Pillar Training
        ('L3', 14, 16),    # Last 3
        ('INJ', 17, 17),   # Injury
        ('REC', 18, 24),   # Recovery
        ('PI1', 25, 25),   # Post-injury 1
        ('PI2', 32, 32),   # Post-injury 2
        ('PI3', 39, 39),   # Post-injury 3
        ('PI4', 46, 46),   # Post-injury 4
        ('RE', 47, 55),    # Rehab Easy
        ('RF', 56, 62),    # Rehab Flat
        ('RP', 63, 69),    # Rehab Pillar
        ('TRC', 70, 70),   # Tracing
        ('PERF', 84, 84),  # Perfusion
    ]
    
    def get_phase_for_day(day):
        """Return phase abbreviation for a given experiment day"""
        for abbrev, phase_start, phase_end in phases:
            if phase_start <= day <= phase_end:
                return abbrev
        return ''
    
    # Scan for existing cohorts
    existing_cohorts = scan_cohort_files_for_planning(script_dir)
    
    # Extract cohort name from subject_ids if not provided
    if cohort_name is None and subject_ids:
        first_id = subject_ids[0]
        if '_' in first_id:
            parts = first_id.split('_')
            cohort_name = f"{parts[0]}_{parts[1]}"
    
    # Add current cohort being created
    all_cohorts = existing_cohorts + [{
        'name': cohort_name or 'NEW',
        'start_date': start_date,
        'file': '(new)'
    }]
    
    # Sort by start date
    all_cohorts.sort(key=lambda x: x['start_date'])
    
    # Determine date range for the Gantt chart
    if all_cohorts:
        earliest_start = min(c['start_date'] for c in all_cohorts)
        latest_start = max(c['start_date'] for c in all_cohorts)
        chart_end = latest_start + timedelta(days=90)
    else:
        earliest_start = start_date
        chart_end = start_date + timedelta(days=90)
    
    # Generate list of dates
    num_days = (chart_end - earliest_start).days + 1
    
    # For very long timelines, group by week instead of day
    if num_days > 120:
        # Weekly view
        dates = []
        current = earliest_start
        while current <= chart_end:
            dates.append(current)
            current += timedelta(days=7)
        date_span = 7
    else:
        # Daily view
        dates = [earliest_start + timedelta(days=i) for i in range(num_days)]
        date_span = 1
    
    # Return data structure for special write function
    return {
        'cohorts': all_cohorts,
        'dates': dates,
        'date_span': date_span,
        'phases': phases,
        'get_phase_for_day': get_phase_for_day,
        '_is_gantt': True  # Flag for write function
    }


def write_1_experiment_planning_gantt(ws, planning_data):
    """
    Write the Gantt-style experiment planning sheet
    """
    cohorts = planning_data['cohorts']
    dates = planning_data['dates']
    get_phase_for_day = planning_data['get_phase_for_day']
    
    # Row 1: "Cohort" header + date headers
    ws.cell(row=1, column=1, value="Cohort")
    ws.cell(row=1, column=2, value="Start Date")
    
    for col_idx, date in enumerate(dates, 3):
        # Format date as MM/DD
        ws.cell(row=1, column=col_idx, value=date.strftime("%m/%d"))
    
    # Row 2+: One row per cohort
    for row_idx, cohort in enumerate(cohorts, 2):
        ws.cell(row=row_idx, column=1, value=cohort['name'])
        ws.cell(row=row_idx, column=2, value=cohort['start_date'].strftime("%Y-%m-%d"))
        
        cohort_start = cohort['start_date']
        
        for col_idx, chart_date in enumerate(dates, 3):
            # Calculate which experiment day this is for this cohort
            exp_day = (chart_date - cohort_start).days
            
            if exp_day >= 0:  # Only show phases after cohort starts
                phase = get_phase_for_day(exp_day)
                if phase:
                    ws.cell(row=row_idx, column=col_idx, value=phase)
    
    # Add legend below the chart
    legend_row = len(cohorts) + 4
    ws.cell(row=legend_row, column=1, value="LEGEND:")
    
    legend_items = [
        ('FD', 'Food Dep (Days 0-3)'),
        ('FT', 'Flat Training (Days 4-6)'),
        ('PT', 'Pillar Training (Days 7-13)'),
        ('L3', 'Last 3/Pre-Injury (Days 14-16)'),
        ('INJ', 'Injury Surgery (Day 17)'),
        ('REC', 'Recovery (Days 18-24)'),
        ('PI1-4', 'Post-Injury Tests (Days 25,32,39,46)'),
        ('RE', 'Rehab Easy (Days 47-55)'),
        ('RF', 'Rehab Flat (Days 56-62)'),
        ('RP', 'Rehab Pillar (Days 63-69)'),
        ('TRC', 'Tracing Surgery (Day 70)'),
        ('PERF', 'Perfusion (Day 84)'),
    ]
    
    for i, (abbrev, description) in enumerate(legend_items):
        ws.cell(row=legend_row + 1 + i, column=1, value=abbrev)
        ws.cell(row=legend_row + 1 + i, column=2, value=description)


def create_3a_manual_ramp(start_date, subject_ids):
    """
    Create 3a_Manual_Ramp sheet for food deprivation ramp (days 0-4)
    
    Wide format: One row per animal, columns grouped by day
    - Day 0: Date, Weight (baseline)
    - Day 1+: Date, Weight, % body weight, Tray Start (g), Tray End (g), Dif, Notes
    
    Formulas:
    - Dif = ABS(Tray End - Tray Start)
    - % body weight = Weight / Baseline Weight (Day 0)
    
    Returns: (headers, data_rows) for manual writing with formulas
    """
    # We'll return structured data instead of DataFrame since we need formulas
    return {
        'start_date': start_date,
        'subject_ids': subject_ids,
        'num_days': 5  # Days 0-4
    }


def write_3a_with_formulas(ws, ramp_data):
    """
    Write 3a_Manual_Ramp with Excel formulas
    
    Structure matches existing files:
    - Col A: Mouse ID
    - Col B: (spacer)
    - Day 0: Date, Weight (baseline)
    - Day 1-4: Date, Weight, % body weight, Tray Start, Tray End, Dif, Notes, (spacer)
    """
    start_date = ramp_data['start_date']
    subject_ids = ramp_data['subject_ids']
    
    # Build headers
    headers = ["Mouse ID", ""]  # A, B
    
    # Day 0 - just baseline
    headers.extend(["Date", "Weight", ""])  # C, D, E (baseline weight in D)
    
    # Days 1-4 - full columns
    for day in range(1, 5):
        headers.extend([
            "Date",           # Date for this day
            "Weight",         # Weight this day
            "% body weight",  # Formula: Weight / Baseline
            "Tray Start (g)", # Manual entry
            "Tray End (g)",   # Manual entry
            "Dif",            # Formula: ABS(End - Start)
            "Notes",          # Manual entry
            ""                # Spacer
        ])
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Column positions (1-indexed)
    COL_MOUSE_ID = 1
    COL_BASELINE_DATE = 3
    COL_BASELINE_WEIGHT = 4
    
    # Each day block after baseline is 8 columns wide
    def get_day_cols(day_num):
        """Get column indices for a given day (1-4)"""
        base = 6 + (day_num - 1) * 8  # Day 1 starts at col 6
        return {
            'date': base,
            'weight': base + 1,
            'pct': base + 2,
            'tray_start': base + 3,
            'tray_end': base + 4,
            'dif': base + 5,
            'notes': base + 6
        }
    
    # Write data rows
    for row_idx, subject_id in enumerate(subject_ids, 2):
        row = row_idx
        
        # Mouse ID
        ws.cell(row=row, column=COL_MOUSE_ID, value=subject_id)
        
        # Day 0 - baseline
        ws.cell(row=row, column=COL_BASELINE_DATE, value=start_date)
        # Weight left blank for manual entry
        
        # Days 1-4
        for day in range(1, 5):
            cols = get_day_cols(day)
            date = start_date + timedelta(days=day)
            
            # Date
            ws.cell(row=row, column=cols['date'], value=date)
            
            # Weight - blank for manual entry
            
            # % body weight formula: this day's weight / baseline weight
            weight_col = openpyxl.utils.get_column_letter(cols['weight'])
            baseline_col = openpyxl.utils.get_column_letter(COL_BASELINE_WEIGHT)
            ws.cell(row=row, column=cols['pct'], 
                    value=f"=IF({weight_col}{row}<>\"\",{weight_col}{row}/${baseline_col}{row},\"\")")
            
            # Tray Start, Tray End - blank for manual entry
            
            # Dif formula: ABS(Tray End - Tray Start)
            start_col = openpyxl.utils.get_column_letter(cols['tray_start'])
            end_col = openpyxl.utils.get_column_letter(cols['tray_end'])
            ws.cell(row=row, column=cols['dif'],
                    value=f"=IF(AND({start_col}{row}<>\"\",{end_col}{row}<>\"\"),ABS({end_col}{row}-{start_col}{row}),\"\")")


def create_3b_manual_tray(start_date, subject_ids):
    """
    Create 3b_Manual_Tray sheet - main pellet scoring sheet
    
    Pre-fills: Date, Animal, Test_Phase, Tray Type/Number
    Leaves blank: Sex (filled from metadata), Weight, pellet scores (1-20), Notes
    
    Includes formula columns for calculations:
    - Weight % (references baseline from 3a_Manual_Ramp)
    - Displaced, Retrieved, Contacted percentages
    - Skill Ratio
    - Average columns (cross-tray for same animal)
    - Max columns
    
    Row order: All animals for tray 1, then all animals for tray 2, etc.
    (This matches the order you'd actually run the experiment)
    """
    rows = []
    
    for day_offset, phase, tray_type, trays_per_day, notes in TIMELINE:
        date = start_date + timedelta(days=day_offset)
        
        # Loop tray first, then animal (so all animals do tray 1, then tray 2, etc.)
        for tray_num in range(1, trays_per_day + 1):
            for subject_id in subject_ids:
                row = {
                    "Date": date,
                    "Animal": subject_id,
                    "Sex": None,  # To be filled from metadata or manually
                    "Weight": None,  # Manual entry
                    "Weight %": None,  # Will be formula in Excel
                    "Test_Phase": phase,
                    "Tray Type/Number": f"{tray_type}{tray_num}",
                }
                # Add pellet columns 1-20 (blank for manual entry)
                for pellet in range(1, 21):
                    row[pellet] = None
                
                # Notes column BEFORE calculations
                row["Notes"] = None
                
                # Calculation columns (will be formulas in Excel)
                row["Displaced"] = None
                row["Retrieved"] = None  
                row["Contacted"] = None
                row["Skill Ratio"] = None
                
                rows.append(row)
    
    df = pd.DataFrame(rows)
    return df


def create_3c_manual_summary(start_date, subject_ids):
    """
    Create 3c_Manual_Summary sheet - wide format summary of Retrieved % by date
    
    Structure (matches existing files):
    - Row 1: Tray type header (Flat, Pillar, Easy, etc.)
    - Row 2: Date header + "retrieved" label in col A
    - Row 3+: Animals with Average_Retrieved for each date
    
    Formulas pull from 3b_Manual_Tray using AVERAGEIFS
    
    Returns dict with metadata for write_3c_with_formulas
    """
    return {
        'start_date': start_date,
        'subject_ids': subject_ids,
        'timeline': TIMELINE
    }


def write_3c_with_formulas(ws, data):
    """
    Write 3c_Manual_Summary with formulas pulling from 3b_Manual_Tray
    
    Wide format: one row per animal, columns for each testing date
    Values are Average_Retrieved (average retrieved % across all trays for that animal on that date)
    """
    start_date = data['start_date']
    subject_ids = data['subject_ids']
    
    # Build list of unique testing dates with their phases
    dates_and_phases = []
    current_phase_type = None
    
    for day_offset, phase, tray_type, trays_per_day, notes in TIMELINE:
        date = start_date + timedelta(days=day_offset)
        
        # Simplify phase name
        if "Flat" in phase and "Training" in phase:
            phase_type = "Flat"
        elif "Pillar" in phase and "Training" in phase:
            phase_type = "Pillar"
        elif "Pre-Injury" in phase or "Last 3" in phase:
            phase_type = "Pillar"
        elif "Post-Injury" in phase:
            phase_type = "Pillar"
        elif "Rehab_Easy" in phase or "Rehab Easy" in phase:
            phase_type = "Easy"
        elif "Rehab_Flat" in phase or "Rehab Flat" in phase:
            phase_type = "Flat"
        elif "Rehab_Pillar" in phase or "Rehab Pillar" in phase:
            phase_type = "Pillar"
        else:
            continue  # Skip non-testing days
        
        # Add spacer when phase changes
        if current_phase_type is not None and phase_type != current_phase_type:
            dates_and_phases.append((None, None))  # Spacer
        
        # Only add unique dates
        if not dates_and_phases or dates_and_phases[-1][1] != date:
            dates_and_phases.append((phase_type, date))
        
        current_phase_type = phase_type
    
    # Write row 1: Phase types ("Retrieved" in col A, then phase names)
    ws.cell(row=1, column=1, value="Retrieved")
    col = 2
    for phase_type, date in dates_and_phases:
        if phase_type:
            ws.cell(row=1, column=col, value=phase_type)
        col += 1
    
    # Write row 2: Dates ("retrieved" in col A, then dates)
    ws.cell(row=2, column=1, value="retrieved")
    col = 2
    for phase_type, date in dates_and_phases:
        if date:
            ws.cell(row=2, column=col, value=date)
        col += 1
    
    # Write animal rows with formulas
    for row_idx, subject_id in enumerate(subject_ids, 3):
        ws.cell(row=row_idx, column=1, value=subject_id)
        
        col = 2
        for phase_type, date in dates_and_phases:
            if date:
                # Formula: Get average Retrieved % for this animal on this date from 3b
                # =AVERAGEIFS('3b_Manual_Tray'!AD:AD,'3b_Manual_Tray'!B:B,A3,'3b_Manual_Tray'!A:A,B2)
                # Where AD = Retrieved column, B = Animal column, A = Date column
                ws.cell(row=row_idx, column=col,
                        value=f"=AVERAGEIFS('3b_Manual_Tray'!$AD:$AD,'3b_Manual_Tray'!$B:$B,$A{row_idx},'3b_Manual_Tray'!$A:$A,{get_column_letter(col)}$2)")
            col += 1


def write_3b_with_formulas(ws, df, num_mice):
    """
    Write 3b_Manual_Tray sheet with Excel formulas
    
    Uses context-based formulas (AVERAGEIFS, MAXIFS, SUMIFS) that work regardless
    of row count, row order, or number of mice. Formulas reference by meaning:
    - "This animal" = match Animal column
    - "This date" = match Date column
    - "This tray" = match Tray column
    
    Column structure:
    - A-G: Date, Animal, Sex, Weight, Weight%, Test_Phase, Tray Type/Number
    - H-AA: Pellets 1-20
    - AB: Notes
    - AC: Displaced (formula)
    - AD: Retrieved (formula)
    - AE: Contacted (formula)
    - AF: Skill Ratio (formula)
    - AG: (spacer)
    - AH-AK: (spacers)
    - AL: Average_Displaced - this animal, this date, all trays
    - AM: Average_Retrieved - this animal, this date, all trays
    - AN: Average_Contacted - this animal, this date, all trays
    - AO: Average_Skill_Ratio - this animal, this date, all trays
    - AP: Average_Average_Retrieved - all animals, this date
    - AQ: Average_Average_Contacted - all animals, this date
    - AR: (spacer)
    - AS: Max_Retrieved - this animal, this date, all trays
    - AT: Max_Contacted - this animal, this date, all trays
    - AU: Average_Max_Retrieved - all animals, this date
    - AV: Average_Max_Contacted - all animals, this date
    
    Args:
        ws: openpyxl worksheet
        df: DataFrame with the data
        num_mice: number of mice (not used for formulas anymore, kept for compatibility)
    """
    # Column letters for reference in formulas
    COL_DATE = 'A'
    COL_ANIMAL = 'B'
    COL_SEX = 'C'
    COL_WEIGHT = 'D'
    COL_WEIGHT_PCT = 'E'
    COL_PHASE = 'F'
    COL_TRAY = 'G'
    # H-AA = Pellets 1-20
    COL_NOTES = 'AB'
    COL_DISPLACED = 'AC'
    COL_RETRIEVED = 'AD'
    COL_CONTACTED = 'AE'
    COL_SKILL_RATIO = 'AF'
    # AG = spacer
    # AH-AK = spacers
    COL_AVG_DISPLACED = 'AL'
    COL_AVG_RETRIEVED = 'AM'
    COL_AVG_CONTACTED = 'AN'
    COL_AVG_SKILL_RATIO = 'AO'
    COL_AVG_AVG_RETRIEVED = 'AP'
    COL_AVG_AVG_CONTACTED = 'AQ'
    # AR = spacer
    COL_MAX_RETRIEVED = 'AS'
    COL_MAX_CONTACTED = 'AT'
    COL_AVG_MAX_RETRIEVED = 'AU'
    COL_AVG_MAX_CONTACTED = 'AV'
    
    # Column indices for writing (1-indexed)
    COL_IDX = {
        'Date': 1, 'Animal': 2, 'Sex': 3, 'Weight': 4, 'Weight %': 5,
        'Test_Phase': 6, 'Tray Type/Number': 7,
        # 8-27 = pellets 1-20
        'Notes': 28,
        'Displaced': 29, 'Retrieved': 30, 'Contacted': 31, 'Skill Ratio': 32,
        # 33 = spacer
        # 34-37 = spacers
        'Average_Displaced': 38, 'Average_Retrieved': 39, 
        'Average_Contacted': 40, 'Average_Skill_Ratio': 41,
        'Average_Average_Retrieved': 42, 'Average_Average_Contacted': 43,
        # 44 = spacer
        'Max_Retrieved': 45, 'Max_Contacted': 46,
        'Average_Max_Retrieved': 47, 'Average_Max_Contacted': 48
    }
    
    # Write header row
    headers = ["Date", "Animal", "Sex", "Weight", "Weight %", "Test_Phase", "Tray Type/Number"]
    headers.extend([str(i) for i in range(1, 21)])  # Pellets 1-20
    headers.append("Notes")
    headers.append("Displaced")
    headers.append("Retrieved")
    headers.append("Contacted")
    headers.append("Skill Ratio")
    headers.append("")  # Spacer
    headers.extend(["", "", "", ""])  # More spacers
    headers.append("Average_Displaced")
    headers.append("Average_Retrieved")
    headers.append("Average_Contacted")
    headers.append("Average_Skill_Ratio")
    headers.append("Average_Average_Retrieved")
    headers.append("Average_Average_Contacted")
    headers.append("")  # Spacer
    headers.append("Max_Retrieved")
    headers.append("Max_Contacted")
    headers.append("Average_Max_Retrieved")
    headers.append("Average_Max_Contacted")
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data rows with formulas
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        row = row_idx  # For formula strings
        
        # Static data columns
        ws.cell(row=row_idx, column=COL_IDX['Date'], value=row_data["Date"])
        ws.cell(row=row_idx, column=COL_IDX['Animal'], value=row_data["Animal"])
        ws.cell(row=row_idx, column=COL_IDX['Sex'], value=row_data["Sex"])
        ws.cell(row=row_idx, column=COL_IDX['Weight'], value=row_data["Weight"])
        ws.cell(row=row_idx, column=COL_IDX['Test_Phase'], value=row_data["Test_Phase"])
        ws.cell(row=row_idx, column=COL_IDX['Tray Type/Number'], value=row_data["Tray Type/Number"])
        
        # Pellet columns 8-27 are left blank for manual entry
        
        # === FORMULAS ===
        
        # Weight % - this animal's weight / baseline weight from ramp sheet
        # Baseline weight is in column D of 3a_Manual_Ramp, Mouse ID in column A
        ws.cell(row=row_idx, column=COL_IDX['Weight %'],
                value=f"=IF({COL_WEIGHT}{row}<>\"\",{COL_WEIGHT}{row}/SUMIF('3a_Manual_Ramp'!$A:$A,{COL_ANIMAL}{row},'3a_Manual_Ramp'!$D:$D),\"\")")
        
        # Displaced % - count of 1s in pellet columns / 20 * 100
        ws.cell(row=row_idx, column=COL_IDX['Displaced'],
                value=f"=COUNTIF(H{row}:AA{row},1)/20*100")
        
        # Retrieved % - count of 2s in pellet columns / 20 * 100
        ws.cell(row=row_idx, column=COL_IDX['Retrieved'],
                value=f"=COUNTIF(H{row}:AA{row},2)/20*100")
        
        # Contacted - sum of Displaced + Retrieved (not percentage, raw sum)
        ws.cell(row=row_idx, column=COL_IDX['Contacted'],
                value=f"=SUM({COL_DISPLACED}{row}:{COL_RETRIEVED}{row})")
        
        # Skill Ratio - Retrieved / Contacted * 100, handle div/0
        ws.cell(row=row_idx, column=COL_IDX['Skill Ratio'],
                value=f"=IF(ISNUMBER(({COL_RETRIEVED}{row}/{COL_CONTACTED}{row})*100),({COL_RETRIEVED}{row}/{COL_CONTACTED}{row})*100,0)")
        
        # Average_Displaced - average Displaced for THIS animal on THIS date (across all trays)
        ws.cell(row=row_idx, column=COL_IDX['Average_Displaced'],
                value=f"=AVERAGEIFS({COL_DISPLACED}:{COL_DISPLACED},{COL_ANIMAL}:{COL_ANIMAL},{COL_ANIMAL}{row},{COL_DATE}:{COL_DATE},{COL_DATE}{row})")
        
        # Average_Retrieved - average Retrieved for THIS animal on THIS date
        ws.cell(row=row_idx, column=COL_IDX['Average_Retrieved'],
                value=f"=AVERAGEIFS({COL_RETRIEVED}:{COL_RETRIEVED},{COL_ANIMAL}:{COL_ANIMAL},{COL_ANIMAL}{row},{COL_DATE}:{COL_DATE},{COL_DATE}{row})")
        
        # Average_Contacted - average Contacted for THIS animal on THIS date
        ws.cell(row=row_idx, column=COL_IDX['Average_Contacted'],
                value=f"=AVERAGEIFS({COL_CONTACTED}:{COL_CONTACTED},{COL_ANIMAL}:{COL_ANIMAL},{COL_ANIMAL}{row},{COL_DATE}:{COL_DATE},{COL_DATE}{row})")
        
        # Average_Skill_Ratio - average Skill Ratio for THIS animal on THIS date
        ws.cell(row=row_idx, column=COL_IDX['Average_Skill_Ratio'],
                value=f"=AVERAGEIFS({COL_SKILL_RATIO}:{COL_SKILL_RATIO},{COL_ANIMAL}:{COL_ANIMAL},{COL_ANIMAL}{row},{COL_DATE}:{COL_DATE},{COL_DATE}{row})")
        
        # Max_Retrieved - max Retrieved for THIS animal on THIS date
        ws.cell(row=row_idx, column=COL_IDX['Max_Retrieved'],
                value=f"=MAXIFS({COL_RETRIEVED}:{COL_RETRIEVED},{COL_ANIMAL}:{COL_ANIMAL},{COL_ANIMAL}{row},{COL_DATE}:{COL_DATE},{COL_DATE}{row})")
        
        # Max_Contacted - max Contacted for THIS animal on THIS date
        ws.cell(row=row_idx, column=COL_IDX['Max_Contacted'],
                value=f"=MAXIFS({COL_CONTACTED}:{COL_CONTACTED},{COL_ANIMAL}:{COL_ANIMAL},{COL_ANIMAL}{row},{COL_DATE}:{COL_DATE},{COL_DATE}{row})")
        
        # Average_Average and Average_Max columns - ONLY in first tray of each day
        # Check if this is the first tray by looking at the tray type/number
        tray_label = row_data["Tray Type/Number"]
        is_first_tray = tray_label.endswith("1")  # F1, P1, E1 are first trays
        
        if is_first_tray:
            # Average_Average_Retrieved - average of all animals' Average_Retrieved on THIS date
            ws.cell(row=row_idx, column=COL_IDX['Average_Average_Retrieved'],
                    value=f"=AVERAGEIF({COL_DATE}:{COL_DATE},{COL_DATE}{row},{COL_AVG_RETRIEVED}:{COL_AVG_RETRIEVED})")
            
            # Average_Average_Contacted - average of all animals' Average_Contacted on THIS date
            ws.cell(row=row_idx, column=COL_IDX['Average_Average_Contacted'],
                    value=f"=AVERAGEIF({COL_DATE}:{COL_DATE},{COL_DATE}{row},{COL_AVG_CONTACTED}:{COL_AVG_CONTACTED})")
            
            # Average_Max_Retrieved - average of all animals' Max_Retrieved on THIS date
            ws.cell(row=row_idx, column=COL_IDX['Average_Max_Retrieved'],
                    value=f"=AVERAGEIF({COL_DATE}:{COL_DATE},{COL_DATE}{row},{COL_MAX_RETRIEVED}:{COL_MAX_RETRIEVED})")
            
            # Average_Max_Contacted - average of all animals' Max_Contacted on THIS date
            ws.cell(row=row_idx, column=COL_IDX['Average_Max_Contacted'],
                    value=f"=AVERAGEIF({COL_DATE}:{COL_DATE},{COL_DATE}{row},{COL_MAX_CONTACTED}:{COL_MAX_CONTACTED})")


def write_3d_weights_flip(ws, df, start_date, subject_ids):
    """
    Write 3d_Weights sheet in flip chart format
    Row 1: Phase names
    Row 2: Dates
    Row 3+: Animals with weight values
    """
    # Build unique dates from timeline
    dates_and_phases = []
    current_phase_type = None
    
    for day_offset, phase, tray_type, trays_per_day, notes in TIMELINE:
        date = start_date + timedelta(days=day_offset)
        
        # Determine phase type for grouping
        if "Flat" in phase and "Training" in phase:
            phase_type = "Flat"
        elif "Pillar" in phase and "Training" in phase:
            phase_type = "Pillar"
        elif "Pre-Injury" in phase:
            phase_type = "Pillar"
        elif "Post-Injury" in phase:
            phase_type = "Pillar"
        elif "Rehab_Easy" in phase:
            phase_type = "Easy"
        elif "Rehab_Flat" in phase:
            phase_type = "Flat"
        elif "Rehab_Pillar" in phase:
            phase_type = "Pillar"
        else:
            phase_type = "Other"
        
        # Check if we need an average column
        if current_phase_type is not None and phase_type != current_phase_type:
            dates_and_phases.append(("Average", None))
        
        # Only add unique dates
        if not dates_and_phases or dates_and_phases[-1][1] != date:
            dates_and_phases.append((phase_type, date))
        
        current_phase_type = phase_type
    
    # Final average
    dates_and_phases.append(("Average", None))
    
    # Write row 1: Phase types
    ws.cell(row=1, column=1, value=None)  # Empty corner cell
    col = 2
    for phase_type, date in dates_and_phases:
        ws.cell(row=1, column=col, value=phase_type)
        col += 1
    
    # Write row 2: Dates
    ws.cell(row=2, column=1, value=None)
    col = 2
    for phase_type, date in dates_and_phases:
        if date:
            ws.cell(row=2, column=col, value=date)
        col += 1
    
    # Write animal rows
    for row_idx, subject_id in enumerate(subject_ids, 3):
        ws.cell(row=row_idx, column=1, value=subject_id)
        
        # Add average formulas for each phase group
        col = 2
        phase_start_col = 2
        current_avg_phase = None
        
        for phase_type, date in dates_and_phases:
            if phase_type == "Average":
                # Write average formula for the preceding phase
                if phase_start_col < col:
                    start_letter = get_column_letter(phase_start_col)
                    end_letter = get_column_letter(col - 1)
                    ws.cell(row=row_idx, column=col, 
                            value=f"=AVERAGE({start_letter}{row_idx}:{end_letter}{row_idx})")
                phase_start_col = col + 1
            col += 1


def create_3d_weights(start_date, subject_ids):
    """
    Create 3d_Weights sheet in flip chart format (animals as rows, dates as columns)
    Matches the existing 3e_Weights structure from Connectome_1
    
    Format:
    Row 0: Phase names (Flat, Pillar, etc.)
    Row 1: Dates
    Row 2+: Animal weights (one animal per row)
    With "Average" columns between phases
    """
    # Build the timeline with phases
    phases_and_dates = []
    
    # Group timeline entries by phase for averages
    current_phase = None
    phase_entries = []
    
    for day_offset, phase, tray_type, trays_per_day, notes in TIMELINE:
        date = start_date + timedelta(days=day_offset)
        
        # Simplify phase name for header
        if "Training" in phase and "Flat" in phase:
            header_phase = "Flat"
        elif "Training" in phase and "Pillar" in phase:
            header_phase = "Pillar"
        elif "Last 3" in phase:
            header_phase = "Pillar"  # Last 3 is pillar testing
        elif "Post-injury" in phase:
            header_phase = "Pillar"
        elif "Rehab Easy" in phase:
            header_phase = "Easy"
        elif "Rehab Flat" in phase:
            header_phase = "Flat"
        elif "Rehab Pillar" in phase:
            header_phase = "Pillar"
        else:
            header_phase = phase
        
        # Check if phase changed - add average column
        if current_phase is not None and header_phase != current_phase:
            phases_and_dates.append(("Average", None, current_phase))
        
        current_phase = header_phase
        phases_and_dates.append((header_phase, date, phase))
    
    # Add final average
    if current_phase:
        phases_and_dates.append(("Average", None, current_phase))
    
    # Build dataframe
    # First column is animal ID, rest are dates/averages
    columns = ["Animal"]
    row0_phases = [None]  # Phase header row
    row1_dates = [None]   # Date header row
    
    for header_phase, date, full_phase in phases_and_dates:
        if date is None:  # Average column
            columns.append(f"Avg_{header_phase}_{len([c for c in columns if 'Avg' in str(c)])}")
            row0_phases.append("Average")
            row1_dates.append(None)
        else:
            columns.append(date.strftime("%Y-%m-%d"))
            row0_phases.append(header_phase)
            row1_dates.append(date)
    
    # Create data rows (one per animal, values blank for manual entry)
    rows = []
    for subject_id in subject_ids:
        row = {"Animal": subject_id}
        for col in columns[1:]:
            row[col] = None
        rows.append(row)
    
    df = pd.DataFrame(rows, columns=columns)
    
    # We'll need to handle the header rows specially when writing to Excel
    # For now, store the phase info as metadata
    df.attrs['phase_row'] = row0_phases
    df.attrs['date_row'] = row1_dates
    
    return df


def create_4_contusion_injury_details(subject_ids, start_date):
    """
    Create 4_Contusion_Injury_Details sheet
    Returns DataFrame - formulas will be added when writing to Excel
    """
    injury_date = start_date + timedelta(days=INJURY_DAY)
    
    rows = []
    for subject_id in subject_ids:
        rows.append({
            "Subject_ID": subject_id,
            "Surgery_Date": injury_date,
            "Surgery_Type": "Contusion",
            "Surgery_Severity": None,  # e.g., "60kd"
            "Contusion_Location": "C5",
            "Subject_Weight (g)": None,
            "Anesthetic": "Ketamine/Xylazine",
            "Anesthetic_Dose": "100mg/10mg/kg",
            "Anesthetic_Volume": None,  # Formula added in write function
            "Analgesic": "Meloxicam",
            "Analgesic_Dose": "5mg/kg",
            "Analgesic_Volume": None,  # Formula added in write function
            "Intended_kd": 60,
            "Intended_Dwell": 0,
            "Stage_Height": None,
            "Actual_kd": None,
            "Actual_displacement": None,
            "Actual_Velocity": None,
            "Actual_Dwell": None,
            "Survived": None  # Y/N
        })
    
    df = pd.DataFrame(rows)
    return df


def write_4_contusion_with_formulas(ws, df):
    """
    Write 4_Contusion_Injury_Details with volume calculation formulas
    """
    # Column indices (1-indexed)
    COL_WEIGHT = 6  # Column F - Subject_Weight (g)
    COL_ANESTHETIC_VOL = 9  # Column I
    COL_ANALGESIC_VOL = 12  # Column L
    
    headers = list(df.columns)
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data rows with formulas
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(headers, 1):
            if col_name == "Anesthetic_Volume":
                # Formula: =F{row}/100&"mL"
                ws.cell(row=row_idx, column=col_idx, 
                        value=f'=IF(F{row_idx}<>"",F{row_idx}/100&"mL","")')
            elif col_name == "Analgesic_Volume":
                # Formula: =F{row}/100&"mL"  
                ws.cell(row=row_idx, column=col_idx,
                        value=f'=IF(F{row_idx}<>"",F{row_idx}/100&"mL","")')
            else:
                ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])


def create_5_sc_injection_details(subject_ids, start_date, max_viruses=3):
    """
    Create 5_SC_Injection_Details sheet for viral injection surgery
    Returns DataFrame - formulas will be added when writing to Excel
    
    Key column: Virus_Group - references a group in 0_Virus_Preparation
    Script 1 will populate virus columns based on Virus_Group assignment.
    
    User enters:
    - Virus_Group (e.g., "ENCR.01") OR fills in virus columns directly
    - Weight, Survived, Signal, Notes
    
    Script 1 populates (from 0_Virus_Preparation based on Virus_Group):
    - Virus_N_Name, Virus_N_Titer, etc.
    - Surgery_Date, Injection_Location (from group assignments)
    
    Args:
        subject_ids: List of subject IDs
        start_date: Start date of experiment
        max_viruses: Max viruses per injection (default 3)
    """
    tracing_date = start_date + timedelta(days=TRACING_DAY)
    
    # Build column list with multi-virus support
    base_cols = [
        "Subject_ID",
        "Virus_Group",  # NEW: References group in 0_Virus_Preparation (e.g., "ENCR.01")
        "Surgery_Date", 
        "Subject_Weight (g)",
        "Surgery_Type",
        "Injection_Location",
        "Depths (D/V)",
        "Coordinates (M/L)",
    ]
    
    # Add per-virus columns
    virus_cols = []
    for v in range(1, max_viruses + 1):
        virus_cols.extend([
            f"Virus_{v}_Name",
            f"Virus_{v}_Titer",
            f"Virus_{v}_Source",
            f"Virus_{v}_Target",
        ])
    
    # Add shared columns after virus info
    end_cols = [
        "Anesthetic",
        "Anesthetic_Dose",
        "Anesthetic_Volume",
        "Analgesic",
        "Analgesic_Dose",
        "Analgesic_Volume",
        "Survived",
        "Signal Post Perfusion",
    ]
    
    all_cols = base_cols + virus_cols + end_cols
    
    rows = []
    for subject_id in subject_ids:
        row = {
            "Subject_ID": subject_id,
            "Virus_Group": None,  # User fills in e.g., "ENCR.01" - Script 1 uses this to populate virus cols
            "Surgery_Date": tracing_date,
            "Subject_Weight (g)": None,
            "Surgery_Type": "Spinal cord virus injection",
            "Injection_Location": "Caudal C6",
            "Depths (D/V)": "-0.6/-0.8",
            "Coordinates (M/L)": "+0.3/-0.3",
            "Anesthetic": "Ketamine/Xylazine",
            "Anesthetic_Dose": "100mg/10mg/kg",
            "Anesthetic_Volume": None,
            "Analgesic": "Meloxicam",
            "Analgesic_Dose": "5mg/kg",
            "Analgesic_Volume": None,
            "Survived": None,
            "Signal Post Perfusion": None,
        }
        # Add empty virus columns
        for v in range(1, max_viruses + 1):
            row[f"Virus_{v}_Name"] = None
            row[f"Virus_{v}_Titer"] = None
            row[f"Virus_{v}_Source"] = None
            row[f"Virus_{v}_Target"] = None
        
        rows.append(row)
    
    df = pd.DataFrame(rows, columns=all_cols)
    return df


def write_5_sc_injection_with_formulas(ws, df):
    """
    Write 5_SC_Injection_Details with volume calculation formulas
    Handles multi-virus column structure
    """
    headers = list(df.columns)
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Find weight column for volume formulas
    weight_col = None
    for i, h in enumerate(headers, 1):
        if 'weight' in h.lower() and 'volume' not in h.lower():
            weight_col = i
            break
    
    # Write data rows with formulas
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(headers, 1):
            if col_name == "Anesthetic_Volume" and weight_col:
                ws.cell(row=row_idx, column=col_idx, 
                        value=f'=IF({get_column_letter(weight_col)}{row_idx}<>"",{get_column_letter(weight_col)}{row_idx}/100&"mL","")')
            elif col_name == "Analgesic_Volume" and weight_col:
                ws.cell(row=row_idx, column=col_idx,
                        value=f'=IF({get_column_letter(weight_col)}{row_idx}<>"",{get_column_letter(weight_col)}{row_idx}/100&"mL","")')
            else:
                ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])


def create_6_ladder(subject_ids):
    """
    Create 6_Ladder sheet for ladder test tracking
    """
    df = pd.DataFrame(columns=[
        "Date",
        "Animal",
        "Test_Type",  # Uninjured/Injured
        "Camera_Type",
        "Total_Steps",
        "Total_Misses",
        "Miss_Pct",
        "Notes"
    ])
    return df


def create_7_stats(subject_ids):
    """
    Create 7_Stats sheet with formulas averaging Retrieved % by phase
    
    Structure:
    - Row 1: Headers (blank, Mouse, Flat Training, Pillar Training, Last 3, etc.)
    - Row 2+: "Retrieved" label + Animal + formulas for each phase
    
    Returns dict for write_7_stats_with_formulas
    """
    return {
        'subject_ids': subject_ids
    }


def write_7_stats_with_formulas(ws, data):
    """
    Write 7_Stats with formulas referencing 3c_Manual_Summary
    
    Each cell averages the relevant columns from 3c for that animal's phase
    """
    subject_ids = data['subject_ids']
    
    # Phase definitions - which columns in 3c correspond to each phase
    # These are the phase groupings for averaging
    phases = [
        "Flat Training",    # Cols B-D in 3c (days 4-6)
        "Pillar Training",  # Cols F-L in 3c (days 7-13)  
        "Last 3",           # Cols N-P in 3c (days 14-16, pre-injury baseline)
        "Post Injury 1",    # Single column (day 25)
        "Post Injury 2-4",  # 3 columns (days 32, 39, 46)
        "Rehab Easy",       # Multiple columns (days 47-55)
        "Rehab Flat",       # Multiple columns (days 56-62)
        "Rehab Pillar"      # Multiple columns (days 63-69)
    ]
    
    # Write header row
    headers = ["", "Mouse"] + phases
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write animal rows
    # The formulas will reference 3c_Manual_Summary
    # Since 3c structure depends on timeline, we use AVERAGEIF with phase matching
    # But 3c doesn't have phase column - it has dates
    # 
    # Alternative approach: Reference specific column ranges in 3c
    # This requires knowing the 3c column layout
    #
    # For now, create placeholder formulas that reference 3c by row
    # The actual column ranges would need to be determined from 3c structure
    
    for row_idx, subject_id in enumerate(subject_ids, 2):
        ws.cell(row=row_idx, column=1, value="Retrieved")
        ws.cell(row=row_idx, column=2, value=subject_id)
        
        # Find this animal's row in 3c (row 3 = first animal, row 4 = second, etc.)
        animal_row_3c = row_idx + 1  # +1 because 3c has 2 header rows
        
        # Phase formulas - these reference column ranges in 3c
        # The actual ranges depend on how many testing days are in each phase
        # Using INDIRECT and MATCH would be more robust but complex
        # For now, placeholder formulas:
        
        # Flat Training (typically first 3 testing days)
        ws.cell(row=row_idx, column=3, 
                value=f"=AVERAGE('3c_Manual_Summary'!B{animal_row_3c}:D{animal_row_3c})")
        
        # Pillar Training (next ~7 days)
        ws.cell(row=row_idx, column=4,
                value=f"=AVERAGE('3c_Manual_Summary'!F{animal_row_3c}:L{animal_row_3c})")
        
        # Last 3 (pre-injury baseline, 3 days)
        ws.cell(row=row_idx, column=5,
                value=f"=AVERAGE('3c_Manual_Summary'!N{animal_row_3c}:P{animal_row_3c})")
        
        # Post Injury 1 (single day)
        ws.cell(row=row_idx, column=6,
                value=f"='3c_Manual_Summary'!R{animal_row_3c}")
        
        # Post Injury 2-4 (3 days)
        ws.cell(row=row_idx, column=7,
                value=f"=AVERAGE('3c_Manual_Summary'!S{animal_row_3c}:U{animal_row_3c})")
        
        # Rehab Easy
        ws.cell(row=row_idx, column=8,
                value=f"=AVERAGE('3c_Manual_Summary'!W{animal_row_3c}:AE{animal_row_3c})")
        
        # Rehab Flat  
        ws.cell(row=row_idx, column=9,
                value=f"=AVERAGE('3c_Manual_Summary'!AG{animal_row_3c}:AM{animal_row_3c})")
        
        # Rehab Pillar
        ws.cell(row=row_idx, column=10,
                value=f"=AVERAGE('3c_Manual_Summary'!AO{animal_row_3c}:AU{animal_row_3c})")


def create_8_brainglobe():
    """
    Create 8_BrainGlobe sheet placeholder for histology import
    """
    df = pd.DataFrame(columns=[
        "Subject_ID",
        "Perfusion_Date",
        "Analysis_Date",
        "Atlas_Used",
        "Total_Cells_Detected",
        "Total_Cells_Left_Hemisphere",
        "Total_Cells_Right_Hemisphere",
        "Quality",
        "Notes"
        # Region-specific columns will be added by import script
    ])
    return df


def create_9_dlc_kinematics(subject_ids, start_date):
    """
    Create 9_DLC_Kinematics sheet - per-day structure matching ASPA output
    
    Structure: One row per animal per day
    ~687 columns:
    - Animal, Date (2)
    - Attention scores: Tray1-4 + Total (5)
    - Per-pellet: 8 metrics × 20 pellets × 4 trays (640)
    - Tray averages: 8 metrics × 4 trays (32)
    - Day averages: 8 metrics (8)
    """
    return {
        'subject_ids': subject_ids,
        'start_date': start_date,
        'timeline': TIMELINE
    }


# Kinematic metrics from ASPA output
KINEMATIC_METRICS = [
    "Swipe_breadth",
    "Swipe_length", 
    "Swipe_area",
    "Swipe_speed",
    "Path_length",
    "Swipe_Duration",
    "Swipe_Duration_Frames",
    "Path_over_Frames"
]


def write_9_dlc_with_formulas(ws, data):
    """
    Write 9_DLC_Kinematics sheet with per-day structure
    Values are blank placeholders for importer; averages have formulas
    """
    subject_ids = data['subject_ids']
    start_date = data['start_date']
    
    # Build column headers
    columns = ["Animal", "Date"]
    
    # Attention scores (5)
    columns.extend([
        "Tray1_Attention_Score", "Tray2_Attention_Score",
        "Tray3_Attention_Score", "Tray4_Attention_Score",
        "Total_Day_Attention_Score"
    ])
    
    # Per-pellet kinematics (640 columns)
    # Tray{1-4}_Pellet{01-20}_{metric}
    for tray in range(1, 5):
        for pellet in range(1, 21):
            for metric in KINEMATIC_METRICS:
                columns.append(f"Tray{tray}_Pellet{pellet:02d}_{metric}")
    
    # Per-tray averages (32 columns)
    for tray in range(1, 5):
        for metric in KINEMATIC_METRICS:
            columns.append(f"Tray{tray}_Avg_{metric}")
    
    # Day averages (8 columns)
    for metric in KINEMATIC_METRICS:
        columns.append(f"Day_Avg_{metric}")
    
    # Create column index lookup
    col_idx = {name: i+1 for i, name in enumerate(columns)}
    
    # Write headers
    for i, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=i, value=col_name)
    
    # Generate rows - one per animal per day
    row_num = 2
    for day_offset, phase, tray_type, trays_per_day, notes in TIMELINE:
        date = start_date + timedelta(days=day_offset)
        
        for subject_id in subject_ids:
            row = row_num
            
            # Animal and Date
            ws.cell(row=row, column=col_idx["Animal"], value=subject_id)
            ws.cell(row=row, column=col_idx["Date"], value=date)
            
            # Attention scores - blank for importer
            # Per-pellet columns - blank for importer
            
            # Tray average formulas
            for tray in range(1, 5):
                for metric in KINEMATIC_METRICS:
                    avg_col = col_idx[f"Tray{tray}_Avg_{metric}"]
                    
                    # Build range of pellet columns for this tray+metric
                    first_pellet_col = col_idx[f"Tray{tray}_Pellet01_{metric}"]
                    last_pellet_col = col_idx[f"Tray{tray}_Pellet20_{metric}"]
                    first_letter = get_column_letter(first_pellet_col)
                    last_letter = get_column_letter(last_pellet_col)
                    
                    # AVERAGE formula across all 20 pellets for this tray+metric
                    ws.cell(row=row, column=avg_col,
                            value=f"=IFERROR(AVERAGE({first_letter}{row}:{last_letter}{row}),\"\")")
            
            # Day average formulas
            for metric in KINEMATIC_METRICS:
                day_avg_col = col_idx[f"Day_Avg_{metric}"]
                
                # Average of 4 tray averages
                tray_avg_refs = [f"{get_column_letter(col_idx[f'Tray{t}_Avg_{metric}'])}{row}" for t in range(1, 5)]
                ws.cell(row=row, column=day_avg_col,
                        value=f"=IFERROR(AVERAGE({','.join(tray_avg_refs)}),\"\")")
            
            row_num += 1
    
    # Autoscale columns
    autoscale_columns(ws)


def create_2_odc_animal_tracking(subject_ids, start_date, project_type='cnt', max_viruses=3):
    """
    Create 2_ODC_Animal_Tracking sheet structure
    Returns dict for write_2_odc_with_formulas
    
    NEW STRUCTURE: One row per animal per DAY (not per tray)
    ~884 columns pulling from all other sheets via formulas
    
    Args:
        subject_ids: List of subject IDs
        start_date: Start date
        project_type: 'cnt' or 'encr'
        max_viruses: Max viruses per injection (default 3)
    """
    return {
        'subject_ids': subject_ids,
        'start_date': start_date,
        'timeline': TIMELINE,
        'project_type': project_type,
        'max_viruses': max_viruses
    }


def write_2_odc_with_formulas(ws, data, cohort_name):
    """
    Write 2_ODC_Animal_Tracking with formulas pulling from other sheets
    
    NEW STRUCTURE: One row per animal per DAY
    - 16 mice × 38 days = 608 rows
    - All 4 trays as columns (80 pellet columns instead of 20) - CNT only
    - Kinematic data section pulls from 9_DLC_Kinematics - CNT only
    - ENCR mode: No behavior columns
    """
    subject_ids = data['subject_ids']
    start_date = data['start_date']
    project_type = data.get('project_type', 'cnt')
    max_viruses = data.get('max_viruses', 3)
    
    # Build column headers
    columns = []
    
    # Section 1: ODC-SCI Required CoDEs (17 columns)
    columns.extend([
        "SubjectID",           # Direct
        "SpeciesTyp",          # Hard-coded
        "SpeciesStrainTyp",    # Hard-coded
        "AnimalSourceNam",     # Hard-coded
        "AgeVal",              # Formula from 0a_Metadata
        "BodyWgtMeasrVal",     # Formula from 3a_Manual_Ramp baseline (CNT) or 0a_Metadata (ENCR)
        "SexTyp",              # Formula from 0a_Metadata
        "InjGroupAssignTyp",   # Cohort name
        "Laboratory",          # Hard-coded
        "StudyLeader",         # Hard-coded
        "Exclusion_in_origin_study",  # Formula from 4_Contusion.Survived
        "Exclusion_reason",    # Formula based on Survived
        "Cause_of_Death",      # Formula based on Survived
        "Injury_type",         # Formula from 4_Contusion
        "Injury_device",       # Hard-coded
        "Injury_level",        # Formula from 4_Contusion
        "Injury_details"       # Formula concatenating 4_Contusion fields
    ])
    
    # Section 2: Contusion Surgery Details (Surgery_1_*) - 19 columns
    columns.extend([
        "Surgery_1_Date", "Surgery_1_Type", "Surgery_1_Severity", "Surgery_1_Location",
        "Surgery_1_Weight_g", "Surgery_1_Anesthetic", "Surgery_1_Anesthetic_Dose",
        "Surgery_1_Anesthetic_Volume", "Surgery_1_Analgesic", "Surgery_1_Analgesic_Dose",
        "Surgery_1_Analgesic_Volume", "Surgery_1_Intended_kd", "Surgery_1_Intended_Dwell",
        "Surgery_1_Stage_Height", "Surgery_1_Actual_kd", "Surgery_1_Actual_Displacement",
        "Surgery_1_Actual_Velocity", "Surgery_1_Actual_Dwell", "Surgery_1_Survived"
    ])
    
    # Section 3: SC Injection Details (Surgery_2_*) with multi-virus
    columns.extend([
        "Surgery_2_Date", "Surgery_2_Weight_g", "Surgery_2_Type", 
        "Surgery_2_Location", "Surgery_2_Depth_DV", "Surgery_2_Coord_ML",
    ])
    # Add per-virus columns
    for v in range(1, max_viruses + 1):
        columns.extend([
            f"Surgery_2_Virus_{v}_Name",
            f"Surgery_2_Virus_{v}_Titer", 
            f"Surgery_2_Virus_{v}_Source",
            f"Surgery_2_Virus_{v}_Target",
        ])
    # Continue with shared injection columns
    columns.extend([
        "Surgery_2_Anesthetic", "Surgery_2_Anesthetic_Dose", "Surgery_2_Anesthetic_Volume",
        "Surgery_2_Analgesic", "Surgery_2_Analgesic_Dose", "Surgery_2_Analgesic_Volume",
        "Surgery_2_Survived", "Surgery_2_Signal_Post_Perfusion"
    ])
    
    # Section 4: BrainGlobe Placeholders (8 columns) - manual/import
    columns.extend([
        "Perfusion_Date", "BrainGlobe_Analysis_Date", "BrainGlobe_Atlas_Used",
        "Total_Cells_Detected", "Total_Cells_Left_Hemisphere", "Total_Cells_Right_Hemisphere",
        "BrainGlobe_Notes", "BrainGlobe_Quality"
    ])
    
    # Section 5: Row-Level Metadata
    if project_type == 'cnt':
        # Full behavior metadata (6 columns - per day)
        columns.extend([
            "Date", "Test_Phase", "Days_Post_Injury", "Tray_Type", "Num_Trays", "Weight"
        ])
        
        # Section 6: Per-Pellet Manual Scores (80 columns)
        # All 4 trays × 20 pellets
        for tray in range(1, 5):
            for pellet in range(1, 21):
                columns.append(f"Tray{tray}_Pellet{pellet:02d}")
        
        # Section 7: Per-Tray Manual Calculations (36 columns)
        for tray in range(1, 5):
            columns.extend([
                f"Tray{tray}_Presented", f"Tray{tray}_Missed", f"Tray{tray}_Displaced",
                f"Tray{tray}_Retrieved", f"Tray{tray}_Contacted", f"Tray{tray}_Miss_Pct",
                f"Tray{tray}_Displaced_Pct", f"Tray{tray}_Retrieved_Pct", f"Tray{tray}_Contacted_Pct"
            ])
        
        # Section 8: Daily Manual Totals/Averages (13 columns)
        columns.extend([
            "Total_Presented", "Total_Missed", "Total_Displaced", "Total_Retrieved", "Total_Contacted",
            "Avg_Miss_Pct", "Avg_Displaced_Pct", "Avg_Retrieved_Pct", "Avg_Contacted_Pct",
            "Max_Retrieved_Pct", "Max_Contacted_Pct", "Min_Retrieved_Pct", "Min_Contacted_Pct"
        ])
        
        # Section 9: Kinematic Attention Scores (5 columns)
        columns.extend([
            "Tray1_Attention_Score", "Tray2_Attention_Score",
            "Tray3_Attention_Score", "Tray4_Attention_Score",
            "Total_Day_Attention_Score"
        ])
        
        # Section 10: Per-Pellet Kinematics (640 columns)
        for tray in range(1, 5):
            for pellet in range(1, 21):
                for metric in KINEMATIC_METRICS:
                    columns.append(f"Tray{tray}_Pellet{pellet:02d}_{metric}")
        
        # Section 11: Kinematic Averages (40 columns)
        for tray in range(1, 5):
            for metric in KINEMATIC_METRICS:
                columns.append(f"Tray{tray}_Avg_{metric}")
        for metric in KINEMATIC_METRICS:
            columns.append(f"Day_Avg_{metric}")
    else:
        # ENCR mode: Minimal metadata, no behavior
        columns.extend([
            "Date", "Days_Post_Surgery", "Notes"
        ])
    
    # Section 12: Source Tracking (3 columns)
    columns.extend(["Source_File", "Source_Sheet", "Row_Notes"])
    
    # Create column index lookup
    col_idx = {name: i+1 for i, name in enumerate(columns)}
    
    # Write headers
    for i, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=i, value=col_name)
    
    # 3b column references for pellet lookup
    # 3b has columns: A=Date, B=Animal, C=Sex, D=Weight, E=Weight%, F=Test_Phase, G=Tray Type/Number, H-AA=Pellets 1-20
    # Pellet columns in 3b: H=1, I=2, J=3, K=4, L=5, M=6, N=7, O=8, P=9, Q=10, R=11, S=12, T=13, U=14, V=15, W=16, X=17, Y=18, Z=19, AA=20
    PELLET_COLS_3B = {i: get_column_letter(i + 7) for i in range(1, 21)}  # 1->H, 2->I, etc.
    
    # Generate rows - one per animal per day
    row_num = 2
    for day_offset, phase, tray_type, trays_per_day, notes in TIMELINE:
        date = start_date + timedelta(days=day_offset)
        
        for subject_id in subject_ids:
            row = row_num
            
            # === SECTION 1: ODC-SCI Required CoDEs ===
            ws.cell(row=row, column=col_idx["SubjectID"], value=subject_id)
            ws.cell(row=row, column=col_idx["SpeciesTyp"], value="Mouse")
            ws.cell(row=row, column=col_idx["SpeciesStrainTyp"], value="C57BL/6J")
            ws.cell(row=row, column=col_idx["AnimalSourceNam"], value="Jackson Laboratory")
            ws.cell(row=row, column=col_idx["InjGroupAssignTyp"], value=cohort_name)
            ws.cell(row=row, column=col_idx["Laboratory"], value="Murray/Blackmore Lab")
            ws.cell(row=row, column=col_idx["StudyLeader"], value="Logan Friedrich")
            ws.cell(row=row, column=col_idx["Injury_device"], value="Infinite Horizon Impactor")
            
            # AgeVal - from 0a_Metadata DOB to current date (weeks)
            # Only calculate if DOB exists (not blank)
            date_col = get_column_letter(col_idx['Date'])
            ws.cell(row=row, column=col_idx["AgeVal"],
                    value=f"=IFERROR(IF(INDEX('0a_Metadata'!$B:$B,MATCH($A{row},'0a_Metadata'!$A:$A,0))=\"\",\"\",DATEDIF(INDEX('0a_Metadata'!$B:$B,MATCH($A{row},'0a_Metadata'!$A:$A,0)),{date_col}{row},\"D\")/7),\"\")")
            
            # BodyWgtMeasrVal - baseline weight
            if project_type == 'cnt':
                # From 3a_Manual_Ramp
                ws.cell(row=row, column=col_idx["BodyWgtMeasrVal"],
                        value=f"=IFERROR(INDEX('3a_Manual_Ramp'!$D:$D,MATCH($A{row},'3a_Manual_Ramp'!$A:$A,0)),\"\")")
            else:
                # From 0a_Metadata for ENCR
                ws.cell(row=row, column=col_idx["BodyWgtMeasrVal"],
                        value=f"=IFERROR(INDEX('0a_Metadata'!$E:$E,MATCH($A{row},'0a_Metadata'!$A:$A,0)),\"\")")
            
            # SexTyp - from 0a_Metadata
            ws.cell(row=row, column=col_idx["SexTyp"],
                    value=f"=IFERROR(INDEX('0a_Metadata'!$D:$D,MATCH($A{row},'0a_Metadata'!$A:$A,0)),\"\")")
            
            # Exclusion fields - based on 4_Contusion.Survived
            ws.cell(row=row, column=col_idx["Exclusion_in_origin_study"],
                    value=f"=IFERROR(IF(INDEX('4_Contusion_Injury_Details'!$T:$T,MATCH($A{row},'4_Contusion_Injury_Details'!$A:$A,0))=\"N\",\"Total exclusion\",\"No\"),\"\")")
            
            excl_col = get_column_letter(col_idx['Exclusion_in_origin_study'])
            ws.cell(row=row, column=col_idx["Exclusion_reason"],
                    value=f"=IF({excl_col}{row}=\"Total exclusion\",\"Died during/after surgery\",\"\")")
            
            ws.cell(row=row, column=col_idx["Cause_of_Death"],
                    value=f"=IF({excl_col}{row}=\"Total exclusion\",\"Died during surgery\",\"Perfusion\")")
            
            # Injury fields from 4_Contusion
            ws.cell(row=row, column=col_idx["Injury_type"],
                    value=f"=IFERROR(INDEX('4_Contusion_Injury_Details'!$C:$C,MATCH($A{row},'4_Contusion_Injury_Details'!$A:$A,0)),\"\")")
            
            ws.cell(row=row, column=col_idx["Injury_level"],
                    value=f"=IFERROR(INDEX('4_Contusion_Injury_Details'!$E:$E,MATCH($A{row},'4_Contusion_Injury_Details'!$A:$A,0)),\"\")")
            
            # Injury_details - concatenate kd, displacement
            ws.cell(row=row, column=col_idx["Injury_details"],
                    value=f"=IFERROR(INDEX('4_Contusion_Injury_Details'!$P:$P,MATCH($A{row},'4_Contusion_Injury_Details'!$A:$A,0))&\"kd, \"&INDEX('4_Contusion_Injury_Details'!$Q:$Q,MATCH($A{row},'4_Contusion_Injury_Details'!$A:$A,0))&\"um disp\",\"\")")
            
            # === SECTION 2: Contusion Surgery Details (Surgery_1_*) ===
            contusion_cols = [
                ("Surgery_1_Date", "B"), ("Surgery_1_Type", "C"), ("Surgery_1_Severity", "D"),
                ("Surgery_1_Location", "E"), ("Surgery_1_Weight_g", "F"), ("Surgery_1_Anesthetic", "G"),
                ("Surgery_1_Anesthetic_Dose", "H"), ("Surgery_1_Anesthetic_Volume", "I"),
                ("Surgery_1_Analgesic", "J"), ("Surgery_1_Analgesic_Dose", "K"),
                ("Surgery_1_Analgesic_Volume", "L"), ("Surgery_1_Intended_kd", "M"),
                ("Surgery_1_Intended_Dwell", "N"), ("Surgery_1_Stage_Height", "O"),
                ("Surgery_1_Actual_kd", "P"), ("Surgery_1_Actual_Displacement", "Q"),
                ("Surgery_1_Actual_Velocity", "R"), ("Surgery_1_Actual_Dwell", "S"),
                ("Surgery_1_Survived", "T")
            ]
            for odc_col, src_col in contusion_cols:
                ws.cell(row=row, column=col_idx[odc_col],
                        value=f"=IFERROR(INDEX('4_Contusion_Injury_Details'!${src_col}:${src_col},MATCH($A{row},'4_Contusion_Injury_Details'!$A:$A,0)),\"\")")
            
            # === SECTION 3: SC Injection Details (Surgery_2_*) ===
            # Base injection columns
            base_injection_cols = [
                ("Surgery_2_Date", "B"), ("Surgery_2_Weight_g", "C"), ("Surgery_2_Type", "D"),
                ("Surgery_2_Location", "E"), ("Surgery_2_Depth_DV", "F"), ("Surgery_2_Coord_ML", "G"),
            ]
            for odc_col, src_col in base_injection_cols:
                ws.cell(row=row, column=col_idx[odc_col],
                        value=f"=IFERROR(INDEX('5_SC_Injection_Details'!${src_col}:${src_col},MATCH($A{row},'5_SC_Injection_Details'!$A:$A,0)),\"\")")
            
            # Multi-virus columns - dynamically map based on 5_SC structure
            # 5_SC has: Subject_ID(A), Surgery_Date(B), Weight(C), Type(D), Location(E), Depth(F), Coord(G),
            #           Virus_1_Name(H), Virus_1_Titer(I), Virus_1_Source(J), Virus_1_Target(K), ...
            virus_start_col = 8  # Column H is Virus_1_Name
            for v in range(1, max_viruses + 1):
                base_col = virus_start_col + (v - 1) * 4
                virus_cols = [
                    (f"Surgery_2_Virus_{v}_Name", get_column_letter(base_col)),
                    (f"Surgery_2_Virus_{v}_Titer", get_column_letter(base_col + 1)),
                    (f"Surgery_2_Virus_{v}_Source", get_column_letter(base_col + 2)),
                    (f"Surgery_2_Virus_{v}_Target", get_column_letter(base_col + 3)),
                ]
                for odc_col, src_col in virus_cols:
                    ws.cell(row=row, column=col_idx[odc_col],
                            value=f"=IFERROR(INDEX('5_SC_Injection_Details'!${src_col}:${src_col},MATCH($A{row},'5_SC_Injection_Details'!$A:$A,0)),\"\")")
            
            # End injection columns (after virus columns)
            # These start at column H + max_viruses*4
            end_start_col = virus_start_col + max_viruses * 4
            end_injection_cols = [
                ("Surgery_2_Anesthetic", get_column_letter(end_start_col)),
                ("Surgery_2_Anesthetic_Dose", get_column_letter(end_start_col + 1)),
                ("Surgery_2_Anesthetic_Volume", get_column_letter(end_start_col + 2)),
                ("Surgery_2_Analgesic", get_column_letter(end_start_col + 3)),
                ("Surgery_2_Analgesic_Dose", get_column_letter(end_start_col + 4)),
                ("Surgery_2_Analgesic_Volume", get_column_letter(end_start_col + 5)),
                ("Surgery_2_Survived", get_column_letter(end_start_col + 6)),
                ("Surgery_2_Signal_Post_Perfusion", get_column_letter(end_start_col + 7)),
            ]
            for odc_col, src_col in end_injection_cols:
                ws.cell(row=row, column=col_idx[odc_col],
                        value=f"=IFERROR(INDEX('5_SC_Injection_Details'!${src_col}:${src_col},MATCH($A{row},'5_SC_Injection_Details'!$A:$A,0)),\"\")")
            
            # === SECTION 4: BrainGlobe ===
            bg_cols = [
                ("Perfusion_Date", "B"), ("BrainGlobe_Analysis_Date", "C"),
                ("BrainGlobe_Atlas_Used", "D"), ("Total_Cells_Detected", "E"),
                ("Total_Cells_Left_Hemisphere", "F"), ("Total_Cells_Right_Hemisphere", "G"),
                ("BrainGlobe_Notes", "I"), ("BrainGlobe_Quality", "H")
            ]
            for odc_col, src_col in bg_cols:
                ws.cell(row=row, column=col_idx[odc_col],
                        value=f"=IFERROR(INDEX('8_BrainGlobe'!${src_col}:${src_col},MATCH($A{row},'8_BrainGlobe'!$A:$A,0)),\"\")")
            
            # === SECTION 5: Row-Level Metadata ===
            ws.cell(row=row, column=col_idx["Date"], value=date)
            
            if project_type == 'cnt':
                # CNT mode: Full behavior metadata
                ws.cell(row=row, column=col_idx["Test_Phase"], value=phase)
                ws.cell(row=row, column=col_idx["Tray_Type"], value=tray_type)
                ws.cell(row=row, column=col_idx["Num_Trays"], value=trays_per_day)
                
                # Days_Post_Injury = Date - Surgery_1_Date
                surgery1_date_col = get_column_letter(col_idx['Surgery_1_Date'])
                ws.cell(row=row, column=col_idx["Days_Post_Injury"],
                        value=f"=IFERROR({date_col}{row}-{surgery1_date_col}{row},\"\")")
                
                # Weight - average from 3b for this animal+date
                ws.cell(row=row, column=col_idx["Weight"],
                        value=f"=IFERROR(AVERAGEIFS('3b_Manual_Tray'!$D:$D,'3b_Manual_Tray'!$B:$B,$A{row},'3b_Manual_Tray'!$A:$A,{date_col}{row}),\"\")")
                
                # === SECTION 6: Per-Pellet Manual Scores ===
                # For each tray 1-4, pull pellet scores from 3b matching animal+date+tray
                for tray_num in range(1, 5):
                    tray_label = f"{tray_type}{tray_num}"
                    for pellet in range(1, 21):
                        odc_col_name = f"Tray{tray_num}_Pellet{pellet:02d}"
                        pellet_col_3b = PELLET_COLS_3B[pellet]
                        
                        # INDEX/MATCH to find row where Animal+Date+Tray match, then get pellet value
                        # Using SUMPRODUCT to find matching row, then INDEX to get value
                        ws.cell(row=row, column=col_idx[odc_col_name],
                                value=f"=IFERROR(INDEX('3b_Manual_Tray'!${pellet_col_3b}:${pellet_col_3b},MATCH(1,('3b_Manual_Tray'!$B:$B=$A{row})*('3b_Manual_Tray'!$A:$A={date_col}{row})*('3b_Manual_Tray'!$G:$G=\"{tray_label}\"),0)),\"\")")
                
                # === SECTION 7: Per-Tray Manual Calculations ===
                for tray_num in range(1, 5):
                    # Get first and last pellet columns for this tray
                    first_pellet = col_idx[f"Tray{tray_num}_Pellet01"]
                    last_pellet = col_idx[f"Tray{tray_num}_Pellet20"]
                    first_letter = get_column_letter(first_pellet)
                    last_letter = get_column_letter(last_pellet)
                    pellet_range = f"{first_letter}{row}:{last_letter}{row}"
                    
                    # Presented = count of non-blank
                    ws.cell(row=row, column=col_idx[f"Tray{tray_num}_Presented"],
                            value=f"=COUNTA({pellet_range})")
                    
                    # Missed = count of 0s
                    ws.cell(row=row, column=col_idx[f"Tray{tray_num}_Missed"],
                            value=f"=COUNTIF({pellet_range},0)")
                    
                    # Displaced = count of 1s
                    ws.cell(row=row, column=col_idx[f"Tray{tray_num}_Displaced"],
                            value=f"=COUNTIF({pellet_range},1)")
                    
                    # Retrieved = count of 2s
                    ws.cell(row=row, column=col_idx[f"Tray{tray_num}_Retrieved"],
                            value=f"=COUNTIF({pellet_range},2)")
                    
                    # Contacted = Displaced + Retrieved
                    disp_col = get_column_letter(col_idx[f"Tray{tray_num}_Displaced"])
                    ret_col = get_column_letter(col_idx[f"Tray{tray_num}_Retrieved"])
                    ws.cell(row=row, column=col_idx[f"Tray{tray_num}_Contacted"],
                            value=f"={disp_col}{row}+{ret_col}{row}")
                    
                    # Percentages
                    pres_col = get_column_letter(col_idx[f"Tray{tray_num}_Presented"])
                    miss_col = get_column_letter(col_idx[f"Tray{tray_num}_Missed"])
                    cont_col = get_column_letter(col_idx[f"Tray{tray_num}_Contacted"])
                
                ws.cell(row=row, column=col_idx[f"Tray{tray_num}_Miss_Pct"],
                        value=f"=IFERROR({miss_col}{row}/{pres_col}{row}*100,\"\")")
                ws.cell(row=row, column=col_idx[f"Tray{tray_num}_Displaced_Pct"],
                        value=f"=IFERROR({disp_col}{row}/{pres_col}{row}*100,\"\")")
                ws.cell(row=row, column=col_idx[f"Tray{tray_num}_Retrieved_Pct"],
                        value=f"=IFERROR({ret_col}{row}/{pres_col}{row}*100,\"\")")
                ws.cell(row=row, column=col_idx[f"Tray{tray_num}_Contacted_Pct"],
                        value=f"=IFERROR({cont_col}{row}/{pres_col}{row}*100,\"\")")
            
                # === SECTION 8: Daily Totals/Averages ===
                # Total columns - sum of tray columns
                for stat in ["Presented", "Missed", "Displaced", "Retrieved", "Contacted"]:
                    refs = [f"{get_column_letter(col_idx[f'Tray{t}_{stat}'])}{row}" for t in range(1, 5)]
                    ws.cell(row=row, column=col_idx[f"Total_{stat}"],
                            value=f"=SUM({','.join(refs)})")
                
                # Average percentages - average of tray percentages
                for stat in ["Miss_Pct", "Displaced_Pct", "Retrieved_Pct", "Contacted_Pct"]:
                    refs = [f"{get_column_letter(col_idx[f'Tray{t}_{stat}'])}{row}" for t in range(1, 5)]
                    ws.cell(row=row, column=col_idx[f"Avg_{stat}"],
                            value=f"=IFERROR(AVERAGE({','.join(refs)}),\"\")")
                
                # Max/Min
                ret_pct_refs = [f"{get_column_letter(col_idx[f'Tray{t}_Retrieved_Pct'])}{row}" for t in range(1, 5)]
                cont_pct_refs = [f"{get_column_letter(col_idx[f'Tray{t}_Contacted_Pct'])}{row}" for t in range(1, 5)]
                
                ws.cell(row=row, column=col_idx["Max_Retrieved_Pct"],
                        value=f"=IFERROR(MAX({','.join(ret_pct_refs)}),\"\")")
                ws.cell(row=row, column=col_idx["Max_Contacted_Pct"],
                        value=f"=IFERROR(MAX({','.join(cont_pct_refs)}),\"\")")
                ws.cell(row=row, column=col_idx["Min_Retrieved_Pct"],
                        value=f"=IFERROR(MIN({','.join(ret_pct_refs)}),\"\")")
                ws.cell(row=row, column=col_idx["Min_Contacted_Pct"],
                        value=f"=IFERROR(MIN({','.join(cont_pct_refs)}),\"\")")
                
                # === SECTION 9: Kinematic Attention Scores (from 9_DLC) ===
                for attn in ["Tray1_Attention_Score", "Tray2_Attention_Score", 
                            "Tray3_Attention_Score", "Tray4_Attention_Score",
                            "Total_Day_Attention_Score"]:
                    # Find column letter in 9_DLC - attention scores are cols 3-7
                    dlc_col_map = {
                        "Tray1_Attention_Score": "C",
                        "Tray2_Attention_Score": "D",
                        "Tray3_Attention_Score": "E",
                        "Tray4_Attention_Score": "F",
                        "Total_Day_Attention_Score": "G"
                    }
                    ws.cell(row=row, column=col_idx[attn],
                            value=f"=IFERROR(INDEX('9_DLC_Kinematics'!${dlc_col_map[attn]}:${dlc_col_map[attn]},MATCH(1,('9_DLC_Kinematics'!$A:$A=$A{row})*('9_DLC_Kinematics'!$B:$B={date_col}{row}),0)),\"\")")
                
                # === SECTION 10 & 11: Per-Pellet Kinematics and Averages (from 9_DLC) ===
                # These pull from 9_DLC_Kinematics matching Animal+Date
                # 9_DLC has same per-day structure so row should match
                # Use INDEX/MATCH by Animal+Date to find the row, then pull each column
                
                for tray in range(1, 5):
                    for pellet in range(1, 21):
                        for metric in KINEMATIC_METRICS:
                            odc_col_name = f"Tray{tray}_Pellet{pellet:02d}_{metric}"
                            # Find DLC column by counting: 2 ID cols + 5 attention + (tray-1)*20*8 + (pellet-1)*8 + metric_index
                            metric_idx = KINEMATIC_METRICS.index(metric)
                            dlc_col_num = 2 + 5 + (tray-1)*20*8 + (pellet-1)*8 + metric_idx + 1
                            dlc_col_letter = get_column_letter(dlc_col_num)
                            
                            ws.cell(row=row, column=col_idx[odc_col_name],
                                    value=f"=IFERROR(INDEX('9_DLC_Kinematics'!${dlc_col_letter}:${dlc_col_letter},MATCH(1,('9_DLC_Kinematics'!$A:$A=$A{row})*('9_DLC_Kinematics'!$B:$B={date_col}{row}),0)),\"\")")
                
                # Kinematic averages from 9_DLC
                # Tray averages start at column: 2 + 5 + 640 + 1 = 648
                for tray in range(1, 5):
                    for metric in KINEMATIC_METRICS:
                        odc_col_name = f"Tray{tray}_Avg_{metric}"
                        metric_idx = KINEMATIC_METRICS.index(metric)
                        dlc_col_num = 2 + 5 + 640 + (tray-1)*8 + metric_idx + 1
                        dlc_col_letter = get_column_letter(dlc_col_num)
                        
                        ws.cell(row=row, column=col_idx[odc_col_name],
                                value=f"=IFERROR(INDEX('9_DLC_Kinematics'!${dlc_col_letter}:${dlc_col_letter},MATCH(1,('9_DLC_Kinematics'!$A:$A=$A{row})*('9_DLC_Kinematics'!$B:$B={date_col}{row}),0)),\"\")")
                
                # Day averages start at column: 2 + 5 + 640 + 32 + 1 = 680
                for metric in KINEMATIC_METRICS:
                    odc_col_name = f"Day_Avg_{metric}"
                    metric_idx = KINEMATIC_METRICS.index(metric)
                    dlc_col_num = 2 + 5 + 640 + 32 + metric_idx + 1
                    dlc_col_letter = get_column_letter(dlc_col_num)
                    
                    ws.cell(row=row, column=col_idx[odc_col_name],
                            value=f"=IFERROR(INDEX('9_DLC_Kinematics'!${dlc_col_letter}:${dlc_col_letter},MATCH(1,('9_DLC_Kinematics'!$A:$A=$A{row})*('9_DLC_Kinematics'!$B:$B={date_col}{row}),0)),\"\")")
                
                # Source sheet for CNT
                ws.cell(row=row, column=col_idx["Source_Sheet"], value="3b_Manual_Tray")
            
            else:
                # ENCR mode: Minimal metadata
                surgery1_date_col = get_column_letter(col_idx['Surgery_1_Date'])
                ws.cell(row=row, column=col_idx["Days_Post_Surgery"],
                        value=f"=IFERROR({date_col}{row}-{surgery1_date_col}{row},\"\")")
                
                # Source sheet for ENCR
                ws.cell(row=row, column=col_idx["Source_Sheet"], value="5_SC_Injection_Details")
            
            # === SECTION 12: Source Tracking (both modes) ===
            ws.cell(row=row, column=col_idx["Source_File"], value=f"=CELL(\"filename\")")
            
            row_num += 1
    
    # Autoscale columns (limited for performance on 800+ columns)
    # Just do first 100 columns
    for col_idx_num in range(1, min(101, len(columns) + 1)):
        header_cell = ws.cell(row=1, column=col_idx_num)
        header_value = header_cell.value
        if header_value:
            width = min(20, max(8, len(str(header_value)) * 1.1))
            ws.column_dimensions[get_column_letter(col_idx_num)].width = width


# =============================================================================
# MAIN FILE GENERATOR
# =============================================================================

def create_new_cohort_file(cohort_name, start_date, num_mice, output_dir=None, 
                           project_type='cnt', max_viruses=3):
    """
    Create a complete new cohort Excel file with all sheets
    
    Args:
        cohort_name: e.g., "CNT_05", "ENCR_01", or just "05"
        start_date: Food deprivation start date (datetime or string YYYY-MM-DD or YYYYMMDD)
        num_mice: Number of mice in cohort (default 16)
        output_dir: Output directory (default: 'generated' subdirectory)
        project_type: 'cnt' (connectome/behavior) or 'encr' (enhancer/brainglobe)
        max_viruses: Maximum viruses per injection (default 3)
    
    Returns:
        Path to created file
    """
    # Parse start_date if string
    if isinstance(start_date, str):
        start_date = parse_date(start_date)
    
    # Generate subject IDs
    subject_ids = generate_subject_ids(cohort_name, num_mice)
    
    # Determine output filename based on project type
    if "_" in cohort_name:
        cohort_num = cohort_name.split("_")[-1]
    else:
        cohort_num = cohort_name
    
    if project_type == 'encr':
        filename = f"Enhancer_{cohort_num}_Animal_Tracking.xlsx"
    else:
        filename = f"Connectome_{cohort_num}_Animal_Tracking.xlsx"
    
    # Use output subdirectory by default
    if output_dir is None:
        output_dir = OUTPUT_SUBDIR
    else:
        output_dir = Path(output_dir)
    
    # Create output directory if it doesn't exist
    output_dir.mkdir(parents=True, exist_ok=True)
    
    output_path = output_dir / filename
    
    print(f"Creating new cohort file: {output_path}")
    print(f"  Cohort: {cohort_name}")
    print(f"  Project type: {project_type.upper()}")
    print(f"  Start date: {start_date.strftime('%Y-%m-%d')}")
    print(f"  Number of mice: {num_mice}")
    print(f"  Max viruses: {max_viruses}")
    print(f"  Subject IDs: {subject_ids[0]} to {subject_ids[-1]}")
    
    # Create all sheets
    sheets = {}
    
    print("\nGenerating sheets...")
    
    print("  - 0a_Metadata")
    sheets["0a_Metadata"] = create_0a_metadata(subject_ids)
    
    print("  - 0_Virus_Preparation")
    sheets["0_Virus_Preparation"] = create_0_virus_preparation(cohort_name, num_groups=2, viruses_per_group=3)
    
    print("  - 0_Injection_Calculations")
    sheets["0_Injection_Calculations"] = create_0_injection_calculations()
    
    print("  - 1_Experiment_Planning")
    sheets["1_Experiment_Planning"] = create_1_experiment_planning(
        start_date, subject_ids, cohort_name=cohort_name, script_dir=SCRIPT_DIR
    )
    
    print("  - 2_ODC_Animal_Tracking")
    sheets["2_ODC_Animal_Tracking"] = create_2_odc_animal_tracking(
        subject_ids, start_date, project_type=project_type, max_viruses=max_viruses
    )
    
    # Behavior sheets - CNT only
    if project_type == 'cnt':
        print("  - 3a_Manual_Ramp")
        sheets["3a_Manual_Ramp"] = create_3a_manual_ramp(start_date, subject_ids)
        
        print("  - 3b_Manual_Tray")
        sheets["3b_Manual_Tray"] = create_3b_manual_tray(start_date, subject_ids)
        
        print("  - 3c_Manual_Summary")
        sheets["3c_Manual_Summary"] = create_3c_manual_summary(start_date, subject_ids)
        
        print("  - 3d_Weights")
        sheets["3d_Weights"] = create_3d_weights(start_date, subject_ids)
    
    print("  - 4_Contusion_Injury_Details")
    sheets["4_Contusion_Injury_Details"] = create_4_contusion_injury_details(subject_ids, start_date)
    
    print("  - 5_SC_Injection_Details")
    sheets["5_SC_Injection_Details"] = create_5_sc_injection_details(
        subject_ids, start_date, max_viruses=max_viruses
    )
    
    # CNT only sheets
    if project_type == 'cnt':
        print("  - 6_Ladder")
        sheets["6_Ladder"] = create_6_ladder(subject_ids)
        
        print("  - 7_Stats")
        sheets["7_Stats"] = create_7_stats(subject_ids)
    
    print("  - 8_BrainGlobe")
    sheets["8_BrainGlobe"] = create_8_brainglobe()
    
    # CNT only
    if project_type == 'cnt':
        print("  - 9_DLC_Kinematics")
        sheets["9_DLC_Kinematics"] = create_9_dlc_kinematics(subject_ids, start_date)
    
    # Write to Excel using openpyxl for formula support
    print(f"\nWriting to {output_path}...")
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Write each sheet
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        
        if sheet_name == "0_Virus_Preparation":
            # Virus preparation planning sheet
            if isinstance(df, dict) and df.get('_is_virus_prep'):
                write_0_virus_preparation(ws, df)
            else:
                # Fallback - shouldn't happen
                pass
        elif sheet_name == "0_Injection_Calculations":
            # Virus prep with calculation formulas
            write_0_injection_calculations_with_formulas(ws, df)
        elif sheet_name == "1_Experiment_Planning":
            # Gantt-style multi-cohort planning
            if isinstance(df, dict) and df.get('_is_gantt'):
                write_1_experiment_planning_gantt(ws, df)
            else:
                # Fallback to standard write
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
        elif sheet_name == "2_ODC_Animal_Tracking":
            # ODC output with formulas pulling from all sheets
            write_2_odc_with_formulas(ws, df, cohort_name)
        elif sheet_name == "3a_Manual_Ramp":
            # Wide format ramp sheet with formulas
            write_3a_with_formulas(ws, df)
        elif sheet_name == "3b_Manual_Tray":
            # Special handling with formulas
            write_3b_with_formulas(ws, df, num_mice)
        elif sheet_name == "3c_Manual_Summary":
            # Wide format summary with formulas pulling from 3b
            write_3c_with_formulas(ws, df)
        elif sheet_name == "3d_Weights":
            # Special flip-chart format
            write_3d_weights_flip(ws, df, start_date, subject_ids)
        elif sheet_name == "4_Contusion_Injury_Details":
            # Surgery sheet with volume formulas
            write_4_contusion_with_formulas(ws, df)
        elif sheet_name == "5_SC_Injection_Details":
            # Surgery sheet with volume formulas and virus references
            write_5_sc_injection_with_formulas(ws, df)
        elif sheet_name == "7_Stats":
            # Stats with formulas referencing 3c
            write_7_stats_with_formulas(ws, df)
        elif sheet_name == "9_DLC_Kinematics":
            # DLC kinematics with formulas for averages
            write_9_dlc_with_formulas(ws, df)
        else:
            # Standard dataframe write
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Autoscale columns to fit header text
        autoscale_columns(ws)
    
    # Save with retry logic in case file is open in Excel
    max_retries = 3
    for attempt in range(max_retries):
        try:
            wb.save(output_path)
            break
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"\n[!] Cannot write to {output_path.name} - file may be open in Excel.")
                print("  Attempting to close it...")
                
                # Try to close the file (Windows-specific)
                try:
                    import subprocess
                    import platform
                    
                    if platform.system() == 'Windows':
                        # Use taskkill to close Excel if it has the file open
                        # This is aggressive but user requested it
                        subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'], 
                                      capture_output=True, timeout=5)
                        print("  Closed Excel. Retrying save...")
                        import time
                        time.sleep(1)  # Give it a moment
                    else:
                        # On Mac/Linux, just inform user
                        print("  Please close the file manually and press Enter...")
                        input()
                except Exception as e:
                    print(f"  Could not auto-close: {e}")
                    print("  Please close the file manually and press Enter...")
                    input()
            else:
                print(f"\n[X] ERROR: Cannot save to {output_path}")
                print("  Please close the file in Excel and run the script again.")
                raise
    
    print(f"\n[OK] Successfully created {output_path}")
    print(f"  Total sheets: {len(sheets)}")
    
    # Summary of key counts
    if "3b_Manual_Tray" in sheets:
        tray_sheet = sheets["3b_Manual_Tray"]
        print(f"  Rows in 3b_Manual_Tray: {len(tray_sheet)}")
    # ODC is now a dict, count rows in generated sheet
    odc_ws = wb['2_ODC_Animal_Tracking']
    print(f"  Rows in 2_ODC_Animal_Tracking: {odc_ws.max_row}")
    
    return output_path


# =============================================================================
# INTERACTIVE INTERFACE
# =============================================================================

def print_header():
    """Print the script header"""
    print()
    print("=" * 60)
    print("  ODC-SCI Cohort File Manager")
    print("  Script 0: Create or Fix Cohort Files")
    print("=" * 60)
    print()


def get_user_choice(prompt, valid_options, default=None):
    """
    Get a choice from the user with validation
    
    Args:
        prompt: The prompt to display
        valid_options: List of valid options
        default: Default value if user presses Enter
    
    Returns:
        The user's validated choice
    """
    while True:
        if default:
            user_input = input(f"{prompt} [default: {default}]: ").strip()
            if user_input == "":
                return default
        else:
            user_input = input(f"{prompt}: ").strip()
        
        if user_input in valid_options:
            return user_input
        else:
            print(f"  Invalid choice. Please enter one of: {', '.join(valid_options)}")


def get_user_input(prompt, default=None, validator=None, error_msg=None):
    """
    Get input from the user with optional validation
    
    Args:
        prompt: The prompt to display
        default: Default value if user presses Enter
        validator: Optional function that returns True if input is valid
        error_msg: Error message to display if validation fails
    
    Returns:
        The user's validated input
    """
    while True:
        if default is not None:
            user_input = input(f"{prompt} [default: {default}]: ").strip()
            if user_input == "":
                return default
        else:
            user_input = input(f"{prompt}: ").strip()
            if user_input == "":
                print("  This field is required. Please enter a value.")
                continue
        
        if validator:
            if validator(user_input):
                return user_input
            else:
                print(f"  {error_msg or 'Invalid input. Please try again.'}")
        else:
            return user_input


def validate_date(date_str):
    """Validate date format YYYY-MM-DD or YYYYMMDD"""
    # Try YYYY-MM-DD format
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        pass
    
    # Try YYYYMMDD format
    try:
        datetime.strptime(date_str, "%Y%m%d")
        return True
    except ValueError:
        return False


def parse_date(date_str):
    """Parse date string in either YYYY-MM-DD or YYYYMMDD format"""
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return datetime.strptime(date_str, "%Y%m%d")


def validate_positive_int(val_str):
    """Validate positive integer"""
    try:
        val = int(val_str)
        return val > 0
    except ValueError:
        return False


def interactive_new_cohort():
    """Interactive mode for creating a new cohort file"""
    print("\n--- Creating New Cohort File ---\n")
    
    # Show where files will be saved
    print(f"Files will be saved to: {OUTPUT_SUBDIR}")
    print()
    
    # Get cohort name
    print("The cohort name will be used to generate SubjectIDs.")
    print("Example: 'CNT_05' will create IDs like CNT_05_01, CNT_05_02, etc.")
    print("         'ENCR_01' will create IDs like ENCR_01_01, ENCR_01_02, etc.")
    cohort_name = get_user_input(
        "Enter cohort name (e.g., CNT_05, ENCR_01, or just 05)"
    )
    
    # Detect project type from cohort name
    project_type = detect_project_type(cohort_name=cohort_name)
    print(f"\n  Detected project type: {project_type.upper()}")
    if project_type == 'encr':
        print("  (ENCR mode: No behavior sheets, focus on BrainGlobe)")
    else:
        print("  (CNT mode: Full behavior tracking + tracing)")
    
    # Get start date - different prompt based on project type
    if project_type == 'encr':
        print("\nThe start date is the surgery/injection date.")
        print("(ENCR projects don't have food deprivation or behavior training)")
        start_date = get_user_input(
            "Enter surgery start date (YYYY-MM-DD or YYYYMMDD)",
            validator=validate_date,
            error_msg="Invalid date format. Please use YYYY-MM-DD or YYYYMMDD (e.g., 2025-02-01 or 20250201)"
        )
    else:
        print("\nThe start date is Day 0 (food deprivation begins).")
        print("All other dates will be calculated from this.")
        start_date = get_user_input(
            "Enter food deprivation start date (YYYY-MM-DD or YYYYMMDD)",
            validator=validate_date,
            error_msg="Invalid date format. Please use YYYY-MM-DD or YYYYMMDD (e.g., 2025-02-01 or 20250201)"
        )
    
    # Get number of mice
    print("\nHow many mice in this cohort?")
    mice_str = get_user_input(
        "Enter number of mice",
        default="16" if project_type == 'cnt' else "8",
        validator=validate_positive_int,
        error_msg="Please enter a positive number."
    )
    num_mice = int(mice_str)
    
    # Get max viruses for ENCR
    max_viruses = 3
    if project_type == 'encr':
        print("\nHow many viruses per injection?")
        viruses_str = get_user_input(
            "Enter max viruses per injection",
            default="4",
            validator=validate_positive_int,
            error_msg="Please enter a positive number."
        )
        max_viruses = int(viruses_str)
    
    # Get output directory
    print("\nWhere should the file be saved?")
    print("Press Enter for current directory, or enter a path.")
    output_dir = get_user_input(
        "Enter output directory",
        default="."
    )
    if output_dir == ".":
        output_dir = None
    
    # Confirm
    print("\n" + "-" * 40)
    print("Please confirm:")
    print(f"  Cohort name: {cohort_name}")
    print(f"  Project type: {project_type.upper()}")
    print(f"  Start date: {start_date}")
    print(f"  Number of mice: {num_mice}")
    if project_type == 'encr':
        print(f"  Max viruses: {max_viruses}")
    print(f"  Output directory: {output_dir or 'current directory'}")
    print("-" * 40)
    
    confirm = get_user_choice(
        "Proceed? (y/n)",
        valid_options=["y", "n", "Y", "N", "yes", "no"],
        default="y"
    ).lower()
    
    if confirm in ["y", "yes"]:
        print()
        return create_new_cohort_file(cohort_name, start_date, num_mice, output_dir,
                                       project_type=project_type, max_viruses=max_viruses)
    else:
        print("\nCancelled. No file created.")
        return None


def interactive_fix_existing():
    """Interactive mode for fixing an existing cohort file"""
    print("\n--- Fix Existing Cohort File ---\n")
    
    # Find Excel files in script directory
    excel_files = list(SCRIPT_DIR.glob("*.xlsx")) + list(SCRIPT_DIR.glob("*.xls"))
    # Exclude files in generated subdirectory
    excel_files = [f for f in excel_files if "generated" not in str(f)]
    
    if not excel_files:
        print("No Excel files found in the script directory.")
        print(f"Looking in: {SCRIPT_DIR}")
        print("\nPlease place your cohort file in the same folder as this script.")
        return None
    
    print("Found Excel files:")
    for i, f in enumerate(excel_files, 1):
        print(f"  {i}. {f.name}")
    print()
    
    # Get user selection
    while True:
        choice = input(f"Select file to fix (1-{len(excel_files)}) or 'q' to quit: ").strip()
        if choice.lower() == 'q':
            return None
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(excel_files):
                selected_file = excel_files[idx]
                break
            else:
                print(f"  Please enter a number between 1 and {len(excel_files)}")
        except ValueError:
            print("  Please enter a valid number")
    
    print(f"\nSelected: {selected_file.name}")
    
    # Run the fixer
    result = fix_existing_file(selected_file)
    
    if result:
        print("\n" + "=" * 60)
        print("Done! Fixed file created.")
        print(f"Location: {result['output_file']}")
        print(f"Report: {result['report_file']}")
        print("=" * 60)
    
    return result


# =============================================================================
# FIXER/CHECKER FUNCTIONALITY
# =============================================================================

# Expected sheets in order
EXPECTED_SHEETS = [
    "0a_Metadata",
    "0_Virus_Preparation",
    "0_Injection_Calculations",
    "1_Experiment_Planning",
    "2_ODC_Animal_Tracking",
    "3a_Manual_Ramp",
    "3b_Manual_Tray",
    "3c_Manual_Summary",
    "3d_Weights",
    "4_Contusion_Injury_Details",
    "5_SC_Injection_Details",
    "6_Ladder",
    "7_Stats",
    "8_BrainGlobe",
    "9_DLC_Kinematics"
]

# Column name mappings (old name -> standard name)
COLUMN_NAME_MAPPINGS = {
    # Animal ID variations
    "Mouse": "Animal",
    "mouse": "Animal",
    "Subject": "Animal",
    "Subject_ID": "Animal",
    "SubjectID": "Animal",
    "Animal_ID": "Animal",
    "animal": "Animal",
    "ID": "Animal",
    
    # Date variations
    "Test_Date": "Date",
    "Testing_Date": "Date",
    "date": "Date",
    "DATE": "Date",
    
    # Sex variations
    "sex": "Sex",
    "SEX": "Sex",
    "Gender": "Sex",
    
    # Weight variations
    "weight": "Weight",
    "WEIGHT": "Weight",
    "Body_Weight": "Weight",
    "BW": "Weight",
    
    # Tray variations
    "Tray": "Tray Type/Number",
    "Tray_Type": "Tray Type/Number",
    "TrayType": "Tray Type/Number",
    "Tray_Number": "Tray Type/Number",
}


class FixerReport:
    """Collects and formats report information"""
    
    def __init__(self, input_file):
        self.input_file = input_file
        self.output_file = None
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.issues = []  # (severity, category, message)
        self.fixes = []   # (category, message)
        self.warnings = []
        self.info = []
        
    def add_issue(self, severity, category, message):
        """Add an issue found. Severity: ERROR, WARNING, INFO"""
        self.issues.append((severity, category, message))
        
    def add_fix(self, category, message):
        """Add a fix that was applied"""
        self.fixes.append((category, message))
        
    def add_warning(self, message):
        """Add a warning"""
        self.warnings.append(message)
        
    def add_info(self, message):
        """Add informational message"""
        self.info.append(message)
    
    def generate_report(self):
        """Generate the full report text"""
        lines = []
        lines.append("=" * 70)
        lines.append("ODC-SCI COHORT FILE CHECKER/FIXER REPORT")
        lines.append("=" * 70)
        lines.append(f"Generated: {self.timestamp}")
        lines.append(f"Input file: {self.input_file}")
        lines.append(f"Output file: {self.output_file}")
        lines.append("")
        
        # Summary
        error_count = sum(1 for s, c, m in self.issues if s == "ERROR")
        warning_count = sum(1 for s, c, m in self.issues if s == "WARNING")
        lines.append("-" * 70)
        lines.append("SUMMARY")
        lines.append("-" * 70)
        lines.append(f"Errors found: {error_count}")
        lines.append(f"Warnings: {warning_count}")
        lines.append(f"Fixes applied: {len(self.fixes)}")
        lines.append("")
        
        # Issues by category
        if self.issues:
            lines.append("-" * 70)
            lines.append("ISSUES FOUND")
            lines.append("-" * 70)
            
            # Group by category
            categories = {}
            for severity, category, message in self.issues:
                if category not in categories:
                    categories[category] = []
                categories[category].append((severity, message))
            
            for category, items in categories.items():
                lines.append(f"\n{category}:")
                for severity, message in items:
                    lines.append(f"  [{severity}] {message}")
        
        # Fixes applied
        if self.fixes:
            lines.append("")
            lines.append("-" * 70)
            lines.append("FIXES APPLIED")
            lines.append("-" * 70)
            
            # Group by category
            categories = {}
            for category, message in self.fixes:
                if category not in categories:
                    categories[category] = []
                categories[category].append(message)
            
            for category, items in categories.items():
                lines.append(f"\n{category}:")
                for message in items:
                    lines.append(f"  [OK] {message}")
        
        # Info
        if self.info:
            lines.append("")
            lines.append("-" * 70)
            lines.append("INFORMATION")
            lines.append("-" * 70)
            for msg in self.info:
                lines.append(f"  - {msg}")
        
        lines.append("")
        lines.append("=" * 70)
        lines.append("END OF REPORT")
        lines.append("=" * 70)
        
        return "\n".join(lines)
    
    def print_report(self):
        """Print report to console"""
        print(self.generate_report())
    
    def save_report(self, output_path):
        """Save report to file"""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(self.generate_report())


def fix_existing_file(input_path, output_dir=None, project_type='cnt', max_viruses=3):
    """
    Check and fix an existing cohort file
    
    Args:
        input_path: Path to the input Excel file
        output_dir: Output directory (default: generated/)
        project_type: 'cnt' (connectome/behavior) or 'encr' (enhancer/brainglobe)
        max_viruses: Maximum viruses per injection (default 3)
    
    Returns:
        dict with 'output_file' and 'report_file' paths, or None if failed
    """
    input_path = Path(input_path)
    
    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        return None
    
    # Setup output directory
    if output_dir is None:
        output_dir = OUTPUT_SUBDIR
    else:
        output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Create output filename (same name as input - no _fixed suffix)
    output_path = output_dir / input_path.name
    report_path = output_dir / f"{input_path.stem}_fix_report.txt"
    
    print(f"\nChecking and fixing: {input_path.name}")
    print(f"Output will be saved to: {output_path}")
    print()
    
    # Initialize report
    report = FixerReport(input_path)
    report.output_file = output_path
    
    # Load the workbook
    try:
        wb = openpyxl.load_workbook(input_path)
        report.add_info(f"Loaded workbook with {len(wb.sheetnames)} sheets")
    except Exception as e:
        print(f"Error loading file: {e}")
        return None
    
    # Get existing sheet names
    existing_sheets = wb.sheetnames
    report.add_info(f"Existing sheets: {', '.join(existing_sheets)}")
    
    # Try to infer cohort info from existing data
    cohort_info = infer_cohort_info(wb, report)
    
    # === CHECK AND FIX SHEETS ===
    
    # 1. Check for missing sheets
    print("Checking sheets...")
    missing_sheets = [s for s in EXPECTED_SHEETS if s not in existing_sheets]
    extra_sheets = [s for s in existing_sheets if s not in EXPECTED_SHEETS]
    
    for sheet in missing_sheets:
        report.add_issue("WARNING", "Missing Sheets", f"Sheet '{sheet}' is missing")
    
    for sheet in extra_sheets:
        report.add_issue("INFO", "Extra Sheets", f"Unexpected sheet '{sheet}' found (will be preserved)")
    
    # 2. Add missing sheets
    if missing_sheets:
        print(f"  Adding {len(missing_sheets)} missing sheets...")
        add_missing_sheets(wb, missing_sheets, cohort_info, report, input_path)
    
    # 3. Check and fix each existing sheet
    print("Checking sheet contents...")
    for sheet_name in existing_sheets:
        if sheet_name in EXPECTED_SHEETS:
            check_and_fix_sheet(wb, sheet_name, cohort_info, report)
    
    # 4. Check if ODC needs restructuring (per-tray to per-day)
    if "2_ODC_Animal_Tracking" in existing_sheets:
        check_and_fix_odc_structure(wb, cohort_info, report)
    
    # 5. Reorder sheets to match expected order
    reorder_sheets(wb, report)
    
    # === SAVE RESULTS ===
    
    print("\nSaving fixed file...")
    try:
        wb.save(output_path)
        print(f"  [OK] Saved: {output_path}")
    except Exception as e:
        print(f"  [X] Error saving: {e}")
        return None
    
    # Save report
    report.save_report(report_path)
    print(f"  [OK] Report: {report_path}")
    
    # Print report to console
    print()
    report.print_report()
    
    return {
        'output_file': output_path,
        'report_file': report_path
    }


def infer_cohort_info(wb, report):
    """
    Try to infer cohort information from existing data
    
    Returns dict with:
        - cohort_name: e.g., "CNT_01"
        - subject_ids: list of subject IDs
        - start_date: experiment start date
        - num_mice: number of mice
    """
    info = {
        'cohort_name': None,
        'subject_ids': [],
        'start_date': None,
        'num_mice': 16  # default
    }
    
    # Try to find subject IDs from various sheets
    subject_id_sources = [
        ("3b_Manual_Tray", ["Animal", "Mouse", "Subject", "SubjectID"]),
        ("0a_Metadata", ["SubjectID", "Animal", "Subject"]),
        ("4_Contusion_Injury_Details", ["SubjectID", "Animal", "Subject"]),
    ]
    
    for sheet_name, possible_cols in subject_id_sources:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Find the subject ID column
            for col in range(1, min(20, ws.max_column + 1)):
                header = ws.cell(1, col).value
                if header in possible_cols:
                    # Extract unique values
                    ids = set()
                    for row in range(2, ws.max_row + 1):
                        val = ws.cell(row, col).value
                        if val and str(val).strip():
                            ids.add(str(val).strip())
                    
                    if ids:
                        info['subject_ids'] = sorted(list(ids))
                        report.add_info(f"Found {len(ids)} subject IDs from {sheet_name}")
                        
                        # Infer cohort name from first ID
                        first_id = info['subject_ids'][0]
                        if "_" in first_id:
                            # e.g., "CNT_01_05" -> "CNT_01"
                            parts = first_id.rsplit("_", 1)
                            info['cohort_name'] = parts[0]
                            report.add_info(f"Inferred cohort name: {info['cohort_name']}")
                        
                        info['num_mice'] = len(ids)
                        break
            if info['subject_ids']:
                break
    
    # Try to find start date from 3b_Manual_Tray
    if "3b_Manual_Tray" in wb.sheetnames:
        ws = wb["3b_Manual_Tray"]
        # Find Date column
        for col in range(1, min(10, ws.max_column + 1)):
            header = ws.cell(1, col).value
            if header in ["Date", "date", "DATE", "Test_Date"]:
                # Get minimum date
                dates = []
                for row in range(2, min(100, ws.max_row + 1)):
                    val = ws.cell(row, col).value
                    if val:
                        if isinstance(val, datetime):
                            dates.append(val)
                        elif isinstance(val, str):
                            try:
                                dates.append(datetime.strptime(val, "%Y-%m-%d"))
                            except:
                                pass
                
                if dates:
                    min_date = min(dates)
                    # Start date is typically 4 days before first testing date
                    info['start_date'] = min_date - timedelta(days=4)
                    report.add_info(f"Inferred start date: {info['start_date'].strftime('%Y-%m-%d')}")
                break
    
    # Set defaults if not found
    if not info['cohort_name']:
        info['cohort_name'] = "CNT_XX"
        report.add_warning("Could not infer cohort name, using 'CNT_XX'")
    
    if not info['subject_ids']:
        info['subject_ids'] = generate_subject_ids(info['cohort_name'], info['num_mice'])
        report.add_warning(f"Could not find subject IDs, generating {info['num_mice']} default IDs")
    
    if not info['start_date']:
        info['start_date'] = datetime.now()
        report.add_warning("Could not infer start date, using today")
    
    return info


def add_missing_sheets(wb, missing_sheets, cohort_info, report, input_path=None):
    """Add missing sheets to workbook
    
    Args:
        wb: Workbook object
        missing_sheets: List of sheet names to add
        cohort_info: Dict with subject_ids, start_date, cohort_name
        report: FixerReport object
        input_path: Path to original file (for data_only loading)
    """
    
    subject_ids = cohort_info['subject_ids']
    start_date = cohort_info['start_date']
    cohort_name = cohort_info['cohort_name']
    
    # For ODC, we need to extract actual data from 3b if it exists
    if '3b_Manual_Tray' in wb.sheetnames and '2_ODC_Animal_Tracking' in missing_sheets:
        # We'll handle ODC specially using the discovery functions
        pass  # The write_2_odc_from_existing_data function will do the discovery
    
    for sheet_name in missing_sheets:
        print(f"    Adding {sheet_name}...")
        ws = wb.create_sheet(title=sheet_name)
        
        if sheet_name == "0a_Metadata":
            df = create_0a_metadata(subject_ids)
            write_dataframe_to_sheet(ws, df)
            report.add_fix("Sheets", f"Created {sheet_name} with {len(subject_ids)} subjects")
            
        elif sheet_name == "0_Injection_Calculations":
            df = create_0_injection_calculations()
            write_0_injection_calculations_with_formulas(ws, df)
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "1_Experiment_Planning":
            data = create_1_experiment_planning(start_date, subject_ids, cohort_name, SCRIPT_DIR)
            if isinstance(data, dict) and data.get('_is_gantt'):
                write_1_experiment_planning_gantt(ws, data)
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "2_ODC_Animal_Tracking":
            # Use discovery approach if 3b exists with data
            if '3b_Manual_Tray' in wb.sheetnames:
                write_2_odc_from_existing_data(ws, wb, subject_ids, None, cohort_name, report, input_path)
            else:
                data = create_2_odc_animal_tracking(subject_ids, start_date)
                write_2_odc_with_formulas(ws, data, cohort_name)
            report.add_fix("Sheets", f"Created {sheet_name} (per-day structure)")
            
        elif sheet_name == "3a_Manual_Ramp":
            data = create_3a_manual_ramp(start_date, subject_ids)
            write_3a_with_formulas(ws, data)
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "3b_Manual_Tray":
            df = create_3b_manual_tray(start_date, subject_ids)
            write_3b_with_formulas(ws, df, len(subject_ids))
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "3c_Manual_Summary":
            data = create_3c_manual_summary(start_date, subject_ids)
            write_3c_with_formulas(ws, data)
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "3d_Weights":
            df = create_3d_weights(start_date, subject_ids)
            write_3d_weights_flip(ws, df, start_date, subject_ids)
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "4_Contusion_Injury_Details":
            df = create_4_contusion_injury_details(subject_ids, start_date)
            write_4_contusion_with_formulas(ws, df)
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "5_SC_Injection_Details":
            df = create_5_sc_injection_details(subject_ids, start_date)
            write_5_sc_injection_with_formulas(ws, df)
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "6_Ladder":
            df = create_6_ladder(subject_ids)
            write_dataframe_to_sheet(ws, df)
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "7_Stats":
            data = create_7_stats(subject_ids)
            write_7_stats_with_formulas(ws, data)
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "8_BrainGlobe":
            df = create_8_brainglobe()
            write_dataframe_to_sheet(ws, df)
            report.add_fix("Sheets", f"Created {sheet_name}")
            
        elif sheet_name == "9_DLC_Kinematics":
            data = create_9_dlc_kinematics(subject_ids, start_date)
            write_9_dlc_with_formulas(ws, data)
            report.add_fix("Sheets", f"Created {sheet_name}")


def write_dataframe_to_sheet(ws, df):
    """Helper to write a dataframe to a worksheet"""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    autoscale_columns(ws)


def discover_3b_structure(ws_3b):
    """
    Discover the column structure of 3b_Manual_Tray by finding key columns.
    
    Returns dict with column indices (1-based) for:
        - animal_col: Animal/SubjectID column
        - date_col: Date column
        - tray_col: Tray Type/Number column
        - sex_col: Sex column
        - weight_col: Weight column
        - phase_col: Test_Phase column
        - pellet_start_col: First pellet column (labeled "1")
        - pellet_end_col: Last pellet column (labeled "20")
    
    Only uses the FIRST occurrence of each column type to avoid duplicates.
    """
    structure = {}
    
    # Scan headers to find columns - only take first occurrence
    for col in range(1, min(60, ws_3b.max_column + 1)):
        header = ws_3b.cell(1, col).value
        if header is None:
            continue
        
        header_str = str(header).strip().lower()
        
        # Animal ID column - only set if not already found
        if header_str in ['animal', 'mouse', 'subject', 'subjectid', 'subject_id', 'mouse id', 'animal_id']:
            if 'animal_col' not in structure:
                structure['animal_col'] = col
        
        # Date column - only set if not already found
        elif header_str == 'date':
            if 'date_col' not in structure:
                structure['date_col'] = col
        
        # Tray column - only set if not already found
        elif header_str in ['tray type/number', 'tray', 'tray_type', 'traytype']:
            if 'tray_col' not in structure:
                structure['tray_col'] = col
        
        # Sex column - only set if not already found
        elif header_str == 'sex':
            if 'sex_col' not in structure:
                structure['sex_col'] = col
        
        # Weight column (but not "weight %") - only set if not already found
        elif header_str == 'weight':
            if 'weight_col' not in structure:
                structure['weight_col'] = col
        
        # Test phase column - only set if not already found
        elif header_str in ['test_phase', 'phase', 'testphase']:
            if 'phase_col' not in structure:
                structure['phase_col'] = col
        
        # Pellet columns - look for "1" through "20" - only set if not already found
        elif header_str == '1':
            if 'pellet_start_col' not in structure:
                structure['pellet_start_col'] = col
        elif header_str == '20':
            if 'pellet_end_col' not in structure:
                structure['pellet_end_col'] = col
    
    return structure


def extract_date_phase_mapping(wb_data):
    """
    Extract date-to-phase mapping from 3c_Manual_Summary sheet.
    
    The summary sheet has:
    - Row 1: Dates as column headers
    - Row 2: Phase names corresponding to each date
    
    Returns:
        dict: {date: phase_name} mapping
    """
    if '3c_Manual_Summary' not in wb_data.sheetnames:
        return {}
    
    ws = wb_data['3c_Manual_Summary']
    date_phase_map = {}
    
    # Scan row 1 for dates, row 2 for phases
    for col in range(1, ws.max_column + 1):
        date_val = ws.cell(1, col).value
        phase_val = ws.cell(2, col).value
        
        if date_val is None or phase_val is None:
            continue
        
        # Convert date to date object
        if isinstance(date_val, datetime):
            date_key = date_val.date()
        elif isinstance(date_val, date):
            date_key = date_val
        else:
            continue  # Skip non-date values
        
        # Store the mapping
        date_phase_map[date_key] = str(phase_val).strip()
    
    return date_phase_map


def extract_all_data_from_workbook(wb_data, structure_3b, report):
    """
    Extract all data from existing workbook sheets using Python.
    
    NOTE: wb_data should be loaded with data_only=True to get cached formula values.
    
    Returns a dict with all extracted data organized by animal.
    """
    data = {
        'animals': {},  # animal_id -> {sex, baseline_weight, dob, dod, ...}
        'testing_days': {},  # animal_id -> [(date, phase, tray_type, weight, trays_data), ...]
        'contusion': {},  # animal_id -> {date, type, severity, location, ...}
        'injection': {},  # animal_id -> {date, virus, titer, ...}
    }
    
    # Extract date-to-phase mapping from 3c_Manual_Summary
    date_phase_map = extract_date_phase_mapping(wb_data)
    if date_phase_map:
        report.add_info(f"Extracted {len(date_phase_map)} date-to-phase mappings from 3c_Manual_Summary")
    
    # Extract from 3b_Manual_Tray
    if '3b_Manual_Tray' in wb_data.sheetnames:
        ws = wb_data['3b_Manual_Tray']
        
        animal_col = structure_3b.get('animal_col', 2)
        date_col = structure_3b.get('date_col', 1)
        sex_col = structure_3b.get('sex_col', 3)
        weight_col = structure_3b.get('weight_col', 4)
        phase_col = structure_3b.get('phase_col', 6)
        tray_col = structure_3b.get('tray_col', 7)
        pellet_start = structure_3b.get('pellet_start_col', 8)
        
        for row in range(2, ws.max_row + 1):
            animal = ws.cell(row, animal_col).value
            if not animal:
                continue
            animal = str(animal).strip()
            
            # Initialize animal if first time seeing it
            if animal not in data['animals']:
                sex = ws.cell(row, sex_col).value
                data['animals'][animal] = {
                    'sex': sex if sex else '',
                    'baseline_weight': None,
                }
                data['testing_days'][animal] = []
            
            # Get date - with data_only=True, formulas return their cached values
            date_val = ws.cell(row, date_col).value
            if date_val is None:
                continue
            
            # Convert to date
            if isinstance(date_val, datetime):
                test_date = date_val.date()
            elif isinstance(date_val, date):
                test_date = date_val
            else:
                # Skip non-date values (shouldn't happen with data_only=True)
                continue
            
            # Get phase - first try from 3b, then fall back to date_phase_map
            phase = ws.cell(row, phase_col).value if phase_col else None
            if not phase and test_date in date_phase_map:
                phase = date_phase_map[test_date]
            
            # Get other values
            weight = ws.cell(row, weight_col).value
            tray_label = ws.cell(row, tray_col).value
            tray_type = tray_label[0] if tray_label and isinstance(tray_label, str) else None
            
            # Extract pellet scores (columns 1-20)
            pellet_scores = []
            for p in range(20):
                val = ws.cell(row, pellet_start + p).value
                pellet_scores.append(val)
            
            # Store this tray's data
            data['testing_days'][animal].append({
                'date': test_date,
                'phase': phase,
                'tray_type': tray_type,
                'tray_label': tray_label,
                'weight': weight,
                'pellet_scores': pellet_scores,
            })
    
    # Extract from 0a_Metadata (DOB, DOD)
    if '0a_Metadata' in wb_data.sheetnames:
        ws = wb_data['0a_Metadata']
        # Find columns
        subj_col, dob_col, dod_col = 1, 2, 3
        for col in range(1, min(10, ws.max_column + 1)):
            header = ws.cell(1, col).value
            if header:
                h = str(header).lower()
                if 'subject' in h or 'animal' in h:
                    subj_col = col
                elif 'birth' in h or 'dob' in h:
                    dob_col = col
                elif 'death' in h or 'dod' in h:
                    dod_col = col
        
        for row in range(2, ws.max_row + 1):
            animal = ws.cell(row, subj_col).value
            if not animal:
                continue
            animal = str(animal).strip()
            
            if animal in data['animals']:
                dob = ws.cell(row, dob_col).value
                dod = ws.cell(row, dod_col).value
                if isinstance(dob, datetime):
                    dob = dob.date()
                if isinstance(dod, datetime):
                    dod = dod.date()
                data['animals'][animal]['dob'] = dob
                data['animals'][animal]['dod'] = dod
    
    # Extract from 3a_Manual_Ramp (baseline weight)
    if '3a_Manual_Ramp' in wb_data.sheetnames:
        ws = wb_data['3a_Manual_Ramp']
        # Find columns
        subj_col, weight_col = 1, 4
        for col in range(1, min(10, ws.max_column + 1)):
            header = ws.cell(1, col).value
            if header:
                h = str(header).lower()
                if 'animal' in h or 'subject' in h:
                    subj_col = col
                elif h == 'weight':
                    weight_col = col
        
        for row in range(2, ws.max_row + 1):
            animal = ws.cell(row, subj_col).value
            if not animal:
                continue
            animal = str(animal).strip()
            
            if animal in data['animals']:
                weight = ws.cell(row, weight_col).value
                if weight and data['animals'][animal]['baseline_weight'] is None:
                    data['animals'][animal]['baseline_weight'] = weight
    
    # Extract from 4_Contusion_Injury_Details
    if '4_Contusion_Injury_Details' in wb_data.sheetnames:
        ws = wb_data['4_Contusion_Injury_Details']
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = ws.cell(1, col).value
            if h:
                headers[str(h).strip()] = col
        
        subj_col = headers.get('Animal', headers.get('SubjectID', headers.get('Subject_ID', 1)))
        
        for row in range(2, ws.max_row + 1):
            animal = ws.cell(row, subj_col).value
            if not animal:
                continue
            animal = str(animal).strip()
            
            contusion_data = {}
            field_mappings = {
                'Surgery_Date': 'date',
                'Surgery_Type': 'type',
                'Surgery_Severity': 'severity',
                'Contusion_Location': 'location',
                'Subject_Weight (g)': 'weight',
                'Anesthetic': 'anesthetic',
                'Anesthetic_Dose': 'anesthetic_dose',
                'Anesthetic_Volume': 'anesthetic_volume',
                'Analgesic': 'analgesic',
                'Analgesic_Dose': 'analgesic_dose',
                'Analgesic_Volume': 'analgesic_volume',
                'Intended_kd': 'intended_kd',
                'Intended_Dwell': 'intended_dwell',
                'Stage_Height': 'stage_height',
                'Actual_kd': 'actual_kd',
                'Actual_displacement': 'actual_displacement',
                'Actual_Velocity': 'actual_velocity',
                'Actual_Dwell': 'actual_dwell',
                'Survived': 'survived',
            }
            
            for sheet_field, data_field in field_mappings.items():
                if sheet_field in headers:
                    val = ws.cell(row, headers[sheet_field]).value
                    if isinstance(val, datetime):
                        val = val.date()
                    contusion_data[data_field] = val
            
            data['contusion'][animal] = contusion_data
    
    # Extract from 5_SC_Injection_Details
    if '5_SC_Injection_Details' in wb_data.sheetnames:
        ws = wb_data['5_SC_Injection_Details']
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = ws.cell(1, col).value
            if h:
                headers[str(h).strip()] = col
        
        subj_col = headers.get('Animal', headers.get('SubjectID', headers.get('Subject_ID', 1)))
        
        for row in range(2, ws.max_row + 1):
            animal = ws.cell(row, subj_col).value
            if not animal:
                continue
            animal = str(animal).strip()
            
            injection_data = {}
            field_mappings = {
                'Surgery_Date': 'date',
                'Subject_Weight (g)': 'weight',
                'Surgery_Type': 'type',
                'Injected_Virus': 'virus',
                'Virus_Titer': 'titer',
                'Injection_Target': 'target',
                'Depths (D/V)': 'depth_dv',
                'Coordinates (M/L)': 'coord_ml',
                'Anesthetic': 'anesthetic',
                'Anesthetic_Dose': 'anesthetic_dose',
                'Anesthetic_Volume': 'anesthetic_volume',
                'Analgesic': 'analgesic',
                'Analgesic_Dose': 'analgesic_dose',
                'Analgesic_Volume': 'analgesic_volume',
                'Survived': 'survived',
                'Signal Post Perfusion': 'signal_post_perfusion',
            }
            
            for sheet_field, data_field in field_mappings.items():
                if sheet_field in headers:
                    val = ws.cell(row, headers[sheet_field]).value
                    if isinstance(val, datetime):
                        val = val.date()
                    injection_data[data_field] = val
            
            data['injection'][animal] = injection_data
    
    report.add_info(f"Extracted data for {len(data['animals'])} animals")
    return data


def compute_odc_rows(extracted_data, cohort_name, source_file, report):
    """
    Compute all ODC rows from extracted data using Python.
    
    Returns list of dicts, each dict is one ODC row with all values computed.
    """
    rows = []
    
    for animal in sorted(extracted_data['animals'].keys()):
        animal_info = extracted_data['animals'][animal]
        contusion = extracted_data['contusion'].get(animal, {})
        injection = extracted_data['injection'].get(animal, {})
        testing_data = extracted_data['testing_days'].get(animal, [])
        
        # Group testing data by date
        days_data = {}  # date -> {phase, tray_type, weight, trays: {label: pellet_scores}}
        for entry in testing_data:
            d = entry['date']
            if d not in days_data:
                days_data[d] = {
                    'phase': entry['phase'],
                    'tray_type': entry['tray_type'],
                    'weight': entry['weight'],
                    'trays': {}
                }
            if entry['tray_label']:
                days_data[d]['trays'][entry['tray_label']] = entry['pellet_scores']
        
        # Contusion date for Days_Post_Injury calculation
        contusion_date = contusion.get('date')
        
        # Create one row per testing day
        for test_date in sorted(days_data.keys()):
            day_info = days_data[test_date]
            
            row = {}
            
            # Section 1: ODC-SCI Required CoDEs
            row['SubjectID'] = animal
            row['SpeciesTyp'] = 'Mouse'
            row['SpeciesStrainTyp'] = 'C57BL/6J'
            row['AnimalSourceNam'] = 'Jackson Laboratory'
            
            # AgeVal - weeks from DOB to test date
            dob = animal_info.get('dob')
            if dob and isinstance(dob, date) and isinstance(test_date, date):
                age_days = (test_date - dob).days
                row['AgeVal'] = round(age_days / 7, 2)
            else:
                row['AgeVal'] = ''
            
            row['BodyWgtMeasrVal'] = animal_info.get('baseline_weight', '')
            row['SexTyp'] = animal_info.get('sex', '')
            row['InjGroupAssignTyp'] = cohort_name
            row['Laboratory'] = 'Murray/Blackmore Lab'
            row['StudyLeader'] = 'Logan Friedrich'
            
            # Exclusion based on survival
            survived = contusion.get('survived', 'Y')
            if survived == 'N':
                row['Exclusion_in_origin_study'] = 'Total exclusion'
                row['Exclusion_reason'] = 'Died during/after surgery'
                row['Cause_of_Death'] = 'Surgery complications'
            else:
                row['Exclusion_in_origin_study'] = 'No'
                row['Exclusion_reason'] = ''
                row['Cause_of_Death'] = ''
            
            # Injury details
            row['Injury_type'] = contusion.get('type', '')
            row['Injury_device'] = 'Infinite Horizon Impactor' if contusion.get('type') else ''
            row['Injury_level'] = contusion.get('location', '')
            
            # Build injury details string
            details_parts = []
            if contusion.get('actual_kd'):
                details_parts.append(f"{contusion['actual_kd']} kdyn")
            if contusion.get('actual_displacement'):
                details_parts.append(f"{contusion['actual_displacement']} mm displacement")
            if contusion.get('actual_velocity'):
                details_parts.append(f"{contusion['actual_velocity']} mm/s")
            row['Injury_details'] = ', '.join(details_parts) if details_parts else ''
            
            # Section 2: Contusion Surgery Details
            row['Contusion_Date'] = contusion.get('date', '')
            row['Contusion_Type'] = contusion.get('type', '')
            row['Contusion_Severity'] = contusion.get('severity', '')
            row['Contusion_Location'] = contusion.get('location', '')
            row['Contusion_Weight_g'] = contusion.get('weight', '')
            row['Contusion_Anesthetic'] = contusion.get('anesthetic', '')
            row['Contusion_Anesthetic_Dose'] = contusion.get('anesthetic_dose', '')
            row['Contusion_Anesthetic_Volume'] = contusion.get('anesthetic_volume', '')
            row['Contusion_Analgesic'] = contusion.get('analgesic', '')
            row['Contusion_Analgesic_Dose'] = contusion.get('analgesic_dose', '')
            row['Contusion_Analgesic_Volume'] = contusion.get('analgesic_volume', '')
            row['Contusion_Intended_kd'] = contusion.get('intended_kd', '')
            row['Contusion_Intended_Dwell'] = contusion.get('intended_dwell', '')
            row['Contusion_Stage_Height'] = contusion.get('stage_height', '')
            row['Contusion_Actual_kd'] = contusion.get('actual_kd', '')
            row['Contusion_Actual_Displacement'] = contusion.get('actual_displacement', '')
            row['Contusion_Actual_Velocity'] = contusion.get('actual_velocity', '')
            row['Contusion_Actual_Dwell'] = contusion.get('actual_dwell', '')
            row['Contusion_Survived'] = contusion.get('survived', '')
            
            # Section 3: SC Injection Details
            row['Injection_Date'] = injection.get('date', '')
            row['Injection_Weight_g'] = injection.get('weight', '')
            row['Injection_Type'] = injection.get('type', '')
            row['Injection_Virus'] = injection.get('virus', '')
            row['Injection_Titer'] = injection.get('titer', '')
            row['Injection_Target'] = injection.get('target', '')
            row['Injection_Depth_DV'] = injection.get('depth_dv', '')
            row['Injection_Coord_ML'] = injection.get('coord_ml', '')
            row['Injection_Anesthetic'] = injection.get('anesthetic', '')
            row['Injection_Anesthetic_Dose'] = injection.get('anesthetic_dose', '')
            row['Injection_Anesthetic_Volume'] = injection.get('anesthetic_volume', '')
            row['Injection_Analgesic'] = injection.get('analgesic', '')
            row['Injection_Analgesic_Dose'] = injection.get('analgesic_dose', '')
            row['Injection_Analgesic_Volume'] = injection.get('analgesic_volume', '')
            row['Injection_Survived'] = injection.get('survived', '')
            row['Injection_Signal_Post_Perfusion'] = injection.get('signal_post_perfusion', '')
            
            # Section 4: BrainGlobe placeholders
            row['Perfusion_Date'] = ''
            row['BrainGlobe_Analysis_Date'] = ''
            row['BrainGlobe_Atlas_Used'] = ''
            row['Total_Cells_Detected'] = ''
            row['Total_Cells_Left_Hemisphere'] = ''
            row['Total_Cells_Right_Hemisphere'] = ''
            row['BrainGlobe_Notes'] = ''
            row['BrainGlobe_Quality'] = ''
            
            # Section 5: Row metadata
            row['Date'] = test_date
            row['Test_Phase'] = day_info['phase']
            row['Tray_Type'] = day_info['tray_type']
            row['Num_Trays'] = len(day_info['trays'])
            row['Weight'] = day_info['weight']
            
            # Days_Post_Injury
            if contusion_date and isinstance(contusion_date, date) and isinstance(test_date, date):
                row['Days_Post_Injury'] = (test_date - contusion_date).days
            else:
                row['Days_Post_Injury'] = ''
            
            # Section 6: Per-pellet scores
            # Determine expected tray labels based on tray type
            tray_type = day_info['tray_type']
            if tray_type == 'F':
                expected_labels = ['F1', 'F2', 'F3', 'F4']
            elif tray_type == 'P':
                expected_labels = ['P1', 'P2', 'P3', 'P4']
            elif tray_type == 'E':
                expected_labels = ['E1', 'E2', 'E3', 'E4']
            else:
                expected_labels = ['F1', 'F2', 'F3', 'F4']
            
            # Initialize per-tray stats
            tray_stats = {}
            
            for tray_num in range(1, 5):
                tray_label = expected_labels[tray_num - 1]
                pellet_scores = day_info['trays'].get(tray_label, [None] * 20)
                
                # Ensure we have 20 pellets
                while len(pellet_scores) < 20:
                    pellet_scores.append(None)
                
                # Write pellet scores
                for pellet_num in range(1, 21):
                    col_name = f"Tray{tray_num}_Pellet{pellet_num:02d}"
                    score = pellet_scores[pellet_num - 1]
                    row[col_name] = score if score is not None else ''
                
                # Calculate tray statistics
                valid_scores = [s for s in pellet_scores if s is not None and s != '']
                presented = len(valid_scores)
                missed = sum(1 for s in valid_scores if s == 0)
                displaced = sum(1 for s in valid_scores if s == 1)
                retrieved = sum(1 for s in valid_scores if s == 2)
                contacted = displaced + retrieved
                
                tray_stats[tray_num] = {
                    'presented': presented,
                    'missed': missed,
                    'displaced': displaced,
                    'retrieved': retrieved,
                    'contacted': contacted,
                }
                
                # Section 7: Per-tray calculations
                row[f'Tray{tray_num}_Presented'] = presented
                row[f'Tray{tray_num}_Missed'] = missed
                row[f'Tray{tray_num}_Displaced'] = displaced
                row[f'Tray{tray_num}_Retrieved'] = retrieved
                row[f'Tray{tray_num}_Contacted'] = contacted
                
                if presented > 0:
                    row[f'Tray{tray_num}_Miss_Pct'] = round(missed / presented * 100, 2)
                    row[f'Tray{tray_num}_Displaced_Pct'] = round(displaced / presented * 100, 2)
                    row[f'Tray{tray_num}_Retrieved_Pct'] = round(retrieved / presented * 100, 2)
                    row[f'Tray{tray_num}_Contacted_Pct'] = round(contacted / presented * 100, 2)
                else:
                    row[f'Tray{tray_num}_Miss_Pct'] = ''
                    row[f'Tray{tray_num}_Displaced_Pct'] = ''
                    row[f'Tray{tray_num}_Retrieved_Pct'] = ''
                    row[f'Tray{tray_num}_Contacted_Pct'] = ''
            
            # Section 8: Daily totals
            total_presented = sum(tray_stats[t]['presented'] for t in range(1, 5))
            total_missed = sum(tray_stats[t]['missed'] for t in range(1, 5))
            total_displaced = sum(tray_stats[t]['displaced'] for t in range(1, 5))
            total_retrieved = sum(tray_stats[t]['retrieved'] for t in range(1, 5))
            total_contacted = sum(tray_stats[t]['contacted'] for t in range(1, 5))
            
            row['Total_Presented'] = total_presented
            row['Total_Missed'] = total_missed
            row['Total_Displaced'] = total_displaced
            row['Total_Retrieved'] = total_retrieved
            row['Total_Contacted'] = total_contacted
            
            # Daily averages (of percentages, not counts)
            pct_values = {'miss': [], 'displaced': [], 'retrieved': [], 'contacted': []}
            for t in range(1, 5):
                if tray_stats[t]['presented'] > 0:
                    pct_values['miss'].append(row[f'Tray{t}_Miss_Pct'])
                    pct_values['displaced'].append(row[f'Tray{t}_Displaced_Pct'])
                    pct_values['retrieved'].append(row[f'Tray{t}_Retrieved_Pct'])
                    pct_values['contacted'].append(row[f'Tray{t}_Contacted_Pct'])
            
            row['Avg_Miss_Pct'] = round(sum(pct_values['miss']) / len(pct_values['miss']), 2) if pct_values['miss'] else ''
            row['Avg_Displaced_Pct'] = round(sum(pct_values['displaced']) / len(pct_values['displaced']), 2) if pct_values['displaced'] else ''
            row['Avg_Retrieved_Pct'] = round(sum(pct_values['retrieved']) / len(pct_values['retrieved']), 2) if pct_values['retrieved'] else ''
            row['Avg_Contacted_Pct'] = round(sum(pct_values['contacted']) / len(pct_values['contacted']), 2) if pct_values['contacted'] else ''
            
            # Max/min
            row['Max_Retrieved_Pct'] = max(pct_values['retrieved']) if pct_values['retrieved'] else ''
            row['Max_Contacted_Pct'] = max(pct_values['contacted']) if pct_values['contacted'] else ''
            row['Min_Retrieved_Pct'] = min(pct_values['retrieved']) if pct_values['retrieved'] else ''
            row['Min_Contacted_Pct'] = min(pct_values['contacted']) if pct_values['contacted'] else ''
            
            # Section 9: Kinematic attention scores (placeholders)
            row['Tray1_Attention_Score'] = ''
            row['Tray2_Attention_Score'] = ''
            row['Tray3_Attention_Score'] = ''
            row['Tray4_Attention_Score'] = ''
            row['Total_Day_Attention_Score'] = ''
            
            # Section 10 & 11: Per-pellet kinematics and averages (placeholders)
            kinematic_metrics = [
                "Swipe_breadth", "Swipe_length", "Swipe_area", "Swipe_speed",
                "Path_length", "Swipe_Duration", "Swipe_Duration_Frames", "Path_over_Frames"
            ]
            
            for tray in range(1, 5):
                for pellet in range(1, 21):
                    for metric in kinematic_metrics:
                        row[f"Tray{tray}_Pellet{pellet:02d}_{metric}"] = ''
            
            for tray in range(1, 5):
                for metric in kinematic_metrics:
                    row[f"Tray{tray}_Avg_{metric}"] = ''
            
            for metric in kinematic_metrics:
                row[f"Day_Avg_{metric}"] = ''
            
            # Section 12: Source tracking
            row['Source_File'] = source_file
            row['Source_Sheet'] = '3b_Manual_Tray'
            row['Row_Notes'] = ''
            
            rows.append(row)
    
    report.add_info(f"Computed {len(rows)} ODC rows")
    return rows


def write_2_odc_from_existing_data(ws, wb, subject_ids, testing_days_unused, cohort_name, report, input_filepath=None):
    """
    Write 2_ODC_Animal_Tracking sheet using Python-extracted data with hard values.
    
    This discovers the structure of existing sheets, extracts all data with Python,
    computes all derived values, and writes hard values (not formulas).
    
    Args:
        ws: The worksheet to write to
        wb: The workbook (used to check sheet existence)
        subject_ids: List of subject IDs (may be used for fallback)
        testing_days_unused: Not used (legacy parameter)
        cohort_name: Name of the cohort
        report: FixReport object
        input_filepath: Path to the original file (needed to reload with data_only=True)
    """
    from openpyxl.utils import get_column_letter
    import openpyxl
    
    # Discover 3b structure from the formula workbook
    if '3b_Manual_Tray' not in wb.sheetnames:
        report.add_error("Cannot create ODC: 3b_Manual_Tray sheet not found")
        return
    
    structure_3b = discover_3b_structure(wb['3b_Manual_Tray'])
    report.add_info(f"Discovered 3b structure: animal={structure_3b.get('animal_col')}, "
                   f"date={structure_3b.get('date_col')}, tray={structure_3b.get('tray_col')}, "
                   f"pellets={structure_3b.get('pellet_start_col')}-{structure_3b.get('pellet_end_col')}")
    
    # Load with data_only=True to get cached formula values
    if input_filepath:
        wb_data = openpyxl.load_workbook(input_filepath, data_only=True)
        report.add_info("Loaded workbook with data_only=True for formula value extraction")
    else:
        # Fallback - use the formula workbook (may miss formula values)
        wb_data = wb
        report.add_warning("No input filepath provided - formula values may not be extracted")
    
    # Extract all data from workbook
    source_file = cohort_name
    extracted_data = extract_all_data_from_workbook(wb_data, structure_3b, report)
    
    # Compute all ODC rows
    odc_rows = compute_odc_rows(extracted_data, cohort_name, source_file, report)
    
    if not odc_rows:
        report.add_warning("No data rows computed for ODC sheet")
        return
    
    # Build headers list (same order as before)
    kinematic_metrics = [
        "Swipe_breadth", "Swipe_length", "Swipe_area", "Swipe_speed",
        "Path_length", "Swipe_Duration", "Swipe_Duration_Frames", "Path_over_Frames"
    ]
    
    headers = [
        # Section 1: ODC-SCI Required CoDEs (17)
        "SubjectID", "SpeciesTyp", "SpeciesStrainTyp", "AnimalSourceNam",
        "AgeVal", "BodyWgtMeasrVal", "SexTyp", "InjGroupAssignTyp",
        "Laboratory", "StudyLeader", "Exclusion_in_origin_study", "Exclusion_reason",
        "Cause_of_Death", "Injury_type", "Injury_device", "Injury_level", "Injury_details",
        # Section 2: Contusion Surgery (19)
        "Contusion_Date", "Contusion_Type", "Contusion_Severity", "Contusion_Location",
        "Contusion_Weight_g", "Contusion_Anesthetic", "Contusion_Anesthetic_Dose",
        "Contusion_Anesthetic_Volume", "Contusion_Analgesic", "Contusion_Analgesic_Dose",
        "Contusion_Analgesic_Volume", "Contusion_Intended_kd", "Contusion_Intended_Dwell",
        "Contusion_Stage_Height", "Contusion_Actual_kd", "Contusion_Actual_Displacement",
        "Contusion_Actual_Velocity", "Contusion_Actual_Dwell", "Contusion_Survived",
        # Section 3: SC Injection (16)
        "Injection_Date", "Injection_Weight_g", "Injection_Type", "Injection_Virus",
        "Injection_Titer", "Injection_Target", "Injection_Depth_DV", "Injection_Coord_ML",
        "Injection_Anesthetic", "Injection_Anesthetic_Dose", "Injection_Anesthetic_Volume",
        "Injection_Analgesic", "Injection_Analgesic_Dose", "Injection_Analgesic_Volume",
        "Injection_Survived", "Injection_Signal_Post_Perfusion",
        # Section 4: BrainGlobe (8)
        "Perfusion_Date", "BrainGlobe_Analysis_Date", "BrainGlobe_Atlas_Used",
        "Total_Cells_Detected", "Total_Cells_Left_Hemisphere", "Total_Cells_Right_Hemisphere",
        "BrainGlobe_Notes", "BrainGlobe_Quality",
        # Section 5: Row metadata (6)
        "Date", "Test_Phase", "Days_Post_Injury", "Tray_Type", "Num_Trays", "Weight",
    ]
    
    # Section 6: Per-pellet manual scores (80)
    for tray in range(1, 5):
        for pellet in range(1, 21):
            headers.append(f"Tray{tray}_Pellet{pellet:02d}")
    
    # Section 7: Per-tray calculations (36 = 9 x 4)
    for tray in range(1, 5):
        headers.extend([
            f"Tray{tray}_Presented", f"Tray{tray}_Missed", f"Tray{tray}_Displaced",
            f"Tray{tray}_Retrieved", f"Tray{tray}_Contacted", f"Tray{tray}_Miss_Pct",
            f"Tray{tray}_Displaced_Pct", f"Tray{tray}_Retrieved_Pct", f"Tray{tray}_Contacted_Pct"
        ])
    
    # Section 8: Daily totals/averages (13)
    headers.extend([
        "Total_Presented", "Total_Missed", "Total_Displaced", "Total_Retrieved", "Total_Contacted",
        "Avg_Miss_Pct", "Avg_Displaced_Pct", "Avg_Retrieved_Pct", "Avg_Contacted_Pct",
        "Max_Retrieved_Pct", "Max_Contacted_Pct", "Min_Retrieved_Pct", "Min_Contacted_Pct"
    ])
    
    # Section 9: Kinematic attention scores (5)
    headers.extend(["Tray1_Attention_Score", "Tray2_Attention_Score", 
                    "Tray3_Attention_Score", "Tray4_Attention_Score", "Total_Day_Attention_Score"])
    
    # Section 10: Per-pellet kinematics (640 = 80 pellets x 8 metrics)
    for tray in range(1, 5):
        for pellet in range(1, 21):
            for metric in kinematic_metrics:
                headers.append(f"Tray{tray}_Pellet{pellet:02d}_{metric}")
    
    # Section 11: Kinematic averages (40 = 32 tray + 8 day)
    for tray in range(1, 5):
        for metric in kinematic_metrics:
            headers.append(f"Tray{tray}_Avg_{metric}")
    for metric in kinematic_metrics:
        headers.append(f"Day_Avg_{metric}")
    
    # Section 12: Source tracking (3)
    headers.extend(["Source_File", "Source_Sheet", "Row_Notes"])
    
    # Write headers
    for c, header in enumerate(headers, 1):
        ws.cell(1, c, value=header)
    
    # Write data rows with hard values
    for row_idx, row_data in enumerate(odc_rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    report.add_info(f"Wrote ODC with {len(odc_rows)} rows, {len(headers)} columns (hard values)")
    autoscale_columns(ws)


def check_and_fix_sheet(wb, sheet_name, cohort_info, report):
    """Check and fix column names in a sheet"""
    ws = wb[sheet_name]
    
    # Get current headers
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        headers.append(val)
    
    # Check for columns that need renaming
    changes = []
    for col_idx, header in enumerate(headers):
        if header in COLUMN_NAME_MAPPINGS:
            new_name = COLUMN_NAME_MAPPINGS[header]
            ws.cell(1, col_idx + 1, value=new_name)
            changes.append(f"{header} -> {new_name}")
    
    if changes:
        report.add_fix("Column Names", f"{sheet_name}: Renamed {len(changes)} columns")
        for change in changes:
            report.add_info(f"  {sheet_name}: {change}")


def check_and_fix_odc_structure(wb, cohort_info, report):
    """
    Check if ODC sheet has old per-tray structure and convert to per-day
    """
    ws = wb["2_ODC_Animal_Tracking"]
    
    # Check row count to determine structure
    row_count = ws.max_row - 1  # Exclude header
    col_count = ws.max_column
    
    num_mice = len(cohort_info['subject_ids'])
    num_days = len(TIMELINE)
    
    expected_per_tray = sum(t[3] for t in TIMELINE) * num_mice  # trays_per_day * num_mice for each day
    expected_per_day = num_days * num_mice
    
    report.add_info(f"ODC sheet: {row_count} rows, {col_count} columns")
    report.add_info(f"Expected per-tray: ~{expected_per_tray} rows")
    report.add_info(f"Expected per-day: {expected_per_day} rows")
    
    # Determine current structure
    if row_count > expected_per_day * 1.5:  # Likely per-tray structure
        report.add_issue("WARNING", "ODC Structure", 
                        f"ODC appears to have per-tray structure ({row_count} rows)")
        report.add_fix("ODC Structure", "Converting ODC from per-tray to per-day structure")
        
        # Delete the old ODC sheet
        del wb["2_ODC_Animal_Tracking"]
        
        # Create new per-day structure
        ws_new = wb.create_sheet(title="2_ODC_Animal_Tracking")
        data = create_2_odc_animal_tracking(cohort_info['subject_ids'], cohort_info['start_date'])
        write_2_odc_with_formulas(ws_new, data, cohort_info['cohort_name'])
        
        report.add_info(f"Created new ODC sheet with {expected_per_day} rows (per-day structure)")
    
    elif col_count < 500:  # Likely old structure without kinematics
        report.add_issue("WARNING", "ODC Structure", 
                        f"ODC appears to be missing kinematic columns ({col_count} columns)")
        report.add_fix("ODC Structure", "Recreating ODC with full column structure including kinematics")
        
        # Delete the old ODC sheet
        del wb["2_ODC_Animal_Tracking"]
        
        # Create new structure
        ws_new = wb.create_sheet(title="2_ODC_Animal_Tracking")
        data = create_2_odc_animal_tracking(cohort_info['subject_ids'], cohort_info['start_date'])
        write_2_odc_with_formulas(ws_new, data, cohort_info['cohort_name'])
        
        report.add_info(f"Created new ODC sheet with ~883 columns (includes kinematics)")
    
    else:
        report.add_info("ODC structure appears correct (per-day with kinematics)")


def reorder_sheets(wb, report):
    """Reorder sheets to match expected order"""
    current_order = wb.sheetnames
    
    # Build new order: expected sheets first (in order), then extras
    new_order = []
    for expected in EXPECTED_SHEETS:
        if expected in current_order:
            new_order.append(expected)
    
    # Add any extra sheets at the end
    for sheet in current_order:
        if sheet not in new_order:
            new_order.append(sheet)
    
    # Check if reordering needed
    if current_order != new_order:
        # Reorder by moving sheets
        for i, sheet_name in enumerate(new_order):
            wb.move_sheet(sheet_name, offset=i - wb.sheetnames.index(sheet_name))
        
        report.add_fix("Sheet Order", "Reordered sheets to standard order")


def interactive_mode():
    """Main interactive interface"""
    print_header()
    
    print(f"Script location: {SCRIPT_DIR}")
    print(f"Output directory: {OUTPUT_SUBDIR}")
    print()
    
    print("What would you like to do?\n")
    print("  1. Create a new cohort file")
    print("  2. Fix an existing cohort file")
    print("  q. Quit")
    print()
    
    choice = get_user_choice(
        "Enter choice (1, 2, or q)",
        valid_options=["1", "2", "q", "Q"]
    )
    
    if choice == "1":
        result = interactive_new_cohort()
        if result:
            print("\n" + "=" * 60)
            print("Done! Your cohort file is ready.")
            print(f"Location: {result}")
            print("=" * 60)
    elif choice == "2":
        interactive_fix_existing()
    elif choice.lower() == "q":
        print("\nGoodbye!")
    
    # Pause before exit so double-click users can see the output
    print()
    input("Press Enter to exit...")


def detect_project_type(cli_type=None, cohort_name=None, filename=None):
    """
    Detect project type with priority: CLI > filename > prompt
    
    Returns: 'cnt' or 'encr'
    """
    # Priority 1: CLI argument
    if cli_type:
        return cli_type.lower()
    
    # Priority 2: Detect from cohort name or filename
    check_strings = []
    if cohort_name:
        check_strings.append(cohort_name.upper())
    if filename:
        check_strings.append(str(filename).upper())
    
    for s in check_strings:
        if 'ENCR' in s or 'ENHANCER' in s:
            return 'encr'
        if 'CNT' in s or 'CONNECTOME' in s:
            return 'cnt'
    
    # Priority 3: Prompt user
    print("\nProject type not detected.")
    print("  1. CNT  - Connectome (behavior + tracing)")
    print("  2. ENCR - Enhancer (BrainGlobe, no behavior)")
    
    while True:
        choice = input("Select project type [1/2]: ").strip()
        if choice == '1' or choice.lower() == 'cnt':
            return 'cnt'
        elif choice == '2' or choice.lower() == 'encr':
            return 'encr'
        print("Invalid choice. Enter 1 or 2.")


# =============================================================================
# COMMAND LINE INTERFACE
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Create or fix cohort Excel files for ODC-SCI data pipeline. "
                    "Run without arguments for interactive mode."
    )
    
    # Use flat argument style for CLI mode
    parser.add_argument("--new", action="store_true", help="Create new cohort file (CLI mode)")
    parser.add_argument("--fix", type=str, default=None, metavar="FILE", help="Fix existing cohort file (CLI mode)")
    parser.add_argument("--cohort", type=str, help="Cohort name (e.g., CNT_05 or ENCR_01)")
    parser.add_argument("--start-date", type=str, help="Food deprivation start date (YYYY-MM-DD)")
    parser.add_argument("--mice", type=int, default=16, help="Number of mice (default: 16)")
    parser.add_argument("--output-dir", type=str, default=None, help="Output directory (default: current)")
    parser.add_argument("--type", type=str, choices=["cnt", "encr"], default=None, 
                        help="Project type: cnt (connectome/behavior) or encr (enhancer/brainglobe)")
    parser.add_argument("--viruses", type=int, default=3, help="Max viruses per injection (default: 3)")
    
    args = parser.parse_args()
    
    # Check if any CLI arguments were provided
    if args.new or args.fix:
        # CLI mode
        if args.new:
            if not args.cohort or not args.start_date:
                print("Error: --new requires --cohort and --start-date")
                print("\nUsage: python 0_Make_or_Fix_Sheets.py --new --cohort CNT_05 --start-date 2025-02-01")
                print("       python 0_Make_or_Fix_Sheets.py --new --cohort ENCR_01 --start-date 2025-02-01 --type encr")
                sys.exit(1)
            
            # Detect project type
            project_type = detect_project_type(args.type, args.cohort, None)
            print(f"Project type: {project_type.upper()}")
            
            create_new_cohort_file(
                args.cohort,
                args.start_date,
                args.mice,
                args.output_dir,
                project_type=project_type,
                max_viruses=args.viruses
            )
        elif args.fix:
            # Detect project type from filename
            project_type = detect_project_type(args.type, None, args.fix)
            print(f"Project type: {project_type.upper()}")
            
            result = fix_existing_file(args.fix, args.output_dir, 
                                       project_type=project_type, 
                                       max_viruses=args.viruses)
            if not result:
                sys.exit(1)
    else:
        # Interactive mode (no arguments provided)
        interactive_mode()


if __name__ == "__main__":
    main()
