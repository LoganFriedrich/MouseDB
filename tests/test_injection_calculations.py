"""0_Injection_Calculations: the lab's batch-block layout, generated and read.

The layout was settled 2026-08-28 (see cohort_tools.make_sheets
INJECTION_CALC_HEADERS). These tests pin it: the generator writes exactly
that layout, and the importer reads a filled-in block back -- including the
real-world quirks (a batch title with a trailing YYYYMMDD, template
placeholders, a Totals row) -- while the two legacy layouts still fall
through to their old parsers.
"""
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from mousedb.cohort_tools.make_sheets import (
    INJECTION_CALC_HEADERS, DISAMBIGUATION_FACTOR,
    create_0_injection_calculations, write_0_injection_calculations_with_formulas,
)
from mousedb.importers import ExcelImporter


class _Query:
    def filter_by(self, **kw):
        return self

    def first(self):
        return None


class FakeSession:
    def __init__(self):
        self.added = []

    def query(self, *a):
        return _Query()

    def add(self, obj):
        self.added.append(obj)


def _importer():
    imp = ExcelImporter.__new__(ExcelImporter)
    imp.warnings, imp.errors = [], []
    imp.imported_counts = {'virus_preps': 0}
    return imp


def _sheet_df(rows):
    """A header=None DataFrame exactly as pd.read_excel would hand it over."""
    width = max(len(r) for r in rows)
    return pd.DataFrame([list(r) + [None] * (width - len(r)) for r in rows])


LAB_BLOCK = [
    ["Batch that didn't qualify on performance metrics - 20260826"],
    INJECTION_CALC_HEADERS,
    ["AAV5-CAG-GFP (lot:v57830) from hearing lab", 100, 6, 3.0, 50.0, 5e12, 3],
    ["AAV-Retro-CAGIG-H2B-V5-mSc (Validated AAV Box 3)", 60, 6, 2.5, 25.0, 2.5e12, None],
    ["Malat BC 7 (Box 40)", 36.5, 6, 0.5, 18.25, 1.825e12, None],
    ["Totals", None, None, 6.0, 78.04, 7.804e12],
]


class TestGenerator:
    def test_writes_the_lab_layout(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        write_0_injection_calculations_with_formulas(
            ws, create_0_injection_calculations(num_batches=1, viruses_per_batch=3))
        assert str(ws.cell(1, 1).value).startswith("Batch")
        assert [ws.cell(2, c + 1).value for c in range(7)] == INJECTION_CALC_HEADERS
        # three virus rows with the two formulas, then Totals
        for r in (3, 4, 5):
            assert ws.cell(r, 5).value == f'=IFERROR(B{r}*D{r}/C{r},"")'
            assert ws.cell(r, 6).value == f'=IFERROR(E{r}*{DISAMBIGUATION_FACTOR},"")'
        assert ws.cell(6, 1).value == "Totals"
        assert ws.cell(6, 4).value == "=SUM(D3:D5)"

    def test_shorthand_conversion_matches_the_lab_convention(self):
        # the lab writes "10" for 1x10^12
        assert 10 * DISAMBIGUATION_FACTOR == 1e12
        assert 50 * DISAMBIGUATION_FACTOR == 5e12

    def test_two_batches_are_separated_by_blank_rows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        write_0_injection_calculations_with_formulas(
            ws, create_0_injection_calculations(num_batches=2, viruses_per_batch=2))
        titles = [r for r in range(1, ws.max_row + 1)
                  if str(ws.cell(r, 1).value or "").startswith("Batch")]
        assert len(titles) == 2
        assert ws.cell(titles[1] - 1, 1).value is None  # blank before batch 2


class TestImporter:
    def test_reads_a_filled_batch_block(self):
        imp = _importer()
        s = FakeSession()
        assert imp._import_injection_batches(_sheet_df(LAB_BLOCK), 'CNT_05', s, False)
        assert imp.imported_counts['virus_preps'] == 1
        vp = s.added[0]
        assert str(vp.prep_date) == "2026-08-26"
        assert vp.preparation_notes == "that didn't qualify on performance metrics"
        assert vp.num_animals == 3
        # summed from the per-virus rows, NOT read from the Totals cell (which
        # is a derived value and, in the real CNT_05 sheet, was stale)
        assert vp.final_titer == pytest.approx(5e12 + 2.5e12 + 1.825e12, rel=1e-6)
        assert vp.virus_name.startswith("AAV5-CAG-GFP") and "Malat BC 7" in vp.virus_name
        assert "parts 3.0/6.0" in vp.calculation_notes

    def test_empty_template_is_batch_layout_but_imports_nothing(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "0_Injection_Calculations"
        write_0_injection_calculations_with_formulas(
            ws, create_0_injection_calculations())
        out = tmp_path / "t.xlsx"
        wb.save(out)
        df = pd.read_excel(out, sheet_name="0_Injection_Calculations", header=None)
        imp = _importer()
        assert imp._import_injection_batches(df, 'CNT_99', FakeSession(), True) is True
        assert imp.imported_counts['virus_preps'] == 0
        assert any("no virus rows" in w for w in imp.warnings)

    def test_batch_without_date_is_skipped_with_a_warning(self):
        rows = [["Batch undated"]] + LAB_BLOCK[1:]
        imp = _importer()
        assert imp._import_injection_batches(_sheet_df(rows), 'CNT_05', FakeSession(), True)
        assert imp.imported_counts['virus_preps'] == 0
        assert any("no YYYYMMDD" in w for w in imp.warnings)

    def test_legacy_layouts_are_not_batch_layout(self):
        vertical = _sheet_df([["Parameter", "Value", "Units"], ["Virus Name", None, None]])
        horizontal = _sheet_df([["Date of surgery", "Virus Name", "Final concentration"]])
        imp = _importer()
        assert imp._import_injection_batches(vertical, 'CNT_05', FakeSession(), True) is False
        assert imp._import_injection_batches(horizontal, 'CNT_01', FakeSession(), True) is False

    def test_the_block_beneath_a_legacy_template_still_wins(self):
        # CNT_05's real sheet: an empty vertical calculator on top, the lab
        # block further down. The block must be found regardless of position.
        rows = [["Parameter", "Value", "Units"], ["Virus Name"], [], []] + LAB_BLOCK
        imp = _importer()
        s = FakeSession()
        assert imp._import_injection_batches(_sheet_df(rows), 'CNT_05', s, False)
        assert len(s.added) == 1
