"""Import ASPA manual pellet scores from the frozen ASPA workbooks.

WHY
---
ASPA is a project exactly like CNT and its videos run through the same
pipeline -- but until 2026-08-28 no ASPA manual score had ever reached the
database, because `mousedb import` reads only Connectome_NN workbooks. With
no bench scores, the bench-vs-algorithm disagreement scan, the manual-vs-
automated accuracy figures and the ODC session export all silently excluded
every ASPA cohort. This closes that.

THE SHEET
---------
Each ASPA workbook (one per cohort letter: I.xlsx, J.xlsx, "K - Contusion
70kd.xlsx" ...) has a tab '1 - ENTER DATA HERE' in a WIDE layout: one row
per animal-day, then four tray blocks side by side --

    Test Date | Test Type | Tray Type | Test Phase | Group | Animal # |
    [Tray # | 1 ... 20 | Displaced | Eaten | hit] x 4

Pellet cells hold 0 (missed), 1 (displaced), 2 (eaten = retrieved) -- the
same vocabulary as the CNT sheets. A tray whose 20 cells are all blank was
not scored and is skipped. Tray Type is Pillar / Easy / Flat -> P / E / F.
Test Type / Test Phase are the operator's labels and are NOT used: the
phase is derived from the cohort's date structure exactly as for CNT.

Animal ids are encoded the pipeline's way (J11 -> ASPA_10_11: cohort =
letter's alphabet position); subjects are created if missing, like every
other import.

USAGE
-----
    python -m mousedb.cohort_tools.import_aspa_scores            # dry run, all cohorts
    python -m mousedb.cohort_tools.import_aspa_scores --cohort J --apply

Every run (with --apply) is recorded in the sheet-import ledger the
Tracking Sheets tab reads, keyed by the cohort id (ASPA_10).
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

SCORE_SHEET = "1 - ENTER DATA HERE"
TRAY_TYPE = {"pillar": "P", "easy": "E", "flat": "F"}
PELLETS = 20


def parse_workbook(path: Path, cohort_id: str, warnings: List[str]) -> List[Tuple]:
    """(subject_id, date, tray_type, tray_number, {pellet: score}) rows."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if SCORE_SHEET not in wb.sheetnames:
            warnings.append("%s: no '%s' tab" % (path.name, SCORE_SHEET))
            return []
        ws = wb[SCORE_SHEET]
        it = ws.iter_rows(values_only=True)
        header = [str(v).strip() if v is not None else "" for v in next(it)]
        blocks = [i for i, h in enumerate(header) if h == "Tray #"]
        if not blocks:
            warnings.append("%s: no 'Tray #' columns in '%s'" % (path.name, SCORE_SHEET))
            return []
        try:
            c_date = header.index("Test Date")
            c_tray = header.index("Tray Type")
            c_animal = header.index("Animal #")
        except ValueError as e:
            warnings.append("%s: missing header column (%s)" % (path.name, e))
            return []

        out = []
        for r in it:
            if r is None or c_date >= len(r) or r[c_date] is None:
                continue
            d = r[c_date]
            d = d.date() if isinstance(d, datetime) else d
            try:
                animal = int(float(r[c_animal]))
            except (TypeError, ValueError):
                continue
            tt = TRAY_TYPE.get(str(r[c_tray] or "").strip().lower())
            if tt is None:
                warnings.append("%s: unknown tray type %r on %s animal %s" % (
                    path.name, r[c_tray], d, animal))
                continue
            subject_id = "%s_%02d" % (cohort_id, animal)
            for b in blocks:
                try:
                    tray_num = int(float(r[b])) if r[b] not in (None, "") else None
                except (TypeError, ValueError):
                    tray_num = None
                cells = list(r[b + 1:b + 1 + PELLETS])
                scores: Dict[int, int] = {}
                for p, v in enumerate(cells, 1):
                    if v in (None, ""):
                        continue
                    try:
                        s = int(float(v))
                    except (TypeError, ValueError):
                        warnings.append("%s: bad score %r (%s %s tray %s pellet %d)" % (
                            path.name, v, subject_id, d, tray_num, p))
                        continue
                    if s not in (0, 1, 2):
                        warnings.append("%s: score %r out of range (%s %s tray %s pellet %d)" % (
                            path.name, v, subject_id, d, tray_num, p))
                        continue
                    scores[p] = s
                if not scores or tray_num is None:
                    continue
                if not 1 <= tray_num <= 4:
                    warnings.append("%s: tray number %s outside 1-4 (%s %s)" % (
                        path.name, tray_num, subject_id, d))
                    continue
                out.append((subject_id, d, tt, tray_num, scores))
        return out
    finally:
        wb.close()


def import_cohort(letter: str, apply: bool = False) -> dict:
    """Import one ASPA cohort's manual scores. Returns a ledger-shaped dict."""
    from ..cohort_sheets import find_aspa_sheet, aspa_cohort_number
    from ..importers import ExcelImporter
    from ..database import init_database

    num = aspa_cohort_number(letter)
    cohort_id = "ASPA_%s" % num
    sheet = find_aspa_sheet(letter)
    entry = {"cohort_id": cohort_id, "sheet_name": sheet.name if sheet else None,
             "sheet_path": str(sheet) if sheet else None,
             "started": datetime.now().isoformat(timespec="seconds"),
             "triggered_by": "aspa-scores", "dry_run": not apply}
    if sheet is None:
        entry.update(success=False, error="no ASPA sheet for cohort %s" % letter)
        return entry
    entry["sheet_mtime"] = datetime.fromtimestamp(sheet.stat().st_mtime).isoformat(timespec="seconds")
    warnings: List[str] = []
    try:
        rows = parse_workbook(sheet, cohort_id, warnings)
        n_pellets = sum(len(s) for *_, s in rows)
        db = init_database()
        imp = ExcelImporter(db)
        imp.imported_counts = {"subjects": 0, "pellet_scores": 0}
        with db.session() as session:
            from ..schema import Cohort, Project
            if not session.query(Cohort).filter_by(cohort_id=cohort_id).first() and rows:
                # A cohort the registration tool skipped (no videos on this
                # machine) can still be created here, because the workbook
                # gives a real start date: the earliest scored session. Never
                # today's date, never a guess.
                if not session.query(Project).filter_by(project_code="ASPA").first():
                    session.add(Project(project_code="ASPA", project_name="ASPA"))
                    session.flush()
                start = min(r[1] for r in rows)
                if not apply:
                    entry["would_create_cohort"] = {"cohort_id": cohort_id, "start_date": str(start)}
                else:
                    session.add(Cohort(cohort_id=cohort_id, project_code="ASPA",
                                       start_date=start,
                                       notes="Registered from %s by the ASPA score import; "
                                             "start date = earliest scored session" % sheet.name))
                    session.flush()
            imp._insert_pellet_rows(session, cohort_id, rows, dry_run=not apply)
            if apply:
                session.commit()
        entry.update(success=True, parsed_trays=len(rows), parsed_pellets=n_pellets,
                     imported=dict(imp.imported_counts),
                     warnings=(warnings + imp.warnings)[:50], error=None)
    except Exception as e:
        import traceback
        entry.update(success=False, error="%s: %s" % (type(e).__name__, e),
                     traceback=traceback.format_exc()[-2000:])
    entry["finished"] = datetime.now().isoformat(timespec="seconds")
    if apply:
        from ..sheet_sync import _append_ledger
        _append_ledger(entry)
    return entry


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--cohort", action="append", metavar="LETTER",
                    help="Only this cohort letter; may be repeated")
    ap.add_argument("--apply", action="store_true", help="Actually write (default: dry run)")
    args = ap.parse_args(argv)

    from ..cohort_sheets import available_aspa_cohorts, aspa_data_dir, describe
    if aspa_data_dir() is None:
        print(describe(), file=sys.stderr)
        return 2
    letters = [c.upper() for c in args.cohort] if args.cohort else available_aspa_cohorts()
    ok = True
    for L in letters:
        e = import_cohort(L, apply=args.apply)
        tag = "OK  " if e.get("success") else "FAIL"
        print("%s %s %-28s trays=%s pellets=%s imported=%s%s" % (
            tag, e["cohort_id"], (e.get("sheet_name") or "-")[:28],
            e.get("parsed_trays", "-"), e.get("parsed_pellets", "-"),
            e.get("imported") or e.get("error"),
            "  [%d warning(s)]" % len(e["warnings"]) if e.get("warnings") else ""))
        for w in (e.get("warnings") or [])[:5]:
            print("      %s" % w)
        ok = ok and bool(e.get("success"))
    if not args.apply:
        print("(dry run -- nothing written; re-run with --apply)")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
