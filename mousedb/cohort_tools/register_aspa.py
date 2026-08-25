#!/usr/bin/env python3
"""Register ASPA cohorts and animals in the database, from the ASPA animal sheets.

WHY
---
ASPA is a project in exactly the way CNT is, and its videos run through the same
pipeline. But no ASPA cohort or subject exists in the database, so
`sync_file_to_database` skips every ASPA video -- it will not write reach data for
an animal it has never heard of. Sixty-plus analysed ASPA sessions have nowhere to
land.

This registers the project, the cohorts, and the animals, so that data can flow.

WHY IT IS NOT JUST "READ THE SHEET"
-----------------------------------
The ASPA workbooks are a frozen, hand-kept record and they are not uniform.
`I.xlsx` has H-cohort animals in its *Weights* and *balance* tabs -- 40 of them,
identical to `H.xlsx`, evidently left over from copying that workbook -- while the
real I animals live in *"Ramp" Training Data*. Reading a fixed sheet by name would
have registered forty H animals as cohort I.

So the rule is: collect ids matching THIS cohort's own letter, anywhere in its
workbook, and cross-check them against the videos that actually exist. Both counts
are reported, and anything that appears in one and not the other is called out
rather than quietly merged.

IDS
---
ASPA names animals `{letter}{subject}` (J11, and loosely J1). The database uses
the pipeline's encoded form so that a video, a features file and a database row
all agree: cohort number = the letter's alphabet position, so J -> cohort 10 and
the subject id is `ASPA_10_11`. Decoding back to `J11` happens on export.

USAGE
-----
    python -m mousedb.cohort_tools.register_aspa                 # dry run, all cohorts
    python -m mousedb.cohort_tools.register_aspa --cohort J      # one cohort
    python -m mousedb.cohort_tools.register_aspa --apply         # actually write

Dry run by default: it prints exactly which cohorts and animals it would create
and changes nothing. Nothing is ever deleted or overwritten -- a cohort or subject
that already exists is left alone.

Requires the ASPA animal sheets to be configured:
    mousedb cohort-sheets --set-aspa "<path>"
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Set

PROJECT = "ASPA"


def animals_in_sheet(path: Path, letter: str) -> Set[int]:
    """Subject numbers for ``letter`` found anywhere in this workbook.

    Deliberately not tied to a sheet name -- see the module docstring for why.
    """
    import openpyxl

    # "J11", loosely "J1", the surgery sheets' "J-6", and the earlier cohorts'
    # "OptD12" -- the same "Opt" misnomer that shows up in their filenames, here
    # on the animal id itself. Without it, D, E, F and G yield no animals at all.
    pattern = re.compile(r"^(?:Opt)?%s[-\s]?(\d{1,2})$" % re.escape(letter), re.I)
    found: Set[int] = set()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        m = pattern.match(cell.strip())
                        if m:
                            n = int(m.group(1))
                            if 1 <= n <= 99:
                                found.add(n)
    finally:
        wb.close()
    return found


def animals_with_videos(letter: str) -> Set[int]:
    """Subject numbers for ``letter`` that have analysed videos on this machine."""
    try:
        from mousedb.cohort_sheets import aspa_cohort_number
    except Exception:
        return set()
    num = aspa_cohort_number(letter)
    if num is None:
        return set()

    roots = []
    try:
        from mousereach.config import Paths
        for r in (Paths.ANALYZED_OUTPUT, Paths.PROCESSING):
            if r:
                roots.append(Path(r))
    except Exception:
        pass

    found: Set[int] = set()
    for root in roots:
        if not root.exists():
            continue
        for f in root.rglob("*%s%s*_features.json" % (PROJECT, num)):
            m = re.search(PROJECT + num + r"(\d{2})", f.name)
            if m:
                found.add(int(m.group(1)))
    return found


def plan(letters: Optional[List[str]] = None) -> List[dict]:
    """What would be registered, per cohort. Reads only."""
    from mousedb.cohort_sheets import (
        available_aspa_cohorts, find_aspa_sheet, aspa_cohort_number)

    letters = letters or available_aspa_cohorts()
    out = []
    for L in letters:
        sheet = find_aspa_sheet(L)
        if sheet is None:
            out.append({"letter": L, "error": "no sheet found"})
            continue
        in_sheet = animals_in_sheet(sheet, L)
        with_video = animals_with_videos(L)
        out.append({
            "letter": L,
            "cohort_id": "%s_%s" % (PROJECT, aspa_cohort_number(L)),
            "sheet": sheet.name,
            "in_sheet": sorted(in_sheet),
            "with_video": sorted(with_video),
            "video_only": sorted(with_video - in_sheet),
            "register": sorted(in_sheet | with_video),
        })
    return out


def register(entries: List[dict], apply: bool = False) -> Dict[str, int]:
    """Create the project, cohorts and subjects. Never deletes or overwrites."""
    from mousedb.database import init_database
    from mousedb.schema import Project, Cohort, Subject

    counts = defaultdict(int)
    db = init_database()
    with db.session() as session:
        project = session.query(Project).filter_by(project_code=PROJECT).first()
        if not project:
            counts["projects"] += 1
            if apply:
                session.add(Project(project_code=PROJECT, project_name=PROJECT))
                session.flush()

        for e in entries:
            if e.get("error") or not e.get("register"):
                continue
            cid = e["cohort_id"]
            cohort = session.query(Cohort).filter_by(cohort_id=cid).first()
            if not cohort:
                counts["cohorts"] += 1
                if apply:
                    session.add(Cohort(cohort_id=cid, project_code=PROJECT,
                                       num_mice=len(e["register"]),
                                       notes="Registered from %s" % e["sheet"]))
                    session.flush()
            for n in e["register"]:
                sid = "%s_%02d" % (cid, n)
                if session.query(Subject).filter_by(subject_id=sid).first():
                    counts["subjects_existing"] += 1
                    continue
                counts["subjects"] += 1
                if apply:
                    session.add(Subject(subject_id=sid, cohort_id=cid))
        if apply:
            session.commit()
    return dict(counts)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        description="Register ASPA cohorts and animals in the database.",
        epilog="Dry run by default. Re-run with --apply to write.")
    ap.add_argument("--cohort", action="append", metavar="LETTER",
                    help="Only this cohort letter; may be repeated")
    ap.add_argument("--apply", action="store_true", help="Actually write")
    args = ap.parse_args(argv)

    try:
        from mousedb.cohort_sheets import aspa_data_dir, describe
    except Exception as e:
        print("ERROR: could not import mousedb (%s)" % e, file=sys.stderr)
        return 2
    if aspa_data_dir() is None:
        print(describe(), file=sys.stderr)
        return 2

    letters = [c.upper() for c in args.cohort] if args.cohort else None
    entries = plan(letters)

    print("%-4s %-9s %-28s %6s %6s %8s" %
          ("", "cohort", "sheet", "sheet", "videos", "register"))
    for e in entries:
        if e.get("error"):
            print("%-4s %-9s %s" % (e["letter"], "-", e["error"]))
            continue
        print("%-4s %-9s %-28s %6d %6d %8d" % (
            e["letter"], e["cohort_id"], e["sheet"][:28],
            len(e["in_sheet"]), len(e["with_video"]), len(e["register"])))
        if e["video_only"]:
            print("       videos exist for animals the sheet does not list: %s"
                  % ", ".join("%s%d" % (e["letter"], n) for n in e["video_only"]))

    counts = register(entries, apply=args.apply)
    print()
    verb = "Created" if args.apply else "Would create"
    print("%s: %d project, %d cohort(s), %d subject(s)"
          % (verb, counts.get("projects", 0), counts.get("cohorts", 0),
             counts.get("subjects", 0)))
    if counts.get("subjects_existing"):
        print("  %d subject(s) already existed and were left alone"
              % counts["subjects_existing"])
    if not args.apply:
        print("\n(dry run -- nothing written; re-run with --apply)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
