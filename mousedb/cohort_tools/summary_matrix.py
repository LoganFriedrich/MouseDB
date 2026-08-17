"""Rebuild the Manual Summary tab as a readable matrix.

The problem
-----------
CNT_05's ``3c_Manual_Summary`` is in tidy/long form -- one row per animal per
session (214 rows). That is fine for a machine and useless for a human: to follow
one mouse you filter and scroll, and nothing is visible at a glance.

CNT_01-03 used a matrix instead, and it worked because BOTH directions carry
meaning: read ACROSS a row to see one mouse's trajectory over the experiment, read
DOWN a column to compare every mouse on a given day. A header row labels each
session's phase (Ramp / Flat / Pillar), so you can see where the protocol changed
and how performance moved with it.

This rebuilds that layout from the long data:

    <metric>   | 2025-07-31 | 2025-08-01 | 2025-08-04 | ...
               | Flat       | Flat       | Flat       | ...
    CNT_05_01  | 80.0       | 76.2       | 62.5       | ...
    ...
    AVG        | ...
    AVG w/o .. | ...        (optional exclusions, as the old sheets did)

One block per metric, matching the old sheets' Retrieved + Contacted blocks.

Values are rounded -- the long tab carries raw floats like 63.7499999999, which
add noise to a table meant to be read at a glance.

The source workbook is never modified; output goes to a new file.

ASCII-only console output (Windows cp1252).
"""
from __future__ import annotations

import argparse
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Sequence

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover
    openpyxl = None

SOURCE_TAB = "3c_Manual_Summary"

# Metric blocks to emit, in order: (display name, column in the long tab).
DEFAULT_METRICS = [
    ("Retrieved", "Retrieved_Pct"),
    ("Contacted", "Contacted_Pct"),
    ("Missed", "Miss_Pct"),
    ("Displaced", "Displaced_Pct"),
]

_DATE_COL, _ANIMAL_COL, _PHASE_COL = "Date", "Animal", "Test_Phase"


def _read_long(path: Path, tab: str = SOURCE_TAB) -> List[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if tab not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"[FAIL] '{tab}' not found in {path.name}. Tabs: {wb.sheetnames}")
    ws = wb[tab]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    out = []
    for r in rows:
        rec = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        if rec.get(_ANIMAL_COL):
            out.append(rec)
    wb.close()
    return out


def _fmt_date(v) -> str:
    try:
        return v.strftime("%Y-%m-%d")
    except AttributeError:
        return str(v)[:10]


def build_matrix(records: Sequence[dict], metric_col: str):
    """-> (sessions, phases, {animal: {session: value}}). Sessions are ordered."""
    sessions: "OrderedDict[str, str]" = OrderedDict()
    table: Dict[str, Dict[str, object]] = defaultdict(dict)
    for rec in records:
        day = _fmt_date(rec.get(_DATE_COL))
        if not day:
            continue
        sessions.setdefault(day, str(rec.get(_PHASE_COL) or ""))
        val = rec.get(metric_col)
        if val is not None:
            table[str(rec[_ANIMAL_COL])][day] = val
    return list(sessions.keys()), sessions, table


def write_matrix_workbook(records, out_path: Path, metrics=None,
                          exclude: Optional[Sequence[str]] = None,
                          cohort_label: str = "") -> Path:
    metrics = metrics or DEFAULT_METRICS
    exclude = [e.strip() for e in (exclude or []) if e.strip()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manual Summary"

    bold = Font(bold=True)
    phase_font = Font(bold=True, italic=True, size=9)
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    avg_fill = PatternFill("solid", fgColor="F2F2F2")
    centre = Alignment(horizontal="center")

    present = {m for _, m in metrics if any(m in r for r in records)}
    row = 1
    for name, col in metrics:
        if col not in present:
            continue
        sessions, phases, table = build_matrix(records, col)
        if not sessions:
            continue

        c = ws.cell(row=row, column=1, value=f"{name}{(' -- ' + cohort_label) if cohort_label else ''}")
        c.font = bold
        c.fill = head_fill
        for j, day in enumerate(sessions, start=2):
            h = ws.cell(row=row, column=j, value=day)
            h.font = bold
            h.fill = head_fill
            h.alignment = centre
        row += 1

        # Phase row -- shows where the protocol changed.
        ws.cell(row=row, column=1, value="Phase").font = phase_font
        for j, day in enumerate(sessions, start=2):
            p = ws.cell(row=row, column=j, value=phases[day])
            p.font = phase_font
            p.alignment = centre
        row += 1

        animals = sorted(table)
        for animal in animals:
            ws.cell(row=row, column=1, value=animal).font = bold
            for j, day in enumerate(sessions, start=2):
                v = table[animal].get(day)
                cell = ws.cell(row=row, column=j,
                               value=round(float(v), 1) if isinstance(v, (int, float)) else v)
                cell.alignment = centre
            row += 1

        def _avg_row(label: str, members: List[str]):
            nonlocal row
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = bold
            lc.fill = avg_fill
            for j, day in enumerate(sessions, start=2):
                vals = [table[a][day] for a in members
                        if isinstance(table[a].get(day), (int, float))]
                cell = ws.cell(row=row, column=j,
                               value=round(sum(vals) / len(vals), 1) if vals else None)
                cell.font = bold
                cell.fill = avg_fill
                cell.alignment = centre
            row += 1

        _avg_row("AVG", animals)
        if exclude:
            kept = [a for a in animals if not any(a.endswith(e) or a == e for e in exclude)]
            if kept and len(kept) != len(animals):
                _avg_row("AVG w/o " + ", ".join(exclude), kept)
        row += 2  # blank rows between metric blocks

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 22
    for j in range(2, ws.max_column + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 11

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        prog="mousedb-summary-matrix",
        description=(
            "Rebuild a cohort's Manual Summary as a readable matrix (animals as rows, "
            "sessions as columns, phase header) instead of the long one-row-per-"
            "animal-per-session layout. Reads the tracking workbook; never modifies it."),
        epilog="Example:\n"
               "  mousedb-summary-matrix --source Connectome_05_Animal_Tracking.xlsx --exclude 14,15\n",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--source", required=True, help="Cohort tracking .xlsx to read.")
    parser.add_argument("--out", default=None,
                        help="Output .xlsx (default: alongside the source, '<name>_summary_matrix.xlsx').")
    parser.add_argument("--tab", default=SOURCE_TAB, help=f"Source tab (default: {SOURCE_TAB}).")
    parser.add_argument("--exclude", default="",
                        help="Comma-separated animal suffixes to add an 'AVG w/o' row for (e.g. 14,15).")
    args = parser.parse_args(argv)

    if openpyxl is None:
        print("[FAIL] openpyxl is required.")
        return 1

    src = Path(args.source)
    if not src.exists():
        print(f"[FAIL] Source not found: {src}")
        return 1
    out = Path(args.out) if args.out else src.with_name(src.stem + "_summary_matrix.xlsx")

    records = _read_long(src, args.tab)
    if not records:
        print(f"[FAIL] No data rows in '{args.tab}'.")
        return 1

    cohort = src.stem.replace("_Animal_Tracking", "")
    written = write_matrix_workbook(
        records, out, exclude=[e for e in args.exclude.split(",") if e.strip()],
        cohort_label=cohort)
    animals = {r.get(_ANIMAL_COL) for r in records}
    sessions = {_fmt_date(r.get(_DATE_COL)) for r in records}
    print(f"[OK] {len(records)} rows -> {len(animals)} animals x {len(sessions)} sessions")
    print(f"Written: {written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
