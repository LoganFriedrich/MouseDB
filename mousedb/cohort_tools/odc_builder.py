"""Populate a cohort's ODC-SCI tracking tab from the rest of its workbook.

Why
---
``2_ODC_Animal_Tracking`` is the 210-column standardised format everything will
eventually be reported in. It has never been filled in for ANY cohort -- 210
columns by hand was never going to happen. But almost every value already exists
in the other tabs, so it can be derived instead of typed.

Filling it is also a useful forcing function: the ODC row IS "everything about one
mouse at one timepoint", so generating it surfaces exactly which fields are
missing, and which are legitimately empty because that phase has not happened yet.

Shape
-----
One row per (subject x session). Subject-level fields repeat on each of that
subject's rows; per-session fields vary:

  cols 1-60    subject / contusion / injection / perfusion+BrainGlobe
  cols 61-67   Date, Test_Phase, Days_Post_Injury, Tray_Type, Weight, ...
  cols 68-163  Tray1-4 x Pellet01-20   (the raw per-pellet outcomes)
  cols 164-200 per-tray and total aggregates, Avg/Max/Min
  cols 201-204 provenance: Source_File, Source_Sheet, Row_Notes, Data_Context
  cols 205-210 Ladder

Sources
-------
  0a_Metadata               subject identity
  3e_Summary                contusion surgery + dosing
  5_SC_Injection_Details    spinal cord injection
  3b_Manual_Tray            per-pellet outcomes (one row per animal/date/tray)
  3d_Weights                session weights
  6_Ladder                  ladder test

Pellet codes in 3b_Manual_Tray: 0 = missed, 1 = displaced, 2 = retrieved.
Contacted = displaced + retrieved. Verified against 3c_Manual_Summary totals.

Phases that have not happened yet (injection, perfusion, BrainGlobe for a cohort
still in rehab) come out EMPTY rather than zero -- an empty cell means "not due
yet", and zero would be a lie.

The source workbook is never modified; output goes to a new file for review.

ASCII-only console output (Windows cp1252).
"""
from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover
    openpyxl = None

ODC_TAB = "2_ODC_Animal_Tracking"
MISSED, DISPLACED, RETRIEVED = 0, 1, 2
MAX_TRAYS, PELLETS_PER_TRAY = 4, 20


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _as_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(str(v).strip()[:10], fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _sheet_records(wb, tab: str) -> List[dict]:
    """Rows of a tab as dicts keyed by header. Empty list if the tab is absent."""
    if tab not in wb.sheetnames:
        return []
    ws = wb[tab]
    it = ws.iter_rows(values_only=True)
    try:
        header = [_norm(h) for h in next(it)]
    except StopIteration:
        return []
    out = []
    for r in it:
        rec = {header[i]: r[i] for i in range(min(len(header), len(r))) if header[i]}
        if any(v is not None and _norm(v) for v in rec.values()):
            out.append(rec)
    return out


_IDENTITY_KEYS = ("SubjectID", "Subject_ID", "Animal", "Subject")

# A column that is filled in on a blank form to show the expected shape, rather
# than because it records anything: the subject list, a placeholder date, and a
# constant label naming the procedure.
_FORM_SCAFFOLD = _IDENTITY_KEYS + ("Surgery_Date", "Surgery_Type", "Date", "Type")


def is_form_tab(records: List[dict]) -> bool:
    """True if a tab is a blank FORM rather than recorded data.

    Some tabs are kept as templates: carried from an older sheet, pre-filled to
    show whoever enters the real data what shape it takes.
    ``5_SC_Injection_Details`` is one -- every row reads Subject_ID +
    2026-02-08 + "SC Injection" + the standard anesthetic/analgesic, six months
    before that surgery is due, with every per-animal field blank.

    Ingesting those as records asserts things that never happened -- an injection
    date, in a data submission.

    Detecting them by "no payload is filled" does NOT work, because a form
    legitimately carries protocol CONSTANTS: the drugs and doses are known in
    advance, so they are pre-filled. The reliable tell is VARIATION. Real records
    differ between animals somewhere -- weights, volumes, outcomes -- because
    animals differ. A form is identical on every row but the subject id.

    So: a tab is a form if no non-identity column holds more than one distinct
    value across its rows.

    NOTE -- this is a WARNING signal only, never an action. It cannot be trusted
    to decide anything, because a form and a genuinely-uniform record are
    structurally identical: 0a_Metadata is uniform too (one litter, all male, one
    cohort, nobody dead yet) and is entirely real. Nothing in the file
    distinguishes "uniform because template" from "uniform because true".

    Which tabs are forms is knowledge the lab has and the workbook does not, so it
    is passed in explicitly (``form_tabs``). This function only flags candidates
    worth asking about.
    """
    if len(records) < 2:
        return False   # cannot judge variation from a single row
    seen: Dict[str, set] = defaultdict(set)
    for rec in records:
        for key, val in rec.items():
            if key in _IDENTITY_KEYS:
                continue
            if val is not None and _norm(val):
                seen[key].add(_norm(val))
    if not seen:
        return True                                   # nothing filled at all
    return all(len(vals) <= 1 for vals in seen.values())


def _index_by_subject(records: List[dict], keys=_IDENTITY_KEYS) -> Dict[str, dict]:
    """Index records by subject. Form tabs are excluded by the CALLER, explicitly
    -- never inferred here, since the inference is not reliable enough to act on."""
    idx = {}
    for rec in records:
        for k in keys:
            if rec.get(k):
                idx.setdefault(_norm(rec[k]), rec)
                break
    return idx


def collect_sessions(tray_rows: List[dict]) -> Dict[tuple, dict]:
    """Group 3b_Manual_Tray rows into sessions keyed by (subject, date).

    Each session gathers up to 4 trays; each tray keeps its 20 pellet codes in
    order, plus the tray's type letter taken from 'Tray Type/Number' (F1 -> F).
    """
    sessions: Dict[tuple, dict] = {}
    for row in tray_rows:
        subject = _norm(row.get("Animal"))
        day = _as_date(row.get("Date"))
        if not subject or day is None:
            continue
        key = (subject, day)
        s = sessions.setdefault(key, {
            "subject": subject, "date": day,
            "test_phase": _norm(row.get("Test_Phase")),
            "weight": row.get("Weight"), "weight_pct": row.get("Weight %"),
            "sex": _norm(row.get("Sex")), "trays": {}, "notes": [],
        })
        label = _norm(row.get("Tray Type/Number"))          # e.g. 'F1', 'P3'
        tray_no = None
        for ch in label:
            if ch.isdigit():
                tray_no = int(ch)
                break
        if tray_no is None or not 1 <= tray_no <= MAX_TRAYS:
            tray_no = len(s["trays"]) + 1
        pellets = []
        for i in range(1, PELLETS_PER_TRAY + 1):
            v = row.get(str(i))
            pellets.append(int(v) if isinstance(v, (int, float)) else None)
        s["trays"][tray_no] = {"type": label[:1].upper() if label else "", "pellets": pellets}
        if _norm(row.get("Notes")):
            s["notes"].append(_norm(row["Notes"]))
    return sessions


def _tray_stats(pellets: List[Optional[int]]) -> dict:
    scored = [p for p in pellets if p is not None]
    presented = len(scored)
    missed = sum(1 for p in scored if p == MISSED)
    displaced = sum(1 for p in scored if p == DISPLACED)
    retrieved = sum(1 for p in scored if p == RETRIEVED)
    contacted = displaced + retrieved
    pct = lambda n: round(100.0 * n / presented, 2) if presented else None
    return {"presented": presented, "missed": missed, "displaced": displaced,
            "retrieved": retrieved, "contacted": contacted,
            "miss_pct": pct(missed), "displaced_pct": pct(displaced),
            "retrieved_pct": pct(retrieved), "contacted_pct": pct(contacted)}


STUDY_CONSTANTS = {
    "SpeciesTyp": "mouse",
    "SpeciesStrainTyp": "C57BL/6J",
    "AnimalSourceNam": "Jackson Labs",
    "StudyLeader": "",          # see --study-leader; left blank rather than guessed
}


def build_rows(wb, source_name: str, form_tabs=(), constants=None) -> List[dict]:
    """Build ODC records (header -> value) for every subject-session.

    ``form_tabs`` names tabs that are blank TEMPLATES rather than records; they
    are read as empty, so their placeholder values never become assertions.
    """
    forms = {t.strip() for t in form_tabs if t and t.strip()}

    def _src(tab: str) -> List[dict]:
        return [] if tab in forms else _sheet_records(wb, tab)

    meta = _index_by_subject(_src("0a_Metadata"))
    contusion = _index_by_subject(_src("3e_Summary"))
    if not contusion:
        contusion = _index_by_subject(_src("4_Contusion_Injury_Details"))
    injection = _index_by_subject(_src("5_SC_Injection_Details"))
    ladder_rows = _src("6_Ladder")
    ladder = _index_by_subject(ladder_rows)

    sessions = collect_sessions(_sheet_records(wb, "3b_Manual_Tray"))

    # Injury date per subject -> Days_Post_Injury. Blank before injury: DPI is
    # only meaningful once the injury has happened.
    injury_date = {s: _as_date(rec.get("Surgery_Date") or rec.get("Contusion_Date") or rec.get("Date"))
                   for s, rec in contusion.items()}

    out = []
    for (subject, day) in sorted(sessions):
        s = sessions[(subject, day)]
        m, c, inj, lad = meta.get(subject, {}), contusion.get(subject, {}), injection.get(subject, {}), ladder.get(subject, {})
        row: Dict[str, object] = {}

        row["SubjectID"] = subject
        row["SexTyp"] = s["sex"] or _norm(m.get("Sex"))
        row["Laboratory"] = "LAB"
        # Study-wide constants. Identical for every animal in every cohort, so they
        # are configuration rather than data -- and were empty across all five
        # cohorts simply because nobody had anywhere to put them.
        for k, v in (constants or {}).items():
            if v:
                row[k] = v
        if m.get("Cohort") is not None:
            row["InjGroupAssignTyp"] = m["Cohort"]
        # AgeVal is derived: ODC wants an age at the session, not a birth date.
        dob = _as_date(m.get("Date_of_Birth"))
        if dob:
            row["AgeVal"] = (day - dob).days
        # Cause_of_Death is a CAUSE; Date_of_Death is a date. Do not conflate them.

        # --- contusion block (3e_Summary) ---
        for src, dst in (
            ("Surgery_Date", "Contusion_Date"), ("Surgery_Type", "Contusion_Type"),
            ("Surgery_Severity", "Contusion_Severity"), ("Contusion_Location", "Contusion_Location"),
            ("Subject_Weight (g)", "Contusion_Weight_g"),
            ("Anesthetic", "Contusion_Anesthetic"), ("Anesthetic_Dose", "Contusion_Anesthetic_Dose"),
            ("Anesthetic_Volume", "Contusion_Anesthetic_Volume"),
            ("Analgesic", "Contusion_Analgesic"), ("Analgesic_Dose", "Contusion_Analgesic_Dose"),
            ("Analgesic_Volume", "Contusion_Analgesic_Volume"),
            ("Intended_kd", "Contusion_Intended_kd"), ("Intended_Dwell", "Contusion_Intended_Dwell"),
            ("Stage_Height", "Contusion_Stage_Height"), ("Actual_kd", "Contusion_Actual_kd"),
            ("Actual_displacement", "Contusion_Actual_Displacement"),
            ("Actual_Velocity", "Contusion_Actual_Velocity"), ("Actual_Dwell", "Contusion_Actual_Dwell"),
            ("Survived", "Contusion_Survived"),
        ):
            if c.get(src) is not None:
                row[dst] = c[src]
        if c:
            row["Injury_type"] = c.get("Surgery_Type") or "Contusion"
            row["Injury_level"] = c.get("Contusion_Location")
            row["Injury_details"] = c.get("Notes")

        # --- injection block (5_SC_Injection_Details) ---
        for src, dst in (
            ("Surgery_Date", "Injection_Date"), ("Subject_Weight (g)", "Injection_Weight_g"),
            ("Surgery_Type", "Injection_Type"), ("Injected_Virus", "Injection_Virus"),
            ("Virus_Titer", "Injection_Titer"), ("Injection_Target", "Injection_Target"),
            ("Depths (D/V)", "Injection_Depth_DV"), ("Coordinates (M/L)", "Injection_Coord_ML"),
            ("Anesthetic", "Injection_Anesthetic"), ("Anesthetic_Dose", "Injection_Anesthetic_Dose"),
            ("Anesthetic_Volume", "Injection_Anesthetic_Volume"),
            ("Analgesic", "Injection_Analgesic"), ("Analgesic_Dose", "Injection_Analgesic_Dose"),
            ("Analgesic_Volume", "Injection_Analgesic_Volume"),
            ("Survived", "Injection_Survived"),
            ("Signal Post Perfusion", "Injection_Signal_Post_Perfusion"),
        ):
            if inj.get(src) is not None:
                row[dst] = inj[src]

        # --- per-session ---
        row["Date"] = day
        row["Test_Phase"] = s["test_phase"]
        idate = injury_date.get(subject)
        row["Days_Post_Injury"] = (day - idate).days if idate and day >= idate else None
        types = {t["type"] for t in s["trays"].values() if t["type"]}
        row["Tray_Type"] = "/".join(sorted(types)) if types else None
        row["Weight"] = s["weight"]
        row["BodyWgtMeasrVal"] = s["weight"]   # ODC's body-weight field for this session
        row["Weight_Pct"] = round(s["weight_pct"] * 100, 1) if isinstance(s["weight_pct"], float) and s["weight_pct"] <= 1 else s["weight_pct"]
        row["Num_Trays"] = len(s["trays"])

        totals = defaultdict(int)
        pcts = defaultdict(list)
        for n in range(1, MAX_TRAYS + 1):
            tray = s["trays"].get(n)
            if not tray:
                continue
            for i, code in enumerate(tray["pellets"], start=1):
                row[f"Tray{n}_Pellet{i:02d}"] = code
            st = _tray_stats(tray["pellets"])
            for k, v in st.items():
                if k.endswith("_pct"):
                    row[f"Tray{n}_{k.replace('_pct','').title()}_Pct"] = v
                    if v is not None:
                        pcts[k].append(v)
                else:
                    row[f"Tray{n}_{k.title()}"] = v
                    totals[k] += v

        for k in ("presented", "missed", "displaced", "retrieved", "contacted"):
            row[f"Total_{k.title()}"] = totals[k]
        tp = totals["presented"]
        for k in ("missed", "displaced", "retrieved", "contacted"):
            label = "Miss" if k == "missed" else k.title()
            row[f"Total_{label}_Pct"] = round(100.0 * totals[k] / tp, 2) if tp else None
        for k, vals in pcts.items():
            label = "Miss" if k == "miss_pct" else k.replace("_pct", "").title()
            row[f"Avg_{label}_Pct"] = round(sum(vals) / len(vals), 2) if vals else None
        for k, label in (("retrieved_pct", "Retrieved"), ("contacted_pct", "Contacted")):
            if pcts.get(k):
                row[f"Max_{label}_Pct"] = max(pcts[k])
                row[f"Min_{label}_Pct"] = min(pcts[k])

        # --- ladder + provenance ---
        for src, dst in (("Date", "Ladder_Date"), ("Total_Steps", "Ladder_Total_Steps"),
                         ("Total_Misses", "Ladder_Total_Misses"), ("Miss_Pct", "Ladder_Miss_Pct")):
            if lad.get(src) is not None:
                row[dst] = lad[src]

        row["Source_File"] = source_name
        row["Source_Sheet"] = "3b_Manual_Tray"
        row["Row_Notes"] = "; ".join(s["notes"]) or None
        row["Data_Context"] = f"generated by mousedb odc_builder from {source_name}"
        out.append(row)
    return out


def write_odc(headers: List[str], rows: List[dict], out_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ODC_TAB[:31]
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")
    for i, rec in enumerate(rows, start=2):
        for j, h in enumerate(headers, start=1):
            v = rec.get(h)
            if v is not None:
                ws.cell(row=i, column=j, value=v)
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 14
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        prog="mousedb-odc-build",
        description=(
            "Populate a cohort's 2_ODC_Animal_Tracking tab by deriving it from the "
            "workbook's other tabs. Reads only; writes a new file for review. "
            "Fields whose phase has not happened yet are left EMPTY, not zero."),
        epilog="Example:\n  mousedb-odc-build --source Connectome_05_Animal_Tracking.xlsx\n",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--source", required=True, help="Cohort tracking .xlsx.")
    parser.add_argument("--out", default=None, help="Output .xlsx (default: <name>_ODC.xlsx alongside).")
    parser.add_argument("--study-leader", default="",
                        help="ODC StudyLeader for every row (e.g. the lab PI).")
    parser.add_argument("--species", default=STUDY_CONSTANTS["SpeciesTyp"])
    parser.add_argument("--strain", default=STUDY_CONSTANTS["SpeciesStrainTyp"])
    parser.add_argument("--source-name", default=STUDY_CONSTANTS["AnimalSourceNam"],
                        help="Animal supplier (AnimalSourceNam).")
    parser.add_argument(
        "--form-tabs", default="",
        help=("Comma-separated tabs that are blank TEMPLATES, not records "
              "(e.g. 5_SC_Injection_Details). Their placeholder values are ignored "
              "rather than emitted as fact. Candidates are reported below; which "
              "ones are really forms cannot be told from the file, so it must be "
              "stated here."))
    args = parser.parse_args(argv)

    if openpyxl is None:
        print("[FAIL] openpyxl is required.")
        return 1
    src = Path(args.source)
    if not src.exists():
        print(f"[FAIL] Source not found: {src}")
        return 1

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if ODC_TAB not in wb.sheetnames:
        print(f"[FAIL] '{ODC_TAB}' not in {src.name}")
        return 1
    headers = [_norm(h) for h in next(wb[ODC_TAB].iter_rows(min_row=1, max_row=1, values_only=True)) if h is not None]
    form_tabs = [t for t in args.form_tabs.split(",") if t.strip()]
    candidates = [t for t in wb.sheetnames
                  if t not in form_tabs and is_form_tab(_sheet_records(wb, t))]
    constants = {"SpeciesTyp": args.species, "SpeciesStrainTyp": args.strain,
                 "AnimalSourceNam": args.source_name, "StudyLeader": args.study_leader}
    rows = build_rows(wb, src.name, form_tabs=form_tabs, constants=constants)
    wb.close()

    if not rows:
        print("[FAIL] No sessions found (is 3b_Manual_Tray populated?).")
        return 1

    out = Path(args.out) if args.out else src.with_name(src.stem + "_ODC.xlsx")
    write_odc(headers, rows, out)

    filled = {h for r in rows for h, v in r.items() if v is not None}
    known = [h for h in headers if h in filled]
    empty = [h for h in headers if h not in filled]
    subjects = {r["SubjectID"] for r in rows}
    print(f"[OK] {len(rows)} rows ({len(subjects)} subjects x sessions)")
    print(f"     columns populated: {len(known)}/{len(headers)}")
    print(f"Written: {out}")
    if form_tabs:
        print(f"     treated as FORMS (ignored): {', '.join(form_tabs)}")
    if candidates:
        print()
        print("[!] Tabs with NO per-animal variation -- possibly blank forms whose")
        print("    placeholder values are being emitted as if they were recorded:")
        for t in candidates:
            print(f"      {t}")
        print("    A genuinely uniform tab looks identical, so this cannot be decided")
        print("    from the file. Pass real ones via --form-tabs.")
    print()
    print("Columns still EMPTY (phase not reached, or no source):")
    for h in empty[:40]:
        print(f"   {h}")
    if len(empty) > 40:
        print(f"   ... and {len(empty)-40} more")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
