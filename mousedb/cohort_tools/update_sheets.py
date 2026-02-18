#!/usr/bin/env python3
"""
Script 1: 1_Update.py

PURPOSE: Populate 2_ODC_Animal_Tracking sheet from source data sheets.

THIS SCRIPT DOES:
- Read 3b_Manual_Tray (pellet scores - one row per tray)
- Group tray rows by (Animal, Date) to create one ODC row per day
- Read 4_Contusion_Injury_Details (surgery data)
- Read 5_SC_Injection_Details (tracing data)
- Read 0a_Metadata (DOB, DOD, sex)
- Read 3a_Manual_Ramp (baseline weights)
- Calculate all derived values (percentages, totals, days post-injury)
- Handle death exclusions
- Populate 2_ODC_Animal_Tracking with ~203 columns
- Update 3c_Manual_Summary
- Update 7_Stats

Usage:
    python 1_Update.py Connectome_1_Animal_Tracking.xlsx
    python 1_Update.py --all  # Process all cohort files in directory
"""

import os
import sys
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

SCRIPT_DIR = Path.cwd()
OUTPUT_SUBDIR = SCRIPT_DIR / "processed"

# ODC column structure (203 columns)
ODC_COLUMNS = [
    # Section 1: ODC-SCI Required CoDEs (17)
    "SubjectID", "SpeciesTyp", "SpeciesStrainTyp", "AnimalSourceNam",
    "AgeVal", "BodyWgtMeasrVal", "SexTyp", "InjGroupAssignTyp",
    "Laboratory", "StudyLeader", "Exclusion_in_origin_study", "Exclusion_reason",
    "Cause_of_Death", "Injury_type", "Injury_device", "Injury_level", "Injury_details",
    
    # Section 2: Contusion Surgery Details (19)
    "Contusion_Date", "Contusion_Type", "Contusion_Severity", "Contusion_Location",
    "Contusion_Weight_g", "Contusion_Anesthetic", "Contusion_Anesthetic_Dose",
    "Contusion_Anesthetic_Volume", "Contusion_Analgesic", "Contusion_Analgesic_Dose",
    "Contusion_Analgesic_Volume", "Contusion_Intended_kd", "Contusion_Intended_Dwell",
    "Contusion_Stage_Height", "Contusion_Actual_kd", "Contusion_Actual_Displacement",
    "Contusion_Actual_Velocity", "Contusion_Actual_Dwell", "Contusion_Survived",
    
    # Section 3: SC Injection Details (16)
    "Injection_Date", "Injection_Weight_g", "Injection_Type", "Injection_Virus",
    "Injection_Titer", "Injection_Target", "Injection_Depth_DV", "Injection_Coord_ML",
    "Injection_Anesthetic", "Injection_Anesthetic_Dose", "Injection_Anesthetic_Volume",
    "Injection_Analgesic", "Injection_Analgesic_Dose", "Injection_Analgesic_Volume",
    "Injection_Survived", "Injection_Signal_Post_Perfusion",
    
    # Section 4: BrainGlobe Placeholders (8)
    "Perfusion_Date", "BrainGlobe_Analysis_Date", "BrainGlobe_Atlas_Used",
    "Total_Cells_Detected", "Total_Cells_Left_Hemisphere", "Total_Cells_Right_Hemisphere",
    "BrainGlobe_Notes", "BrainGlobe_Quality",
    
    # Section 5: Row-Level Metadata (7)
    "Date", "Test_Phase", "Days_Post_Injury", "Tray_Type", "Weight", "Weight_Pct", "Num_Trays",
]

# Add per-pellet columns (80 = 4 trays × 20 pellets)
for tray in range(1, 5):
    for pellet in range(1, 21):
        ODC_COLUMNS.append(f"Tray{tray}_Pellet{pellet:02d}")

# Add per-tray calculations (36 = 4 trays × 9 metrics)
for tray in range(1, 5):
    ODC_COLUMNS.extend([
        f"Tray{tray}_Presented", f"Tray{tray}_Missed", f"Tray{tray}_Displaced",
        f"Tray{tray}_Retrieved", f"Tray{tray}_Contacted",
        f"Tray{tray}_Miss_Pct", f"Tray{tray}_Displaced_Pct",
        f"Tray{tray}_Retrieved_Pct", f"Tray{tray}_Contacted_Pct"
    ])

# Add daily totals and averages (17)
ODC_COLUMNS.extend([
    "Total_Presented", "Total_Missed", "Total_Displaced", "Total_Retrieved", "Total_Contacted",
    "Total_Miss_Pct", "Total_Displaced_Pct", "Total_Retrieved_Pct", "Total_Contacted_Pct",
    "Avg_Miss_Pct", "Avg_Displaced_Pct", "Avg_Retrieved_Pct", "Avg_Contacted_Pct",
    "Max_Retrieved_Pct", "Max_Contacted_Pct", "Min_Retrieved_Pct", "Min_Contacted_Pct"
])

# Add source tracking (3)
ODC_COLUMNS.extend(["Source_File", "Source_Sheet", "Row_Notes"])

# Create column index lookup
ODC_COL_IDX = {name: i for i, name in enumerate(ODC_COLUMNS)}

# Hard-coded values
PROJECT_DEFAULTS = {
    "SpeciesTyp": "Mouse",
    "SpeciesStrainTyp": "C57BL/6J",
    "AnimalSourceNam": "Jackson Laboratory",
    "Laboratory": "Murray/Blackmore Lab",
    "StudyLeader": "Adam Murray",
    "Injury_device": "Infinite Horizon Impactor",
}


# =============================================================================
# DATA EXTRACTION FUNCTIONS
# =============================================================================

def find_column(ws, possible_names, max_col=20):
    """Find column index by checking header row for possible names."""
    for col in range(1, min(max_col, ws.max_column + 1)):
        header = ws.cell(1, col).value
        if header:
            header_lower = str(header).strip().lower()
            for name in possible_names:
                if name.lower() == header_lower:
                    return col
    return None


def extract_metadata(wb):
    """Extract metadata from 0a_Metadata sheet."""
    metadata = {}  # animal_id -> {dob, dod, sex, cohort}
    
    if '0a_Metadata' not in wb.sheetnames:
        return metadata
    
    ws = wb['0a_Metadata']
    
    # Find columns
    id_col = find_column(ws, ['SubjectID', 'Subject_ID', 'Animal', 'Mouse'])
    dob_col = find_column(ws, ['Date_of_Birth', 'DOB', 'Birth_Date'])
    dod_col = find_column(ws, ['Date_of_Death', 'DOD', 'Death_Date'])
    sex_col = find_column(ws, ['Sex', 'Gender'])
    cohort_col = find_column(ws, ['Cohort'])
    
    if not id_col:
        return metadata
    
    for row in range(2, ws.max_row + 1):
        animal_id = ws.cell(row, id_col).value
        if not animal_id:
            continue
        
        animal_id = str(animal_id).strip()
        metadata[animal_id] = {
            'dob': ws.cell(row, dob_col).value if dob_col else None,
            'dod': ws.cell(row, dod_col).value if dod_col else None,
            'sex': ws.cell(row, sex_col).value if sex_col else None,
            'cohort': ws.cell(row, cohort_col).value if cohort_col else None,
        }
    
    return metadata


def extract_baseline_weights(wb):
    """Extract baseline weights from 3a_Manual_Ramp sheet."""
    weights = {}  # animal_id -> baseline_weight
    
    if '3a_Manual_Ramp' not in wb.sheetnames:
        return weights
    
    ws = wb['3a_Manual_Ramp']
    
    # Find columns
    id_col = find_column(ws, ['Mouse ID', 'Animal', 'SubjectID', 'Subject_ID'])
    weight_col = find_column(ws, ['Weight'])
    
    if not id_col:
        return weights
    
    # Weight is typically in column D (index 4)
    if not weight_col:
        weight_col = 4
    
    for row in range(2, ws.max_row + 1):
        animal_id = ws.cell(row, id_col).value
        if not animal_id:
            continue
        
        animal_id = str(animal_id).strip()
        weight = ws.cell(row, weight_col).value
        if weight and isinstance(weight, (int, float)):
            weights[animal_id] = weight
    
    return weights


def extract_contusion_data(wb):
    """Extract contusion surgery data from 4_Contusion_Injury_Details."""
    contusion = {}  # animal_id -> {all fields}
    
    if '4_Contusion_Injury_Details' not in wb.sheetnames:
        return contusion
    
    ws = wb['4_Contusion_Injury_Details']
    
    # Build header map
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(1, col).value
        if h:
            headers[str(h).strip()] = col
    
    # Find ID column
    id_col = None
    for name in ['Subject_ID', 'SubjectID', 'Animal', 'Mouse']:
        if name in headers:
            id_col = headers[name]
            break
    
    if not id_col:
        return contusion
    
    # Field mappings
    field_map = {
        'Surgery_Date': 'date',
        'Surgery_Type': 'type',
        'Surgery_Severity': 'severity',
        'Contusion_Location': 'location',
        'Subject_Weight (g)': 'weight',
        'Anesthetic': 'anesthetic',
        'Anesthetic_Dose': 'anesthetic_dose',
        'Anesthetic_Volume': 'anesthetic_volume',
        'Analgesic': 'analgesic',
        'Analgesic_Dose': 'analgesic_dose',
        'Analgesic_Volume': 'analgesic_volume',
        'Intended_kd': 'intended_kd',
        'Intended_Dwell': 'intended_dwell',
        'Stage_Height': 'stage_height',
        'Actual_kd': 'actual_kd',
        'Actual_displacement': 'actual_displacement',
        'Actual_Velocity': 'actual_velocity',
        'Actual_Dwell': 'actual_dwell',
        'Survived': 'survived',
    }
    
    for row in range(2, ws.max_row + 1):
        animal_id = ws.cell(row, id_col).value
        if not animal_id:
            continue
        
        animal_id = str(animal_id).strip()
        contusion[animal_id] = {}
        
        for sheet_field, data_field in field_map.items():
            if sheet_field in headers:
                val = ws.cell(row, headers[sheet_field]).value
                contusion[animal_id][data_field] = val
    
    return contusion


def extract_injection_data(wb):
    """Extract SC injection data from 5_SC_Injection_Details."""
    injection = {}  # animal_id -> {all fields}
    
    if '5_SC_Injection_Details' not in wb.sheetnames:
        return injection
    
    ws = wb['5_SC_Injection_Details']
    
    # Build header map
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(1, col).value
        if h:
            headers[str(h).strip()] = col
    
    # Find ID column
    id_col = None
    for name in ['Subject_ID', 'SubjectID', 'Animal', 'Mouse']:
        if name in headers:
            id_col = headers[name]
            break
    
    if not id_col:
        return injection
    
    # Field mappings
    field_map = {
        'Surgery_Date': 'date',
        'Subject_Weight (g)': 'weight',
        'Surgery_Type': 'type',
        'Injected_Virus': 'virus',
        'Virus_Titer': 'titer',
        'Injection_Target': 'target',
        'Depths (D/V)': 'depth_dv',
        'Coordinates (M/L)': 'coord_ml',
        'Anesthetic': 'anesthetic',
        'Anesthetic_Dose': 'anesthetic_dose',
        'Anesthetic_Volume': 'anesthetic_volume',
        'Analgesic': 'analgesic',
        'Analgesic_Dose': 'analgesic_dose',
        'Analgesic_Volume': 'analgesic_volume',
        'Survived': 'survived',
        'Signal Post Perfusion': 'signal_post_perfusion',
    }
    
    for row in range(2, ws.max_row + 1):
        animal_id = ws.cell(row, id_col).value
        if not animal_id:
            continue
        
        animal_id = str(animal_id).strip()
        injection[animal_id] = {}
        
        for sheet_field, data_field in field_map.items():
            if sheet_field in headers:
                val = ws.cell(row, headers[sheet_field]).value
                injection[animal_id][data_field] = val
    
    return injection


def extract_tray_data(wb):
    """
    Extract tray data from 3b_Manual_Tray and group by (Animal, Date).
    
    Returns:
        dict: {(animal_id, date): {'phase': str, 'weight': float, 'trays': {label: [scores]}}}
    """
    grouped = defaultdict(lambda: {'phase': None, 'weight': None, 'sex': None, 'trays': {}})
    
    if '3b_Manual_Tray' not in wb.sheetnames:
        return grouped
    
    ws = wb['3b_Manual_Tray']
    
    # Find columns dynamically
    date_col = find_column(ws, ['Date'])
    animal_col = find_column(ws, ['Animal', 'Mouse', 'Subject', 'SubjectID', 'Mouse ID'])
    sex_col = find_column(ws, ['Sex'])
    weight_col = find_column(ws, ['Weight'])
    phase_col = find_column(ws, ['Test_Phase', 'Phase'])
    tray_col = find_column(ws, ['Tray Type/Number', 'Tray', 'Tray_Type'])
    
    # Find first pellet column (should be labeled "1")
    pellet_start = None
    for col in range(1, min(30, ws.max_column + 1)):
        h = ws.cell(1, col).value
        if h and str(h).strip() == '1':
            pellet_start = col
            break
    
    if not all([date_col, animal_col, tray_col, pellet_start]):
        print(f"  Warning: Could not find all required columns in 3b_Manual_Tray")
        print(f"    date_col={date_col}, animal_col={animal_col}, tray_col={tray_col}, pellet_start={pellet_start}")
        return grouped
    
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row, date_col).value
        animal_id = ws.cell(row, animal_col).value
        
        if not date_val or not animal_id:
            continue
        
        animal_id = str(animal_id).strip()
        
        # Normalize date
        if isinstance(date_val, datetime):
            date_key = date_val.date()
        elif hasattr(date_val, 'date'):
            date_key = date_val
        else:
            continue
        
        key = (animal_id, date_key)
        
        # Get phase, weight, sex
        if phase_col:
            phase = ws.cell(row, phase_col).value
            if phase:
                grouped[key]['phase'] = str(phase).strip()
        
        if weight_col:
            weight = ws.cell(row, weight_col).value
            if weight and isinstance(weight, (int, float)):
                grouped[key]['weight'] = weight
        
        if sex_col:
            sex = ws.cell(row, sex_col).value
            if sex:
                grouped[key]['sex'] = str(sex).strip()
        
        # Get tray label and pellet scores
        tray_label = ws.cell(row, tray_col).value
        if not tray_label:
            continue
        
        tray_label = str(tray_label).strip()
        
        # Extract pellet scores (columns 1-20)
        scores = []
        for p in range(20):
            val = ws.cell(row, pellet_start + p).value
            if val is not None and val != '':
                try:
                    scores.append(int(val))
                except (ValueError, TypeError):
                    scores.append(None)
            else:
                scores.append(None)
        
        grouped[key]['trays'][tray_label] = scores
    
    return grouped


# =============================================================================
# ODC ROW COMPUTATION
# =============================================================================

def compute_tray_stats(scores):
    """Compute statistics for a single tray's pellet scores."""
    valid = [s for s in scores if s is not None]
    
    presented = len(valid)
    missed = sum(1 for s in valid if s == 0)
    displaced = sum(1 for s in valid if s == 1)
    retrieved = sum(1 for s in valid if s == 2)
    contacted = displaced + retrieved
    
    if presented > 0:
        miss_pct = round(missed / presented * 100, 2)
        displaced_pct = round(displaced / presented * 100, 2)
        retrieved_pct = round(retrieved / presented * 100, 2)
        contacted_pct = round(contacted / presented * 100, 2)
    else:
        miss_pct = displaced_pct = retrieved_pct = contacted_pct = None
    
    return {
        'presented': presented,
        'missed': missed,
        'displaced': displaced,
        'retrieved': retrieved,
        'contacted': contacted,
        'miss_pct': miss_pct,
        'displaced_pct': displaced_pct,
        'retrieved_pct': retrieved_pct,
        'contacted_pct': contacted_pct,
    }


def build_odc_row(animal_id, date, day_data, metadata, baseline_weights, contusion, injection, source_file):
    """
    Build a single ODC row from grouped day data.
    
    Returns:
        list: Values for all ODC columns
    """
    row = [None] * len(ODC_COLUMNS)
    
    # Get supporting data
    meta = metadata.get(animal_id, {})
    cont = contusion.get(animal_id, {})
    inj = injection.get(animal_id, {})
    baseline = baseline_weights.get(animal_id)
    
    # Infer cohort from animal_id
    cohort = animal_id.rsplit('_', 1)[0] if '_' in animal_id else 'Unknown'
    
    # === Section 1: ODC-SCI Required CoDEs ===
    row[ODC_COL_IDX['SubjectID']] = animal_id
    row[ODC_COL_IDX['SpeciesTyp']] = PROJECT_DEFAULTS['SpeciesTyp']
    row[ODC_COL_IDX['SpeciesStrainTyp']] = PROJECT_DEFAULTS['SpeciesStrainTyp']
    row[ODC_COL_IDX['AnimalSourceNam']] = PROJECT_DEFAULTS['AnimalSourceNam']
    
    # AgeVal - weeks from DOB
    dob = meta.get('dob')
    if dob and isinstance(dob, datetime):
        age_days = (datetime.combine(date, datetime.min.time()) - dob).days
        row[ODC_COL_IDX['AgeVal']] = round(age_days / 7, 2)
    
    row[ODC_COL_IDX['BodyWgtMeasrVal']] = baseline
    row[ODC_COL_IDX['SexTyp']] = meta.get('sex') or day_data.get('sex')
    row[ODC_COL_IDX['InjGroupAssignTyp']] = cohort
    row[ODC_COL_IDX['Laboratory']] = PROJECT_DEFAULTS['Laboratory']
    row[ODC_COL_IDX['StudyLeader']] = PROJECT_DEFAULTS['StudyLeader']
    
    # Exclusion handling
    survived = cont.get('survived', 'Y')
    if survived and str(survived).upper() == 'N':
        row[ODC_COL_IDX['Exclusion_in_origin_study']] = 'Total exclusion'
        row[ODC_COL_IDX['Exclusion_reason']] = 'Died during/after surgery'
        row[ODC_COL_IDX['Cause_of_Death']] = 'Surgery complications'
    else:
        row[ODC_COL_IDX['Exclusion_in_origin_study']] = 'No'
        row[ODC_COL_IDX['Exclusion_reason']] = ''
        row[ODC_COL_IDX['Cause_of_Death']] = 'Perfusion'
    
    # Injury details
    row[ODC_COL_IDX['Injury_type']] = cont.get('type', 'Contusion')
    row[ODC_COL_IDX['Injury_device']] = PROJECT_DEFAULTS['Injury_device']
    row[ODC_COL_IDX['Injury_level']] = cont.get('location', '')
    
    # Build injury details string
    details = []
    if cont.get('actual_kd'):
        details.append(f"{cont['actual_kd']} kdyn")
    if cont.get('actual_displacement'):
        details.append(f"{cont['actual_displacement']} µm displacement")
    row[ODC_COL_IDX['Injury_details']] = ', '.join(details) if details else ''
    
    # === Section 2: Contusion Surgery Details ===
    row[ODC_COL_IDX['Contusion_Date']] = cont.get('date')
    row[ODC_COL_IDX['Contusion_Type']] = cont.get('type')
    row[ODC_COL_IDX['Contusion_Severity']] = cont.get('severity')
    row[ODC_COL_IDX['Contusion_Location']] = cont.get('location')
    row[ODC_COL_IDX['Contusion_Weight_g']] = cont.get('weight')
    row[ODC_COL_IDX['Contusion_Anesthetic']] = cont.get('anesthetic')
    row[ODC_COL_IDX['Contusion_Anesthetic_Dose']] = cont.get('anesthetic_dose')
    row[ODC_COL_IDX['Contusion_Anesthetic_Volume']] = cont.get('anesthetic_volume')
    row[ODC_COL_IDX['Contusion_Analgesic']] = cont.get('analgesic')
    row[ODC_COL_IDX['Contusion_Analgesic_Dose']] = cont.get('analgesic_dose')
    row[ODC_COL_IDX['Contusion_Analgesic_Volume']] = cont.get('analgesic_volume')
    row[ODC_COL_IDX['Contusion_Intended_kd']] = cont.get('intended_kd')
    row[ODC_COL_IDX['Contusion_Intended_Dwell']] = cont.get('intended_dwell')
    row[ODC_COL_IDX['Contusion_Stage_Height']] = cont.get('stage_height')
    row[ODC_COL_IDX['Contusion_Actual_kd']] = cont.get('actual_kd')
    row[ODC_COL_IDX['Contusion_Actual_Displacement']] = cont.get('actual_displacement')
    row[ODC_COL_IDX['Contusion_Actual_Velocity']] = cont.get('actual_velocity')
    row[ODC_COL_IDX['Contusion_Actual_Dwell']] = cont.get('actual_dwell')
    row[ODC_COL_IDX['Contusion_Survived']] = cont.get('survived')
    
    # === Section 3: SC Injection Details ===
    row[ODC_COL_IDX['Injection_Date']] = inj.get('date')
    row[ODC_COL_IDX['Injection_Weight_g']] = inj.get('weight')
    row[ODC_COL_IDX['Injection_Type']] = inj.get('type')
    row[ODC_COL_IDX['Injection_Virus']] = inj.get('virus')
    row[ODC_COL_IDX['Injection_Titer']] = inj.get('titer')
    row[ODC_COL_IDX['Injection_Target']] = inj.get('target')
    row[ODC_COL_IDX['Injection_Depth_DV']] = inj.get('depth_dv')
    row[ODC_COL_IDX['Injection_Coord_ML']] = inj.get('coord_ml')
    row[ODC_COL_IDX['Injection_Anesthetic']] = inj.get('anesthetic')
    row[ODC_COL_IDX['Injection_Anesthetic_Dose']] = inj.get('anesthetic_dose')
    row[ODC_COL_IDX['Injection_Anesthetic_Volume']] = inj.get('anesthetic_volume')
    row[ODC_COL_IDX['Injection_Analgesic']] = inj.get('analgesic')
    row[ODC_COL_IDX['Injection_Analgesic_Dose']] = inj.get('analgesic_dose')
    row[ODC_COL_IDX['Injection_Analgesic_Volume']] = inj.get('analgesic_volume')
    row[ODC_COL_IDX['Injection_Survived']] = inj.get('survived')
    row[ODC_COL_IDX['Injection_Signal_Post_Perfusion']] = inj.get('signal_post_perfusion')
    
    # === Section 4: BrainGlobe (placeholders) ===
    # Left empty - populated by BrainGlobe import script
    
    # === Section 5: Row-Level Metadata ===
    row[ODC_COL_IDX['Date']] = date
    row[ODC_COL_IDX['Test_Phase']] = day_data.get('phase')
    
    # Days post-injury
    contusion_date = cont.get('date')
    if contusion_date:
        if isinstance(contusion_date, datetime):
            contusion_date = contusion_date.date()
        try:
            dpi = (date - contusion_date).days
            row[ODC_COL_IDX['Days_Post_Injury']] = dpi
        except:
            pass
    
    # Tray type (first letter of first tray label)
    trays = day_data.get('trays', {})
    if trays:
        first_label = list(trays.keys())[0]
        row[ODC_COL_IDX['Tray_Type']] = first_label[0] if first_label else None
    
    row[ODC_COL_IDX['Weight']] = day_data.get('weight')
    
    # Weight percentage
    if day_data.get('weight') and baseline:
        row[ODC_COL_IDX['Weight_Pct']] = round(day_data['weight'] / baseline, 3)
    
    row[ODC_COL_IDX['Num_Trays']] = len(trays)
    
    # === Section 6: Per-Pellet Scores ===
    # Map tray labels to tray numbers: F1/P1/E1 -> Tray1, F2/P2/E2 -> Tray2, etc.
    tray_stats = {}
    
    for tray_label, scores in trays.items():
        # Extract tray number from label (F1->1, P2->2, E3->3, etc.)
        try:
            tray_num = int(tray_label[1:])
        except (ValueError, IndexError):
            continue
        
        if not 1 <= tray_num <= 4:
            continue
        
        # Write pellet scores
        for pellet_idx, score in enumerate(scores):
            col_name = f"Tray{tray_num}_Pellet{pellet_idx+1:02d}"
            if col_name in ODC_COL_IDX:
                row[ODC_COL_IDX[col_name]] = score
        
        # Compute tray statistics
        stats = compute_tray_stats(scores)
        tray_stats[tray_num] = stats
        
        # Write tray stats
        row[ODC_COL_IDX[f'Tray{tray_num}_Presented']] = stats['presented']
        row[ODC_COL_IDX[f'Tray{tray_num}_Missed']] = stats['missed']
        row[ODC_COL_IDX[f'Tray{tray_num}_Displaced']] = stats['displaced']
        row[ODC_COL_IDX[f'Tray{tray_num}_Retrieved']] = stats['retrieved']
        row[ODC_COL_IDX[f'Tray{tray_num}_Contacted']] = stats['contacted']
        row[ODC_COL_IDX[f'Tray{tray_num}_Miss_Pct']] = stats['miss_pct']
        row[ODC_COL_IDX[f'Tray{tray_num}_Displaced_Pct']] = stats['displaced_pct']
        row[ODC_COL_IDX[f'Tray{tray_num}_Retrieved_Pct']] = stats['retrieved_pct']
        row[ODC_COL_IDX[f'Tray{tray_num}_Contacted_Pct']] = stats['contacted_pct']
    
    # === Section 8: Daily Totals ===
    if tray_stats:
        total_presented = sum(s['presented'] for s in tray_stats.values())
        total_missed = sum(s['missed'] for s in tray_stats.values())
        total_displaced = sum(s['displaced'] for s in tray_stats.values())
        total_retrieved = sum(s['retrieved'] for s in tray_stats.values())
        total_contacted = sum(s['contacted'] for s in tray_stats.values())
        
        row[ODC_COL_IDX['Total_Presented']] = total_presented
        row[ODC_COL_IDX['Total_Missed']] = total_missed
        row[ODC_COL_IDX['Total_Displaced']] = total_displaced
        row[ODC_COL_IDX['Total_Retrieved']] = total_retrieved
        row[ODC_COL_IDX['Total_Contacted']] = total_contacted
        
        if total_presented > 0:
            row[ODC_COL_IDX['Total_Miss_Pct']] = round(total_missed / total_presented * 100, 2)
            row[ODC_COL_IDX['Total_Displaced_Pct']] = round(total_displaced / total_presented * 100, 2)
            row[ODC_COL_IDX['Total_Retrieved_Pct']] = round(total_retrieved / total_presented * 100, 2)
            row[ODC_COL_IDX['Total_Contacted_Pct']] = round(total_contacted / total_presented * 100, 2)
        
        # Averages across trays
        pcts = {'miss': [], 'displaced': [], 'retrieved': [], 'contacted': []}
        for s in tray_stats.values():
            if s['miss_pct'] is not None:
                pcts['miss'].append(s['miss_pct'])
            if s['displaced_pct'] is not None:
                pcts['displaced'].append(s['displaced_pct'])
            if s['retrieved_pct'] is not None:
                pcts['retrieved'].append(s['retrieved_pct'])
            if s['contacted_pct'] is not None:
                pcts['contacted'].append(s['contacted_pct'])
        
        if pcts['miss']:
            row[ODC_COL_IDX['Avg_Miss_Pct']] = round(sum(pcts['miss']) / len(pcts['miss']), 2)
        if pcts['displaced']:
            row[ODC_COL_IDX['Avg_Displaced_Pct']] = round(sum(pcts['displaced']) / len(pcts['displaced']), 2)
        if pcts['retrieved']:
            row[ODC_COL_IDX['Avg_Retrieved_Pct']] = round(sum(pcts['retrieved']) / len(pcts['retrieved']), 2)
            row[ODC_COL_IDX['Max_Retrieved_Pct']] = max(pcts['retrieved'])
            row[ODC_COL_IDX['Min_Retrieved_Pct']] = min(pcts['retrieved'])
        if pcts['contacted']:
            row[ODC_COL_IDX['Avg_Contacted_Pct']] = round(sum(pcts['contacted']) / len(pcts['contacted']), 2)
            row[ODC_COL_IDX['Max_Contacted_Pct']] = max(pcts['contacted'])
            row[ODC_COL_IDX['Min_Contacted_Pct']] = min(pcts['contacted'])
    
    # === Source Tracking ===
    row[ODC_COL_IDX['Source_File']] = source_file
    row[ODC_COL_IDX['Source_Sheet']] = '3b_Manual_Tray'
    
    return row


# =============================================================================
# MAIN PROCESSING
# =============================================================================

def process_file(input_path, output_dir=None):
    """Process a single cohort file and populate ODC sheet."""
    
    input_path = Path(input_path)
    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        return None
    
    if output_dir is None:
        output_dir = OUTPUT_SUBDIR
    else:
        output_dir = Path(output_dir)
    
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{input_path.stem}_updated.xlsx"
    
    print(f"\n{'='*60}")
    print(f"Processing: {input_path.name}")
    print(f"{'='*60}")
    
    # Load workbook
    print("Loading workbook...")
    wb = openpyxl.load_workbook(input_path)
    
    # Also load with data_only for formula values
    print("Loading formula values...")
    try:
        wb_data = openpyxl.load_workbook(input_path, data_only=True)
    except:
        wb_data = wb
    
    # Extract data from source sheets
    print("Extracting metadata...")
    metadata = extract_metadata(wb_data)
    print(f"  Found {len(metadata)} animals in metadata")
    
    print("Extracting baseline weights...")
    baseline_weights = extract_baseline_weights(wb_data)
    print(f"  Found {len(baseline_weights)} baseline weights")
    
    print("Extracting contusion data...")
    contusion = extract_contusion_data(wb_data)
    print(f"  Found {len(contusion)} contusion records")
    
    print("Extracting injection data...")
    injection = extract_injection_data(wb_data)
    print(f"  Found {len(injection)} injection records")
    
    print("Extracting tray data...")
    grouped_trays = extract_tray_data(wb_data)
    print(f"  Found {len(grouped_trays)} (animal, date) combinations")
    
    if not grouped_trays:
        print("Error: No tray data found!")
        return None
    
    # Build ODC rows
    print("\nBuilding ODC rows...")
    odc_rows = []
    source_file = input_path.stem
    
    for (animal_id, date), day_data in sorted(grouped_trays.items()):
        row = build_odc_row(
            animal_id, date, day_data,
            metadata, baseline_weights, contusion, injection,
            source_file
        )
        odc_rows.append(row)
    
    print(f"  Built {len(odc_rows)} ODC rows")
    
    # Create or replace ODC sheet
    print("\nWriting ODC sheet...")
    
    if '2_ODC_Animal_Tracking' in wb.sheetnames:
        del wb['2_ODC_Animal_Tracking']
    
    ws = wb.create_sheet('2_ODC_Animal_Tracking')
    
    # Write headers
    for col_idx, header in enumerate(ODC_COLUMNS, 1):
        ws.cell(1, col_idx, value=header)
    
    # Write data rows
    for row_idx, row_data in enumerate(odc_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value=value)
    
    # Auto-scale first 50 columns
    for col in range(1, min(51, len(ODC_COLUMNS) + 1)):
        header = ws.cell(1, col).value
        if header:
            width = min(20, max(8, len(str(header)) * 1.1))
            ws.column_dimensions[get_column_letter(col)].width = width
    
    # Move ODC sheet to correct position (index 3, after planning sheet)
    sheets = wb.sheetnames
    if '2_ODC_Animal_Tracking' in sheets:
        current_idx = sheets.index('2_ODC_Animal_Tracking')
        target_idx = 3  # After 0a, 0_, 1_
        if current_idx != target_idx:
            wb.move_sheet('2_ODC_Animal_Tracking', offset=target_idx - current_idx)
    
    # Save
    print(f"\nSaving to: {output_path}")
    wb.save(output_path)
    
    print(f"\n{'='*60}")
    print(f"[OK] Successfully processed {input_path.name}")
    print(f"     ODC sheet: {len(odc_rows)} rows x {len(ODC_COLUMNS)} columns")
    print(f"     Output: {output_path}")
    print(f"{'='*60}")
    
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Populate 2_ODC_Animal_Tracking from source data sheets."
    )
    parser.add_argument("file", nargs="?", help="Cohort file to process")
    parser.add_argument("--all", action="store_true", help="Process all cohort files in directory")
    parser.add_argument("--output-dir", type=str, help="Output directory")
    
    args = parser.parse_args()
    
    if args.all:
        # Process all Connectome files
        files = list(SCRIPT_DIR.glob("Connectome_*_Animal_Tracking*.xlsx"))
        files = [f for f in files if not f.name.startswith('~')]
        
        if not files:
            print("No Connectome files found.")
            return
        
        print(f"Found {len(files)} files to process:")
        for f in files:
            print(f"  - {f.name}")
        
        for f in files:
            try:
                process_file(f, args.output_dir)
            except Exception as e:
                print(f"Error processing {f.name}: {e}")
    
    elif args.file:
        process_file(args.file, args.output_dir)
    
    else:
        # Interactive mode
        print("\n" + "=" * 60)
        print("  Script 1: Update ODC Sheet")
        print("=" * 60)
        
        files = list(SCRIPT_DIR.glob("*.xlsx"))
        files = [f for f in files if not f.name.startswith('~')]
        
        if not files:
            print("\nNo Excel files found in script directory.")
            return
        
        print("\nExcel files:")
        for i, f in enumerate(files, 1):
            print(f"  {i}. {f.name}")
        
        print(f"\n  a. Process all Connectome files")
        
        choice = input("\nSelect file number (or 'a' for all): ").strip()
        
        if choice.lower() == 'a':
            cohort_files = [f for f in files if 'Connectome' in f.name]
            for f in cohort_files:
                try:
                    process_file(f)
                except Exception as e:
                    print(f"Error: {e}")
        else:
            try:
                idx = int(choice) - 1
                if 0 <= idx < len(files):
                    process_file(files[idx])
            except ValueError:
                print("Invalid selection")
        
        input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
