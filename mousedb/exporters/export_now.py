"""Minimal export - no external dependencies beyond pandas/openpyxl."""
import pandas as pd
from pathlib import Path
import openpyxl
import re


def main():
    """Main export function."""
    base_dir = Path.cwd()

    # Load existing data
    print("Loading data...")
    pellet_df = pd.read_csv(base_dir / 'generated' / 'all_cohorts_pellet_level.csv', low_memory=False)
    tray_df = pd.read_csv(base_dir / 'generated' / 'all_cohorts_tray_summaries.csv')
    print(f"  Pellets: {len(pellet_df):,}")
    print(f"  Trays: {len(tray_df):,}")

    # Load surgery data directly
    print("\nLoading surgery metadata...")
    surgery_rows = []

    for fpath in sorted(base_dir.glob('Connectome_*_Animal_Tracking*.xlsx')):
        if fpath.name.startswith('~') or '_fixed' in fpath.name or '(' in fpath.name:
            continue

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)

            # Extract cohort
            match = re.search(r'Connectome_(\d+)_', fpath.name)
            cohort = f'CNT_{int(match.group(1)):02d}' if match else None

            if '4_Contusion_Injury_Details' in wb.sheetnames:
                ws = wb['4_Contusion_Injury_Details']
                headers = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column+1) if ws.cell(1, c).value}

                id_col = headers.get('Subject_ID') or headers.get('Animal')
                if id_col:
                    for row in range(2, ws.max_row + 1):
                        animal = ws.cell(row, id_col).value
                        if not animal:
                            continue

                        surgery_rows.append({
                            'mouse_id': str(animal).replace('_', ''),
                            'cohort': cohort,
                            'surgery_date': ws.cell(row, headers.get('Surgery_Date', 999)).value if 'Surgery_Date' in headers else None,
                            'surgery_type': ws.cell(row, headers.get('Surgery_Type', 999)).value if 'Surgery_Type' in headers else None,
                            'surgery_severity': ws.cell(row, headers.get('Surgery_Severity', 999)).value if 'Surgery_Severity' in headers else None,
                            'injury_location': ws.cell(row, headers.get('Contusion_Location', 999)).value if 'Contusion_Location' in headers else None,
                            'actual_kd': ws.cell(row, headers.get('Actual_kd', 999)).value if 'Actual_kd' in headers else None,
                            'actual_displacement': ws.cell(row, headers.get('Actual_displacement', 999)).value if 'Actual_displacement' in headers else None,
                            'survived': ws.cell(row, headers.get('Survived', 999)).value if 'Survived' in headers else None,
                        })
            wb.close()
            print(f"  {fpath.name}")
        except Exception as e:
            print(f"  SKIP {fpath.name}: {e}")

    surgery_df = pd.DataFrame(surgery_rows)
    print(f"  Total: {len(surgery_df)} mice with surgery data")

    # Merge surgery into tray data
    tray_df['_norm'] = tray_df['Animal'].str.replace('_', '').str.upper()
    surgery_df['_norm'] = surgery_df['mouse_id'].str.upper()

    tray_merged = tray_df.merge(
        surgery_df.drop(columns=['mouse_id', 'cohort']),
        on='_norm',
        how='left'
    ).drop(columns=['_norm'])

    # Add days post injury
    tray_merged['Date'] = pd.to_datetime(tray_merged['Date'])
    tray_merged['surgery_date'] = pd.to_datetime(tray_merged['surgery_date'], errors='coerce')
    tray_merged['days_post_injury'] = (tray_merged['Date'] - tray_merged['surgery_date']).dt.days

    # Add timepoint category
    def get_timepoint(phase):
        if pd.isna(phase): return None
        p = str(phase)
        if 'Training' in p: return 'Training'
        if 'Pre-Injury' in p: return 'Pre-Injury'
        if 'Post-Injury' in p: return 'Post-Injury'
        if 'Rehab_Easy' in p: return 'Rehab_Easy'
        if 'Rehab_Flat' in p: return 'Rehab_Flat'
        if 'Rehab_Pillar' in p: return 'Rehab_Pillar'
        return p

    tray_merged['Timepoint'] = tray_merged['Test_Phase'].apply(get_timepoint)

    # Export
    out = base_dir / 'generated' / 'mousedb_export.xlsx'
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        tray_merged.to_excel(w, 'Tray_with_Surgery', index=False)
        surgery_df.to_excel(w, 'Surgery_Metadata', index=False)

    print(f"\n{'='*50}")
    print(f"EXPORTED: {out}")
    print(f"{'='*50}")
    print(f"  {len(tray_merged):,} tray records")
    print(f"  {tray_merged['Animal'].nunique()} animals")
    print(f"  Cohorts: {sorted(tray_merged['Cohort'].unique())}")
    print(f"  Timepoints: {sorted(tray_merged['Timepoint'].dropna().unique())}")
    print(f"  Days post injury available: {tray_merged['days_post_injury'].notna().sum()}")


if __name__ == "__main__":
    main()
